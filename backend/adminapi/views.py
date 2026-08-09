import csv
from datetime import date, timedelta
from decimal import Decimal

from django.db.models import Count, Prefetch, Q, Sum
from django.db import transaction
from django.db.models.functions import TruncDate, TruncMonth, TruncYear
from django.utils import timezone
from django.core.cache import cache
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from django.http import HttpResponse

from accounts.models import AuditLog, CustomUser, FinancialTransaction, LedgerEntry, Wallet, WalletAccount, WalletTransaction, WalletUploadRequest, ConsumerVoucher, UserKYC, WithdrawalRequest, SupportTicket, SupportTicketMessage, AgencyRegionAssignment
from coupons.models import Coupon, CouponCode, CouponSubmission, CouponBatch
from uploads.models import FileUpload, DashboardCard, HomeCard, LuckyDrawSubmission
from market.models import Product, PurchaseRequest, Banner, BannerItem, BannerPurchaseRequest
from business.models import UserMatrixProgress, AutoPoolAccount, DailyReport, CommissionConfig, PromoPurchase
from mlm_ranks.models import RankUpgrade
from .permissions import IsAdminOrStaff, IsSuperAdmin, HasAdminModuleAccess, HasAnyPermission, has_admin_module_access, MODULE_KEYS
from .serializers import AdminUserNodeSerializer, AdminKYCSerializer, AdminWithdrawalSerializer, AdminMatrixProgressSerializer, AdminSupportTicketSerializer, AdminSupportTicketMessageSerializer, AdminUserEditSerializer, AdminAutopoolTxnSerializer, AdminAutopoolConfigSerializer
from .dynamic import field_meta_from_serializer


class AdminMetricsView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_basic")]

    def get(self, request):
        today = timezone.now().date()

        # Return cached metrics unless explicitly bypassed with ?refresh=1
        try:
            refresh = str(request.query_params.get("refresh") or "").lower()
        except Exception:
            refresh = ""
        cache_key = "admin_metrics_v3"
        if refresh not in ("1", "true", "yes"):
            cached = cache.get(cache_key)
            if cached is not None:
                return Response(cached, status=status.HTTP_200_OK)

        # Ensure CommissionConfig exists (seed if missing)
        try:
            CommissionConfig.get_solo()
        except Exception:
            pass

        # Users (condensed into a single aggregate where possible)
        users_agg = CustomUser.objects.aggregate(
            total=Count("id"),
            active=Count("id", filter=Q(account_active=True)),
            blocked=Count("id", filter=Q(is_active=False)),
            todayNew=Count("id", filter=Q(date_joined__date=today)),
            consumers_without_kyc=Count("id", filter=Q(category="consumer") & Q(kyc__isnull=True)),
        )

        # KYC pending: users with KYC not verified + users without KYC (consumers)
        kyc_unverified = UserKYC.objects.filter(verified=False).count()
        kyc_pending = (kyc_unverified or 0) + (users_agg.get("consumers_without_kyc") or 0)

        users_block = {
            "total": users_agg.get("total") or 0,
            "active": users_agg.get("active") or 0,
            "inactive": (users_agg.get("total") or 0) - (users_agg.get("active") or 0),
            "blocked": users_agg.get("blocked") or 0,
            "todayNew": users_agg.get("todayNew") or 0,
            "kycPending": int(kyc_pending),
        }

        # KYC aggregate block for dashboard (submitted, pending, approved)
        # "submitted" counts only KYC where user has actually provided details (bank_name / account / IFSC not empty)
        qs_submitted = UserKYC.objects.filter(
            (Q(bank_name__isnull=False) & ~Q(bank_name=""))
            | (Q(bank_account_number__isnull=False) & ~Q(bank_account_number=""))
            | (Q(ifsc_code__isnull=False) & ~Q(ifsc_code=""))
        )
        submitted_count = qs_submitted.count()
        approved_count = qs_submitted.filter(verified=True).count()
        kyc_block = {
            "submitted": int(submitted_count),
            "approved": int(approved_count),
            # Pending among those who actually submitted (submitted - approved)
            "pending": int(max(submitted_count - approved_count, 0)),
            # Unverified KYC rows (matches AdminKYC list with status=pending)
            "unverified": int(kyc_unverified or 0),
            # Consumers without any KYC row
            "missing": int(users_agg.get("consumers_without_kyc") or 0),
            # Combined pending signal for high-level use
            "pending_all": int((kyc_unverified or 0) + (users_agg.get("consumers_without_kyc") or 0)),
        }

        # Wallets (aggregate total balance and count together)
        from accounts.models import WalletAccount
        from accounts.finance_constants import WalletTypes
        total_balance = WalletAccount.objects.filter(
            wallet_type__in=[WalletTypes.MAIN, WalletTypes.SELF_PACKAGE_POCKET]
        ).aggregate(s=Sum("current_balance"))["s"] or Decimal("0.00")
        total_wallets_count = Wallet.objects.count()
        wallets_block = {
            "totalBalance": float(total_balance),
            "transactionsToday": WalletTransaction.objects.filter(created_at__date=today).count(),
            "count": total_wallets_count,
        }

        # Withdrawals (single aggregate for count and amount)
        wagg2 = WithdrawalRequest.objects.filter(status="pending").aggregate(c=Count("id"), s=Sum("amount"))
        withdrawals_block = {
            "pendingCount": wagg2.get("c") or 0,
            "pendingAmount": float(wagg2.get("s") or Decimal("0.00")),
        }

        # Coupons (single aggregate for multiple counts)
        cagg = CouponCode.objects.aggregate(
            total=Count("id"),
            assigned=Count("id", filter=Q(status__in=["ASSIGNED_AGENCY", "ASSIGNED_EMPLOYEE"])),
            redeemed=Count("id", filter=Q(status="REDEEMED")),
        )
        # Pending submissions considered as waiting for approvals (SUBMITTED or EMPLOYEE_APPROVED)
        pending_submissions = CouponSubmission.objects.filter(status__in=["SUBMITTED", "EMPLOYEE_APPROVED"]).count()
        coupons_block = {
            "total": cagg.get("total") or 0,
            "assigned": cagg.get("assigned") or 0,
            "redeemed": cagg.get("redeemed") or 0,
            "pendingSubmissions": pending_submissions,
        }

        # Uploads (single aggregate for total and today's new)
        uagg = FileUpload.objects.aggregate(
            total=Count("id"),
            todayNew=Count("id", filter=Q(created_at__date=today)),
        )
        uploads_block = {
            "total": uagg.get("total") or 0,
            "todayNew": uagg.get("todayNew") or 0,
            "failed": 0,
        }

        # Uploads models
        uploads_models_block = {
            "dashboardCards": DashboardCard.objects.count(),
            "homeCards": HomeCard.objects.count(),
            "luckyDrawSubmissions": LuckyDrawSubmission.objects.count(),
            "luckyDrawPendingTRE": LuckyDrawSubmission.objects.filter(status="SUBMITTED").count(),
            "luckyDrawPendingAgency": LuckyDrawSubmission.objects.filter(status="TRE_APPROVED").count(),
        }

        # Market
        market_block = {
            "products": Product.objects.count(),
            "purchaseRequests": PurchaseRequest.objects.count(),
            "purchaseRequestsPending": PurchaseRequest.objects.filter(status=PurchaseRequest.STATUS_PENDING).count(),
            "banners": Banner.objects.count(),
            "bannerItems": BannerItem.objects.count(),
            "bannerPurchaseRequests": BannerPurchaseRequest.objects.count(),
            "bannerPurchaseRequestsPending": BannerPurchaseRequest.objects.filter(status=BannerPurchaseRequest.STATUS_PENDING).count(),
        }

        # Reports aggregate (today and total in one query)
        ragg = DailyReport.objects.aggregate(
            today=Count("id", filter=Q(date=today)),
            total=Count("id"),
        )

        def wallet_credit_stats(qs):
            today_filter = Q(created_at__date=today)
            agg = qs.aggregate(
                today_count=Count("user", filter=today_filter, distinct=True),
                today_amount=Sum("amount", filter=today_filter),
                total_count=Count("user", distinct=True),
                total_amount=Sum("amount"),
            )
            return {
                "todayCount": int(agg.get("today_count") or 0),
                "todayAmount": float(agg.get("today_amount") or Decimal("0.00")),
                "totalCount": int(agg.get("total_count") or 0),
                "totalAmount": float(agg.get("total_amount") or Decimal("0.00")),
            }

        upload_sources = ["WALLET_UPLOAD", "UPLOAD_TO_WALLET", "PACKAGE_UPLOAD", "PACKAGE_BUY_UPLOAD"]
        wallet_pocket_stats = {
            # User-approved add-money uploads. This is money added into Add Money pocket, not money spent.
            "addMoney": wallet_credit_stats(WalletTransaction.objects.filter(
                source_type__in=upload_sources,
                amount__gt=0,
            )),
            # Normal internal/self package pocket credits, excluding add-money upload credits.
            "selfPackagePocket": wallet_credit_stats(WalletTransaction.objects.filter(
                type="INTERNAL_WALLET_CREDIT",
                amount__gt=0,
            ).exclude(source_type__in=upload_sources)),
            # Package purchase coupon wallet credits.
            "packageCouponPocket": wallet_credit_stats(WalletTransaction.objects.filter(
                type__in=["PACKAGE_COUPON_WALLET_CREDIT", "VOUCHER_REDEEM_CREDIT"],
                amount__gt=0,
            )),
            # General coupon pocket credits.
            "couponPocket": wallet_credit_stats(WalletTransaction.objects.filter(
                type__in=["COUPON_WALLET_CREDIT", "COUPON_WALLET_REFUND"],
                amount__gt=0,
            )),
            "shoppingWallet": wallet_credit_stats(WalletTransaction.objects.filter(
                type="SHOPPING_WALLET_CREDIT",
                amount__gt=0,
            )),
            "withdrawalWallet": wallet_credit_stats(WalletTransaction.objects.filter(
                type="WITHDRAWAL_WALLET_CREDIT",
                amount__gt=0,
            )),
        }

        def promo_package_stats_for_queryset(base):
            today_filter = Q(approved_at__date=today)

            def summarize(qs):
                agg = qs.aggregate(
                    today_count=Count("user", filter=today_filter, distinct=True),
                    today_amount=Sum("amount_paid", filter=today_filter),
                    total_count=Count("user", distinct=True),
                    total_amount=Sum("amount_paid"),
                )
                return {
                    "todayCount": int(agg.get("today_count") or 0),
                    "todayAmount": float(agg.get("today_amount") or Decimal("0.00")),
                    "totalCount": int(agg.get("total_count") or 0),
                    "totalAmount": float(agg.get("total_amount") or Decimal("0.00")),
                }

            return {
                "overall": summarize(base),
                # Internal self-package pocket: wallet-paid purchase using normal internal package balance.
                "selfPackagePocket": summarize(base.filter(
                    payment_mode="WALLET",
                    wallet_debit_tx__type="INTERNAL_WALLET_DEBIT",
                    wallet_debit_tx__source_type="PROMO_PURCHASE",
                )),
                # Add Money pocket: wallet-paid purchase using funds uploaded/approved into Add Money.
                "addMoney": summarize(base.filter(
                    payment_mode="WALLET",
                    wallet_debit_tx__type="INTERNAL_WALLET_DEBIT",
                    wallet_debit_tx__source_type="WALLET_UPLOAD",
                )),
                # Package coupon pocket: wallet-paid purchase using package-purchase coupon balance.
                "couponPocket": summarize(base.filter(
                    payment_mode="WALLET",
                    wallet_debit_tx__type="PACKAGE_COUPON_WALLET_DEBIT",
                )),
            }

        def promo_package_stats(price):
            price = Decimal(str(price))
            lo = price - Decimal("0.50")
            hi = price + Decimal("0.50")
            return promo_package_stats_for_queryset(PromoPurchase.objects.filter(
                status="APPROVED",
                package__price__gte=lo,
                package__price__lte=hi,
            ))

        smart_product_qs = PromoPurchase.objects.filter(status="APPROVED").filter(
            Q(package__type="MONTHLY") |
            Q(package__price__gte=Decimal("999.50"), package__price__lte=Decimal("1000.50"))
        )

        def rank_upgrade_stats(qs):
            today_filter = Q(created_at__date=today)
            agg = qs.aggregate(
                today_count=Count("user", filter=today_filter, distinct=True),
                today_amount=Sum("upgrade_amount", filter=today_filter),
                total_count=Count("user", distinct=True),
                total_amount=Sum("upgrade_amount"),
            )
            return {
                "todayCount": int(agg.get("today_count") or 0),
                "todayAmount": float(agg.get("today_amount") or Decimal("0.00")),
                "totalCount": int(agg.get("total_count") or 0),
                "totalAmount": float(agg.get("total_amount") or Decimal("0.00")),
            }

        digital_education_prime_stats = {
            "pendingApproval": rank_upgrade_stats(
                RankUpgrade.objects.filter(payment_status=RankUpgrade.STATUS_INITIATED)
            ),
            "approved": rank_upgrade_stats(
                RankUpgrade.objects.filter(payment_status=RankUpgrade.STATUS_SUCCESS)
            ),
            "overall": rank_upgrade_stats(RankUpgrade.objects.all()),
        }

        # Autopool aggregates (single DB hit)
        acc_by_status = list(
            AutoPoolAccount.objects.values("status")
            .annotate(c=Count("id"))
            .order_by()
        )

        payload = {
            "users": users_block,
            "wallets": wallets_block,
            "withdrawals": withdrawals_block,
            "kyc": kyc_block,
            "coupons": coupons_block,
            "uploads": uploads_block,
            "uploadsModels": uploads_models_block,
            "market": market_block,
            "autopool": {
                "total": sum(row["c"] for row in acc_by_status),
                "byStatus": {row["status"]: row["c"] for row in acc_by_status},
            },
            "reports": {
                "dailyReportsToday": ragg.get("today") or 0,
                "dailyReportsTotal": ragg.get("total") or 0,
            },
            "packageStats": {
                "subscription750": promo_package_stats(750),
                "smartProduct1000": promo_package_stats_for_queryset(smart_product_qs),
                "digitalEducationPrime": digital_education_prime_stats,
            },
            "walletPocketStats": wallet_pocket_stats,
            "commission": {
                "configs": CommissionConfig.objects.count(),
            },
        }
        # Cache for a short duration to avoid heavy repeated aggregation under load
        try:
            cache.set(cache_key, payload, timeout=20)  # seconds
        except Exception:
            pass
        return Response(payload, status=status.HTTP_200_OK)


class AdminUserTreeRoot(APIView):
    """
    Resolve a root user for the hierarchy tree.
    identifier can be: id (int), username, email, or unique_id.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("users")]

    def get(self, request):
        identifier = (request.query_params.get("identifier") or "").strip()
        if not identifier:
            return Response({"detail": "identifier is required"}, status=400)

        user = None
        # Normalize digits for id/phone queries
        digits = "".join([c for c in identifier if c.isdigit()])

        # Prefer exact prefixed_id (sponsor code)
        user = CustomUser.objects.filter(prefixed_id__iexact=identifier).first() or user

        # Try by id (numeric)
        if not user and digits and digits == identifier and digits.isdigit():
            user = CustomUser.objects.filter(id=int(digits)).first()

        # Try by username/email/unique_id and phone digits (avoid matching children by sponsor_id here)
        if not user:
            q = (
                Q(username__iexact=identifier)
                | Q(email__iexact=identifier)
                | Q(unique_id__iexact=identifier)
            )
            if digits:
                q = q | Q(phone__iexact=digits) | Q(username__iexact=digits)
            user = CustomUser.objects.filter(q).first()

        if not user:
            return Response({"detail": "User not found"}, status=404)

        # Annotate direct_count and has_children
        node = (
            CustomUser.objects.filter(id=user.id)
            .annotate(
                direct_count=Count("registrations", distinct=True),
            )
            .first()
        )
        has_children = (getattr(node, "direct_count", 0) or 0) > 0

        data = AdminUserNodeSerializer({
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role,
            "category": user.category,
            "phone": user.phone,
            "state": user.state,
            "pincode": user.pincode,
            "direct_count": getattr(node, "direct_count", 0) or 0,
            "has_children": has_children,
        }).data
        return Response(data, status=200)


class AdminUserTreeDefaultRoot(APIView):
    """
    Return default root user for hierarchy tree (first superuser by id; fallback to first staff; else earliest user).
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("users")]

    def get(self, request):
        user = None
        try:
            from business.models import RootConsumerConfig
            cfg = RootConsumerConfig.get_solo()
            user = cfg.get_root_user()
        except Exception:
            user = None

        if not user:
            user = (
                CustomUser.objects.filter(is_superuser=True).order_by("id").first()
                or CustomUser.objects.filter(is_staff=True).order_by("id").first()
                or CustomUser.objects.order_by("id").first()
            )

        if not user:
            return Response({"detail": "No users found"}, status=404)

        node = (
            CustomUser.objects.filter(id=user.id)
            .annotate(direct_count=Count("registrations", distinct=True))
            .first()
        )
        has_children = (getattr(node, "direct_count", 0) or 0) > 0

        data = AdminUserNodeSerializer({
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role,
            "category": user.category,
            "phone": user.phone,
            "state": user.state,
            "pincode": user.pincode,
            "direct_count": getattr(node, "direct_count", 0) or 0,
            "has_children": has_children,
        }).data
        return Response(data, status=200)


class AdminUserTreeChildren(APIView):
    """
    Return direct children for given userId (registered_by relationship).
    Supports pagination with page and page_size.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("users")]

    def get(self, request):
        try:
            user_id = int(request.query_params.get("userId") or "0")
        except ValueError:
            return Response({"detail": "Invalid or missing userId"}, status=400)

        page = max(int(request.query_params.get("page") or 1), 1)
        page_size = min(max(int(request.query_params.get("page_size") or 20), 1), 100)

        qs = (
            CustomUser.objects.filter(registered_by_id=user_id)
            .annotate(direct_count=Count("registrations", distinct=True))
            .order_by("-date_joined")
        )

        # Fast count path to avoid heavy COUNT(*) over annotated/prefetched queryset
        cat_param = (request.query_params.get("category") or "").strip()
        if cat_param:
            c_key = cat_param.lower().replace(" ", "_").replace("-", "_")
            try:
                if "cordinator" in c_key:
                    c_key = c_key.replace("cordinator", "coordinator")
                if "subfranchise" in c_key and "sub_franchise" not in c_key:
                    c_key = c_key.replace("subfranchise", "sub_franchise")
            except Exception:
                pass
            try:
                cat_values = {str(k).lower(): k for k, _ in getattr(CustomUser, "CATEGORY_CHOICES", [])}
                cat_labels = {str(v).lower().replace(" ", "_").replace("-", "_"): k for k, v in getattr(CustomUser, "CATEGORY_CHOICES", [])}
                c_norm = cat_values.get(c_key) or cat_labels.get(c_key) or c_key
            except Exception:
                c_norm = c_key
            total = CustomUser.objects.filter(category__iexact=c_norm).count()
        else:
            total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        items = []
        for u in qs[start:end]:
            items.append({
                "id": u.id,
                "username": u.username,
                "full_name": u.full_name,
                "role": u.role,
                "category": u.category,
                "phone": u.phone,
                "state": u.state,
                "pincode": u.pincode,
                "direct_count": getattr(u, "direct_count", 0) or 0,
                "has_children": (getattr(u, "direct_count", 0) or 0) > 0,
            })

        data = {
            "count": total,
            "page": page,
            "page_size": page_size,
            "results": AdminUserNodeSerializer(items, many=True).data,
        }
        return Response(data, status=200)


class AdminUsersList(ListAPIView):
    """
    Admin users list with powerful filters: role, phone, category, pincode, state, kyc.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("users")]
    serializer_class = AdminUserNodeSerializer

    def _consumer_columns_preset(self):
        return str(self.request.query_params.get("consumer_columns") or "").strip().lower()

    def _needs_package_summary(self):
        return self._consumer_columns_preset() in ("package", "all")

    def _needs_wallet_summary(self):
        return self._consumer_columns_preset() in ("wallet", "coupons", "all")

    def get_queryset(self):
        select_related = ["country", "state", "city", "wallet", "kyc", "registered_by", "merchant_profile"]
        only_fields = [
            # Base fields used by AdminUserNodeSerializer and filters
            "id", "username", "full_name", "email", "role", "category",
            "phone", "pincode", "date_joined", "is_active", "account_active",
            "prefixed_id", "sponsor_id", "unique_id", "first_purchase_activated_at",
            "avatar",
            # Geo relations displayed in serializer
            "country__id", "country__name",
            "state__id", "state__name",
            "city__id", "city__name",
            # Registered by (for sponsor display)
            "registered_by__username", "registered_by__prefixed_id", "registered_by__full_name",
            # Wallet relation placeholder for select_related traversal
            "wallet__id",
            # KYC status in list grid
            "kyc__verified", "kyc__verified_at",
            # Merchant profile fields used in Admin grids
            "merchant_profile",
            "merchant_profile__business_name",
            "merchant_profile__business_category",
            "merchant_profile__address",
            "merchant_profile__commission_percent",
            "merchant_profile__service_mode",
            # Password metadata (status/algo only)
            "password", "last_password_encrypted",
        ]
        select_related.append("reward_points_account")
        only_fields.append("reward_points_account__balance_points")
        prefetches = [
            Prefetch("wallet_accounts", queryset=WalletAccount.objects.only("user_id", "wallet_type", "current_balance"), to_attr="prefetched_wallet_accounts"),
            Prefetch("promo_purchases", queryset=PromoPurchase.objects.select_related("package").filter(status="APPROVED").only("id", "user_id", "package_id", "status", "approved_at", "requested_at", "tri_app_slug", "package__id", "package__code", "package__name", "package__type", "package__price").order_by("-approved_at", "-id"), to_attr="approved_promo_purchases"),
            Prefetch(
                "region_assignments",
                queryset=AgencyRegionAssignment.objects.select_related("state").order_by("level", "state__name", "district", "pincode"),
                to_attr="prefetched_agency_assignments",
            ),
            Prefetch("matrix_progress", queryset=UserMatrixProgress.objects.only("id", "user_id", "level_reached"), to_attr="matrix_progress_list"),
        ]

        qs = (
            CustomUser.objects
            .select_related(*select_related)
        )
        if prefetches:
            qs = qs.prefetch_related(*prefetches)
        role = (self.request.query_params.get("role") or "").strip()
        phone = (self.request.query_params.get("phone") or "").strip()
        category = (self.request.query_params.get("category") or "").strip()
        pincode = (self.request.query_params.get("pincode") or "").strip()
        state_id = (self.request.query_params.get("state") or "").strip()
        kyc = (self.request.query_params.get("kyc") or "").strip()
        search = (self.request.query_params.get("search") or "").strip()
        activated = (self.request.query_params.get("activated") or "").strip().lower()
        is_active = (self.request.query_params.get("is_active") or "").strip().lower()

        # Normalize role/category to be case-insensitive and accept human labels/tokens
        if role:
            r = str(role).strip()
            r_key = r.lower()
            try:
                role_keys = {str(k).lower(): k for k, _ in getattr(CustomUser, "ROLE_CHOICES", [])}
                role_labels = {str(v).lower().replace(" ", "_").replace("-", "_"): k for k, v in getattr(CustomUser, "ROLE_CHOICES", [])}
                r_norm = role_keys.get(r_key) or role_labels.get(r_key.replace(" ", "_").replace("-", "_")) or r
            except Exception:
                r_norm = r
            qs = qs.filter(role__iexact=r_norm)
        if category:
            c = str(category).strip()
            c_key = c.lower().replace(" ", "_").replace("-", "_")
            # Fix common misspellings/synonyms to make URL query robust
            try:
                # Accept '*_cordinator' -> '*_coordinator'
                if "cordinator" in c_key:
                    c_key = c_key.replace("cordinator", "coordinator")
                # Accept 'subfranchise' -> 'sub_franchise'
                if "subfranchise" in c_key and "sub_franchise" not in c_key:
                    c_key = c_key.replace("subfranchise", "sub_franchise")
            except Exception:
                pass
            try:
                cat_values = {str(k).lower(): k for k, _ in getattr(CustomUser, "CATEGORY_CHOICES", [])}
                cat_labels = {str(v).lower().replace(" ", "_").replace("-", "_"): k for k, v in getattr(CustomUser, "CATEGORY_CHOICES", [])}
                c_norm = cat_values.get(c_key) or cat_labels.get(c_key) or c_key
            except Exception:
                c_norm = c_key
            qs = qs.filter(category__iexact=c_norm)
        if phone:
            qs = qs.filter(phone__icontains=phone)
        if pincode:
            qs = qs.filter(pincode__icontains=pincode)
        if state_id and state_id.isdigit():
            qs = qs.filter(state_id=int(state_id))
        if kyc:
            if kyc == "pending":
                qs = qs.filter(Q(kyc__verified=False) | Q(kyc__isnull=True))
            elif kyc == "verified":
                qs = qs.filter(kyc__verified=True)

        # New: filter by explicit account_active (admin "Account status")
        account_active = (self.request.query_params.get("account_active") or "").strip().lower()
        if account_active in ("1", "true", "yes", "active"):
            qs = qs.filter(account_active=True)
        elif account_active in ("0", "false", "no", "inactive"):
            qs = qs.filter(account_active=False)

        if is_active in ("1", "true", "yes", "active", "enabled"):
            qs = qs.filter(is_active=True)
        elif is_active in ("0", "false", "no", "inactive", "blocked", "disabled"):
            qs = qs.filter(is_active=False)

        if activated in ("1", "true", "yes", "activated"):
            qs = qs.filter(first_purchase_activated_at__isnull=False)
        elif activated in ("0", "false", "no", "inactive", "not_activated", "unactivated", "notactivated"):
            qs = qs.filter(first_purchase_activated_at__isnull=True)
        if search:
            qs = qs.filter(
                Q(username__icontains=search)
                | Q(full_name__icontains=search)
                | Q(email__icontains=search)
                | Q(unique_id__icontains=search)
            )

        ordering = (self.request.query_params.get("ordering") or "-date_joined").strip()
        if ordering in ("-date_joined", "date_joined"):
            qs = qs.order_by("-date_joined", "-id")
        else:
            qs = qs.order_by(ordering)
        return qs

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        # List view context: keep response lightweight and avoid decryption and heavy work
        ctx["purpose"] = "list"
        ctx["consumer_columns"] = self._consumer_columns_preset()
        return ctx

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        try:
            page = int(request.query_params.get("page") or 1)
        except Exception:
            page = 1
        try:
            page_size = int(request.query_params.get("page_size") or 25)
        except Exception:
            page_size = 25
        page = max(1, page)
        page_size = max(1, min(page_size, 200))

        # Fast mode: skip COUNT(*) and compute has_next via one extra row
        fast_param = str(request.query_params.get("fast") or "").strip().lower()
        if fast_param in ("1", "true", "yes"):
            start = (page - 1) * page_size
            limit = page_size + 1
            items = list(qs[start:start + limit])
            has_next = len(items) > page_size
            if has_next:
                items = items[:page_size]
            serializer = self.get_serializer(items, many=True)
            return Response(
                {
                    "count": None,
                    "page": page,
                    "page_size": page_size,
                    "has_next": bool(has_next),
                    "has_prev": start > 0,
                    "results": serializer.data,
                },
                status=200,
            )

        # Compute total with a lightweight queryset (no prefetch/annotate) to avoid slow COUNT(*)
        base = CustomUser.objects.all()
        role = (request.query_params.get("role") or "").strip()
        phone = (request.query_params.get("phone") or "").strip()
        category = (request.query_params.get("category") or "").strip()
        pincode = (request.query_params.get("pincode") or "").strip()
        state_id = (request.query_params.get("state") or "").strip()
        kyc = (request.query_params.get("kyc") or "").strip()
        search = (request.query_params.get("search") or "").strip()
        activated = (request.query_params.get("activated") or "").strip().lower()
        account_active = (request.query_params.get("account_active") or "").strip().lower()

        if role:
            r = str(role).strip()
            r_key = r.lower()
            try:
                role_keys = {str(k).lower(): k for k, _ in getattr(CustomUser, "ROLE_CHOICES", [])}
                role_labels = {str(v).lower().replace(" ", "_").replace("-", "_"): k for k, v in getattr(CustomUser, "ROLE_CHOICES", [])}
                r_norm = role_keys.get(r_key) or role_labels.get(r_key.replace(" ", "_").replace("-", "_")) or r
            except Exception:
                r_norm = r
            base = base.filter(role__iexact=r_norm)

        if category:
            c = str(category).strip()
            c_key = c.lower().replace(" ", "_").replace("-", "_")
            try:
                if "cordinator" in c_key:
                    c_key = c_key.replace("cordinator", "coordinator")
                if "subfranchise" in c_key and "sub_franchise" not in c_key:
                    c_key = c_key.replace("subfranchise", "sub_franchise")
            except Exception:
                pass
            try:
                cat_values = {str(k).lower(): k for k, _ in getattr(CustomUser, "CATEGORY_CHOICES", [])}
                cat_labels = {str(v).lower().replace(" ", "_").replace("-", "_"): k for k, v in getattr(CustomUser, "CATEGORY_CHOICES", [])}
                c_norm = cat_values.get(c_key) or cat_labels.get(c_key) or c_key
            except Exception:
                c_norm = c_key
            base = base.filter(category__iexact=c_norm)

        if phone:
            base = base.filter(phone__icontains=phone)
        if pincode:
            base = base.filter(pincode__icontains=pincode)
        if state_id and state_id.isdigit():
            base = base.filter(state_id=int(state_id))
        if kyc:
            if kyc == "pending":
                base = base.filter(Q(kyc__verified=False) | Q(kyc__isnull=True))
            elif kyc == "verified":
                base = base.filter(kyc__verified=True)

        if account_active in ("1", "true", "yes", "active"):
            base = base.filter(account_active=True)
        elif account_active in ("0", "false", "no", "inactive"):
            base = base.filter(account_active=False)

        if activated in ("1", "true", "yes", "activated"):
            base = base.filter(first_purchase_activated_at__isnull=False)
        elif activated in ("0", "false", "no", "inactive", "not_activated", "unactivated", "notactivated"):
            base = base.filter(first_purchase_activated_at__isnull=True)

        if search:
            base = base.filter(
                Q(username__icontains=search)
                | Q(full_name__icontains=search)
                | Q(email__icontains=search)
                | Q(unique_id__icontains=search)
            )

        total = base.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = self.get_serializer(qs[start:end], many=True)
        total_pages = (total + page_size - 1) // page_size
        return Response(
            {
                "count": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "has_next": end < total,
                "has_prev": start > 0,
                "results": serializer.data,
            },
            status=200,
        )


class AdminUserCategoryCountsView(APIView):
    """
    Aggregate user counts for key categories in a single request.
    Returns a map of category -> count.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("users")]

    def get(self, request):
        categories = [
            "consumer",
            "agency_state_coordinator",
            "agency_state",
            "agency_sub_franchise",
            "employee",
            "merchant",
        ]
        agg = CustomUser.objects.aggregate(
            **{c: Count("id", filter=Q(category=c)) for c in categories}
        )
        return Response(agg, status=200)


class AdminUsersExportXLSX(APIView):
    """
    Export Admin Users grid to XLSX with full details.
    - Applies the same filters as AdminUsersList (via query params).
    - Excludes usernames in the "9000000" series by default.
      Pass include_9000000=1 to include them.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("users"), HasAnyPermission("manage_users", "show_users")]

    def get(self, request):
        # Lazy import so server still boots if openpyxl is missing
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except Exception:
            return Response({"detail": "openpyxl is not installed on the server"}, status=400)

        # Reuse list view queryset with current filters
        view = AdminUsersList()
        view.request = request
        qs = view.get_queryset()

        # Exclude usernames that are in 9000000 series by default
        # (covers usernames that start with or contain 9000000, e.g. test seeds)
        include_9m = str(request.query_params.get("include_9000000") or "").lower() in ("1", "true", "yes")
        if not include_9m:
            qs = qs.exclude(Q(username__startswith="9000000") | Q(username__icontains="9000000"))

        # Serialize with 'detail' purpose to compute richer fields (e.g., kyc, wallet summary)
        ser = AdminUserNodeSerializer(qs, many=True, context={"request": request, "purpose": "detail"})
        items = ser.data

        # Column order mirrors Admin Users grid (plus a few essentials)
        headers = [
            "SI No", "Edit/View", "Login", "Active/Inactive", "Joining Date", "Package Buy Date",
            "System Serial Number", "User ID", "User Name", "Sponsor ID & Name", "Address & Pincode",
            "KYC Profile", "Subscription 1", "Subscription 2", "Subscription 3", "Smart Product Package",
            "Digital Education", "Certified Training Certificate", "Package Purchase Gift Card Received",
            "New Tour Package", "Team Consumer Self Package ID Count", "Team Consumer Block ID",
            "Total Earning", "Main Wallet", "Coupon Pocket", "Self Package Pocket", "Withdrawal Pocket",
            "Redeem Points", "Add Money", "Withdrawal To Pocket", "Coupon Pocket Breakdown",
            "Self Package Pocket Details", "Coupon Admin Charges", "Withdrawal Admin Charges",
            "Package Admin Charges", "Franchisee Reference Reward", "Zonal Reward",
            "Direct Sponsor Benefit", "Level Income Benefit", "Smart Product Pocket", "SPP Spin & Win",
            "Digital Education Prime Package Spin & Win", "Shopping Self Re-birth ID Count",
            "Franchisee Self Rebirth ID Count", "Caption/Coupon Self Re-birth ID Count",
            "Raw ID", "Username", "Email", "Role", "Category", "Phone", "Wallet Status",
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(headers)

        for idx, obj in enumerate(items, start=1):
            ws.append([
                idx,
                "Available in admin UI",
                "Available in admin UI",
                "Active" if obj.get("account_active") else "Inactive",
                obj.get("date_joined") or "",
                obj.get("package_buy_date") or "",
                obj.get("system_serial_number") or obj.get("id") or "",
                obj.get("user_code") or obj.get("username") or "",
                obj.get("full_name") or obj.get("username") or "",
                obj.get("sponsor_display") or obj.get("sponsor_id") or "",
                obj.get("address_pincode") or obj.get("pincode") or "",
                obj.get("kyc_status") or "",
                obj.get("subscription_1") or "",
                obj.get("subscription_2") or "",
                obj.get("subscription_3") or "",
                obj.get("smart_product_package") or "",
                obj.get("digital_education") or "",
                "Needs source",
                "Needs source",
                obj.get("new_tour_package") or "",
                "Needs business rule",
                "Needs business rule",
                obj.get("total_earning") if obj.get("total_earning") not in (None, "") else "",
                obj.get("main_wallet") if obj.get("main_wallet") not in (None, "") else "",
                obj.get("coupon_pocket") if obj.get("coupon_pocket") not in (None, "") else "",
                obj.get("self_package_pocket") if obj.get("self_package_pocket") not in (None, "") else "",
                obj.get("withdrawal_pocket") if obj.get("withdrawal_pocket") not in (None, "") else "",
                obj.get("redeem_points") if obj.get("redeem_points") not in (None, "") else "",
                obj.get("add_money_pocket") if obj.get("add_money_pocket") not in (None, "") else "",
                obj.get("withdrawal_to_pocket") if obj.get("withdrawal_to_pocket") not in (None, "") else "",
                obj.get("coupon_pocket_breakdown") or "",
                obj.get("self_package_pocket_details") or "",
                "Needs charge rule",
                "Needs charge rule",
                "Needs charge rule",
                "Needs aggregation",
                "Needs aggregation",
                "Needs aggregation",
                "Needs aggregation",
                "Needs aggregation",
                "Needs source",
                "Needs source",
                "Needs business rule",
                "Needs business rule",
                "Needs business rule",
                obj.get("id"),
                obj.get("username") or "",
                obj.get("email") or "",
                obj.get("role") or "",
                obj.get("category") or "",
                obj.get("phone") or "",
                obj.get("wallet_status") or "",
            ])

        # Auto-size columns
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                val = "" if row[0].value is None else str(row[0].value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"admin_users_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

class AdminUserEditMetaView(APIView):
    """
    Return dynamic field metadata for Admin user edit dialog based on AdminUserEditSerializer.
    """
    permission_classes = [IsAdminOrStaff]

    def get(self, request):
        from .serializers import AdminUserEditSerializer
        try:
            meta = field_meta_from_serializer(AdminUserEditSerializer) or []

            # Normalize base meta
            for f in meta:
                name = f.get("name")
                if name == "password":
                    f["type"] = "PasswordField"
                    f["required"] = False
                    f["label"] = f.get("label") or "Set New Password"
                if name == "sponsor_id":
                    # Allow superuser or admins with manage_users/edit_users permissions to modify Sponsor ID
                    is_super = bool(getattr(request.user, "is_superuser", False))
                    can_edit = is_super
                    if not can_edit:
                        try:
                            from .permissions import get_effective_permissions
                            perms = get_effective_permissions(request.user)
                            can_edit = ("*" in perms) or ("manage_users" in perms) or ("edit_users" in perms)
                        except Exception:
                            can_edit = False
                    f["read_only"] = not can_edit
                    if not can_edit:
                        hint = "Only superuser or admins with Manage/Edit Users permission can modify Sponsor ID"
                        if f.get("help_text"):
                            f["help_text"] = f"{f['help_text']} | {hint}"
                        else:
                            f["help_text"] = hint

            # Restrict to allowed fields only (Admin Users editor shows geo fields via custom UI)
            allowed = {"full_name", "email", "phone", "sponsor_id", "account_active"}
            meta = [f for f in meta if f.get("name") in allowed]

            # Reorder and add helpful labels/hints
            order = ["full_name", "email", "phone", "sponsor_id", "account_active"]
            meta_map = {f["name"]: f for f in meta}
            out = []
            for key in order:
                if key in meta_map:
                    f = meta_map[key]
                    if key == "full_name":
                        f["label"] = "Full Name"
                    if key == "email":
                        f["label"] = "Email"
                    if key == "phone":
                        f["label"] = "Phone"
                        hint = "On save, Username is auto-synced to this phone. Downline users whose Sponsor ID equals the old username will be updated to the new one."
                        f["help_text"] = (f.get("help_text") + " | " + hint).strip(" |") if f.get("help_text") else hint
                    if key == "sponsor_id":
                        f["label"] = "Sponsor ID"
                        # read_only already set above based on permissions; keep any help_text from earlier
                    if key == "account_active":
                        f["label"] = "Account Active"
                    out.append(f)
            meta = out
        except Exception:
            # Fallback minimal meta (restricted to allowed fields)
            is_super = bool(getattr(request.user, "is_superuser", False))
            can_edit = is_super
            if not can_edit:
                try:
                    from .permissions import get_effective_permissions
                    perms = get_effective_permissions(request.user)
                    can_edit = ("*" in perms) or ("manage_users" in perms) or ("edit_users" in perms)
                except Exception:
                    can_edit = False
            meta = [
                {"name": "full_name", "type": "CharField", "required": False, "label": "Full Name"},
                {"name": "email", "type": "EmailField", "required": False, "label": "Email"},
                {"name": "phone", "type": "CharField", "required": False, "label": "Phone", "help_text": "Username will be synced to this phone on save; sponsor cascades accordingly."},
                {"name": "sponsor_id", "type": "CharField", "required": False, "label": "Sponsor ID", "read_only": (not can_edit)},
                {"name": "account_active", "type": "BooleanField", "required": False, "label": "Account Active"},
            ]
        return Response({"fields": meta}, status=200)


class AdminUserDetail(APIView):
    """
    Retrieve/Update admin-editable fields for a specific user.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("users"), HasAnyPermission("manage_users", "show_users", "edit_users")]

    def get(self, request, pk: int):
        profile = str(request.query_params.get("profile") or "").strip().lower() in ("1", "true", "yes")
        qs = CustomUser.objects.filter(pk=pk)
        if profile:
            qs = qs.select_related("country", "state", "city", "wallet", "kyc", "registered_by", "merchant_profile", "reward_points_account").prefetch_related(
                Prefetch("wallet_accounts", queryset=WalletAccount.objects.only("user_id", "wallet_type", "current_balance"), to_attr="prefetched_wallet_accounts"),
                Prefetch("promo_purchases", queryset=PromoPurchase.objects.select_related("package").filter(status="APPROVED").order_by("-approved_at", "-id"), to_attr="approved_promo_purchases"),
            )
        user = qs.first()
        if not user:
            return Response({"detail": "Not found"}, status=404)
        if profile:
            data = AdminUserNodeSerializer(user, context={"request": request, "purpose": "detail"}).data
            return Response(data, status=200)
        data = AdminUserEditSerializer(user).data
        return Response(data, status=200)

    def patch(self, request, pk: int):
        user = CustomUser.objects.filter(pk=pk).first()
        if not user:
            return Response({"detail": "Not found"}, status=404)
        # RBAC: edit requires edit_users or manage_users (superuser bypass)
        try:
            from .permissions import get_effective_permissions
            u = request.user
            if not getattr(u, "is_superuser", False):
                perms = get_effective_permissions(u)
                if "*" not in perms and (("edit_users" not in perms) and ("manage_users" not in perms)):
                    return Response({"detail": "Forbidden"}, status=403)
        except Exception:
            return Response({"detail": "Forbidden"}, status=403)
        old_sponsor = getattr(user, "registered_by", None)
        serializer = AdminUserEditSerializer(user, data=request.data, partial=True, context={"request": request})
        if serializer.is_valid():
            obj = serializer.save()
            new_sponsor = getattr(obj, "registered_by", None)
            if new_sponsor and (not old_sponsor or old_sponsor.id != new_sponsor.id):
                try:
                    from business.models import AutoPoolAccount
                    AutoPoolAccount.reanchor_user_to_sponsor(obj, new_sponsor, "FIVE_150")
                    AutoPoolAccount.reanchor_user_to_sponsor(obj, new_sponsor, "THREE_150")
                except Exception:
                    pass
                try:
                    from mlm_ranks.services.five_matrix import FiveMatrixService
                    FiveMatrixService.reparent_rank_matrix_node(obj, new_sponsor)
                except Exception:
                    pass
            return Response(AdminUserEditSerializer(obj).data, status=200)
        return Response(serializer.errors, status=400)

    def delete(self, request, pk: int):
        """
        Permanently delete a user. Guards:
          - Cannot delete self
          - Cannot delete superuser
        Returns 204 No Content on success.
        """
        if not getattr(request.user, "is_superuser", False):
            return Response({"detail": "Only superuser can delete users."}, status=403)
        user = CustomUser.objects.filter(pk=pk).first()
        if not user:
            return Response({"detail": "Not found"}, status=404)
        # Prevent accidental self-delete
        if getattr(request.user, "id", None) == user.id:
            return Response({"detail": "You cannot delete your own admin account."}, status=400)
        # Prevent deleting superusers
        if getattr(user, "is_superuser", False):
            return Response({"detail": "Cannot delete a superuser."}, status=400)
        try:
            user.delete()
            return Response(status=204)
        except Exception as e:
            return Response({"detail": str(e)}, status=400)


class AdminUserImpersonateView(APIView):
    """
    Admin-only: mint JWT tokens for a specific user to allow 'view as' login.
    """
    permission_classes = [IsSuperAdmin]

    def post(self, request, pk: int):
        user = CustomUser.objects.filter(pk=pk).first()
        if not user:
            return Response({"detail": "Not found"}, status=404)
        if not getattr(user, "is_active", False):
            return Response({"detail": "User is blocked. Unblock before login."}, status=400)
        try:
            refresh = RefreshToken.for_user(user)
            data = {
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "role": getattr(user, "role", "user") or "user",
                "username": getattr(user, "username", None),
                "id": getattr(user, "id", None),
            }
            return Response(data, status=200)
        except Exception as e:
            return Response({"detail": str(e)}, status=400)


class AdminUserSetTempPasswordView(APIView):
    """
    Admin-only: generate a secure temporary password for a user, set it,
    and store an encrypted copy for display in Admin Users grid.
    Response: { "temp_password": "..." }
    """
    permission_classes = [IsSuperAdmin]

    def post(self, request, pk: int):
        user = CustomUser.objects.filter(pk=pk).first()
        if not user:
            return Response({"detail": "Not found"}, status=404)

        # Generate a strong temporary password: 8 alnum + 1 special + 1 digit
        try:
            import secrets, string
            base = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
            special = secrets.choice('!@#$%^&*')
            digit = str(secrets.randbelow(10))
            pwd = f"{base}{special}{digit}"
        except Exception:
            # Extremely rare, fallback
            pwd = "Tr!k0-" + str(int(timezone.now().timestamp()))[-6:]

        try:
            user.set_password(pwd)
            # Store encrypted plaintext for admin visibility
            try:
                from core.crypto import encrypt_string
                enc = encrypt_string(pwd)
            except Exception:
                enc = None

            if enc:
                user.last_password_encrypted = enc
                user.save(update_fields=["password", "last_password_encrypted"])
            else:
                user.save(update_fields=["password"])
        except Exception as e:
            return Response({"detail": str(e)}, status=400)

        return Response({"temp_password": pwd}, status=200)


class AdminUserWalletAdjustView(APIView):
    """
    Admin-only: adjust a user's wallet by credit or debit.
    Body: { "action": "credit" | "debit", "amount": number, "note": "optional", "type": "optional override type" }
    - Credits go to main balance with no withholding.
    - Debits reduce balances using Wallet.debit rules.
    Response: { balance, main_balance, withdrawable_balance }
    """
    permission_classes = [IsSuperAdmin]

    def post(self, request, pk: int):
        user = CustomUser.objects.filter(pk=pk).first()
        if not user:
            return Response({"detail": "Not found"}, status=404)

        data = request.data or {}
        action = str(data.get("action") or "").strip().lower()
        note = str(data.get("note") or "").strip()
        tx_type_in = str(data.get("type") or "").strip()

        # Coerce amount
        from decimal import Decimal as D
        try:
            amount = D(str(data.get("amount")))
        except Exception:
            return Response({"detail": "amount must be a number"}, status=400)
        if amount <= 0:
            return Response({"detail": "amount must be > 0"}, status=400)

        if action not in ("credit", "debit"):
            return Response({"detail": "action must be 'credit' or 'debit'"}, status=400)

        w = Wallet.get_or_create_for_user(user)
        meta = {"note": note, "by_admin": getattr(request.user, "username", None) or ""}

        try:
            if action == "credit":
                tx_type = tx_type_in or "ADJUSTMENT_CREDIT"
                # No withholding on manual admin credit
                w.credit(amount, tx_type=tx_type, meta={**meta, "no_withhold": True}, source_type="ADMIN", source_id=str(getattr(request.user, "id", "")))
            else:
                tx_type = tx_type_in or "ADJUSTMENT_DEBIT"
                w.debit(amount, tx_type=tx_type, meta=meta, source_type="ADMIN", source_id=str(getattr(request.user, "id", "")))
        except Exception as e:
            return Response({"detail": str(e)}, status=400)

        # Fresh balances
        try:
            w_refreshed = Wallet.objects.get(pk=w.pk)
        except Exception:
            w_refreshed = w
        return Response(
            {
                "balance": float(w_refreshed.balance or 0),
                "main_balance": float(w_refreshed.main_balance or 0),
                "withdrawable_balance": float(w_refreshed.withdrawable_balance or 0),
            },
            status=200,
        )


def _wallet_bucket_sums(user):
    return _wallet_bucket_sums_from_qs(WalletTransaction.objects.filter(user=user))


def _wallet_bucket_sums_from_qs(tx):

    def sum_types(credit_types, debit_types=None):
        credit = tx.filter(type__in=list(credit_types or []), amount__gt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        debit = tx.filter(type__in=list(debit_types or []), amount__lt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        return (Decimal(str(credit)) + Decimal(str(debit))).quantize(Decimal("0.01"))

    direct = tx.filter(type__in=["DIRECT_REF_BONUS"], amount__gt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    level = tx.filter(type__in=["LEVEL_BONUS", "AUTOPOOL_BONUS_THREE", "AUTOPOOL_BONUS_FIVE"], amount__gt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    service_charge = tx.filter(source_type="ADMIN_SERVICE_CHARGE", amount__lt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    return {
        "coupon": sum_types(["COUPON_WALLET_CREDIT", "COUPON_WALLET_REFUND"], ["COUPON_WALLET_DEBIT", "VOUCHER_CREATE_DEBIT"]),
        "shopping": sum_types(["SHOPPING_WALLET_CREDIT"], ["SHOPPING_WALLET_DEBIT"]),
        "self_package": sum_types(["INTERNAL_WALLET_CREDIT"], ["INTERNAL_WALLET_DEBIT"]),
        "package_purchase_coupon": sum_types(["PACKAGE_COUPON_WALLET_CREDIT", "VOUCHER_REDEEM_CREDIT"], ["PACKAGE_COUPON_WALLET_DEBIT"]),
        "wallet_to_wallet": sum_types(["WALLET_TO_WALLET_IN"], ["WALLET_TO_WALLET_OUT"]),
        "direct_benefit": Decimal(str(direct)).quantize(Decimal("0.01")),
        "level_benefit": Decimal(str(level)).quantize(Decimal("0.01")),
        "admin_service_charges": abs(Decimal(str(service_charge))).quantize(Decimal("0.01")),
    }


def _empty_wallet_buckets():
    return {
        "coupon": Decimal("0.00"),
        "shopping": Decimal("0.00"),
        "self_package": Decimal("0.00"),
        "package_purchase_coupon": Decimal("0.00"),
        "wallet_to_wallet": Decimal("0.00"),
        "direct_benefit": Decimal("0.00"),
        "level_benefit": Decimal("0.00"),
        "admin_service_charges": Decimal("0.00"),
    }


def _wallet_bucket_sums_for_users(user_ids):
    ids = [int(x) for x in (user_ids or []) if x]
    if not ids:
        return {}
    rows = (
        WalletTransaction.objects
        .filter(user_id__in=ids)
        .values("user_id")
        .annotate(
            coupon_credit=Sum("amount", filter=Q(type__in=["COUPON_WALLET_CREDIT", "COUPON_WALLET_REFUND"], amount__gt=0)),
            coupon_debit=Sum("amount", filter=Q(type__in=["COUPON_WALLET_DEBIT", "VOUCHER_CREATE_DEBIT"], amount__lt=0)),
            shopping_credit=Sum("amount", filter=Q(type__in=["SHOPPING_WALLET_CREDIT"], amount__gt=0)),
            shopping_debit=Sum("amount", filter=Q(type__in=["SHOPPING_WALLET_DEBIT"], amount__lt=0)),
            self_package_credit=Sum("amount", filter=Q(type__in=["INTERNAL_WALLET_CREDIT"], amount__gt=0)),
            self_package_debit=Sum("amount", filter=Q(type__in=["INTERNAL_WALLET_DEBIT"], amount__lt=0)),
            package_coupon_credit=Sum("amount", filter=Q(type__in=["PACKAGE_COUPON_WALLET_CREDIT", "VOUCHER_REDEEM_CREDIT"], amount__gt=0)),
            package_coupon_debit=Sum("amount", filter=Q(type__in=["PACKAGE_COUPON_WALLET_DEBIT"], amount__lt=0)),
            wallet_transfer_credit=Sum("amount", filter=Q(type__in=["WALLET_TO_WALLET_IN"], amount__gt=0)),
            wallet_transfer_debit=Sum("amount", filter=Q(type__in=["WALLET_TO_WALLET_OUT"], amount__lt=0)),
            direct_benefit=Sum("amount", filter=Q(type__in=["DIRECT_REF_BONUS"], amount__gt=0)),
            level_benefit=Sum("amount", filter=Q(type__in=["LEVEL_BONUS", "AUTOPOOL_BONUS_THREE", "AUTOPOOL_BONUS_FIVE"], amount__gt=0)),
            admin_service_charges=Sum("amount", filter=Q(source_type="ADMIN_SERVICE_CHARGE", amount__lt=0)),
        )
    )
    out = {}
    for row in rows:
        def dec(key):
            return Decimal(str(row.get(key) or "0.00"))

        out[row["user_id"]] = {
            "coupon": (dec("coupon_credit") + dec("coupon_debit")).quantize(Decimal("0.01")),
            "shopping": (dec("shopping_credit") + dec("shopping_debit")).quantize(Decimal("0.01")),
            "self_package": (dec("self_package_credit") + dec("self_package_debit")).quantize(Decimal("0.01")),
            "package_purchase_coupon": (dec("package_coupon_credit") + dec("package_coupon_debit")).quantize(Decimal("0.01")),
            "wallet_to_wallet": (dec("wallet_transfer_credit") + dec("wallet_transfer_debit")).quantize(Decimal("0.01")),
            "direct_benefit": dec("direct_benefit").quantize(Decimal("0.01")),
            "level_benefit": dec("level_benefit").quantize(Decimal("0.01")),
            "admin_service_charges": abs(dec("admin_service_charges")).quantize(Decimal("0.01")),
        }
    return out


def _wallet_summary_payload(user, buckets=None, create_missing_wallet=True):
    if create_missing_wallet:
        w = Wallet.get_or_create_for_user(user)
    else:
        try:
            w = user.wallet
        except Exception:
            w = None
    buckets = buckets if buckets is not None else _wallet_bucket_sums(user)
    return {
        "user_id": user.id,
        "username": user.username,
        "prefixed_id": getattr(user, "prefixed_id", "") or "",
        "full_name": getattr(user, "full_name", "") or "",
        "phone": getattr(user, "phone", "") or "",
        "category": getattr(user, "category", "") or "",
        "balance": str(getattr(w, "balance", 0) or 0),
        "main_balance": str(getattr(w, "main_balance", 0) or 0),
        "withdrawable_balance": str(getattr(w, "withdrawable_balance", 0) or 0),
        "self_account_balance": str(getattr(w, "self_account_balance", 0) or 0),
        "pockets": {k: str(v) for k, v in buckets.items()},
        "updated_at": getattr(w, "updated_at", None),
    }


def _tx_payload(tx):
    return {
        "id": tx.id,
        "user_id": tx.user_id,
        "username": getattr(tx.user, "username", "") if getattr(tx, "user", None) else "",
        "amount": str(tx.amount),
        "balance_after": str(tx.balance_after),
        "type": tx.type,
        "source_type": tx.source_type or "",
        "source_id": tx.source_id or "",
        "meta": tx.meta or {},
        "created_at": tx.created_at,
    }


def _admin_name(user):
    if not user:
        return ""
    return getattr(user, "username", "") or getattr(user, "prefixed_id", "") or str(getattr(user, "id", ""))


def _entry_payload(entry):
    tx = getattr(entry, "financial_transaction", None)
    account = getattr(entry, "wallet_account", None)
    return {
        "id": entry.id,
        "entry_ref": entry.entry_ref,
        "transaction_id": getattr(tx, "id", None),
        "transaction_ref": getattr(tx, "transaction_ref", ""),
        "wallet_account_id": getattr(account, "id", None),
        "wallet_type": getattr(account, "wallet_type", ""),
        "user_id": entry.user_id,
        "username": _admin_name(getattr(entry, "user", None)),
        "direction": entry.direction,
        "amount": str(entry.amount),
        "balance_before": str(entry.balance_before),
        "balance_after": str(entry.balance_after),
        "status": entry.status,
        "remarks": entry.remarks,
        "metadata": entry.metadata or {},
        "created_at": entry.created_at,
    }


def _finance_tx_payload(tx, *, include_entries=False):
    entries = list(getattr(tx, "prefetched_ledger_entries", []) or [])
    if not entries and include_entries:
        entries = list(tx.ledger_entries.select_related("wallet_account", "user").order_by("id")[:100])
    wallet_types = []
    before_balance = ""
    after_balance = ""
    source_destination = []
    for entry in entries:
        account = getattr(entry, "wallet_account", None)
        wallet_type = getattr(account, "wallet_type", "")
        if wallet_type and wallet_type not in wallet_types:
            wallet_types.append(wallet_type)
        source_destination.append(f"{entry.direction}:{_admin_name(getattr(entry, 'user', None))}:{wallet_type}")
        if before_balance == "":
            before_balance = str(entry.balance_before)
        after_balance = str(entry.balance_after)
    meta = tx.metadata or {}
    return {
        "id": tx.id,
        "transaction_id": tx.transaction_ref,
        "transaction_ref": tx.transaction_ref,
        "flow_id": tx.flow_id,
        "user_id": tx.user_id,
        "username": _admin_name(getattr(tx, "user", None)),
        "wallet_type": ", ".join(wallet_types),
        "wallet_types": wallet_types,
        "transaction_type": tx.category,
        "category": tx.category,
        "mlm_income_type": meta.get("mlm_income_type") or meta.get("reward_type") or meta.get("income_type") or "",
        "source_module": tx.source_module,
        "source_id": tx.source_id,
        "destination_module": tx.destination_module,
        "source_destination": " | ".join(source_destination),
        "gross_amount": str(tx.gross_amount),
        "service_charge": str(tx.charges_amount),
        "charges_amount": str(tx.charges_amount),
        "gst_amount": str(tx.gst_amount),
        "tds_amount": str(tx.tds_amount),
        "net_amount": str(tx.net_amount),
        "before_balance": before_balance,
        "after_balance": after_balance,
        "status": tx.status,
        "approval_status": tx.approval_status,
        "created_by": _admin_name(getattr(tx, "created_by", None)),
        "approved_by": _admin_name(getattr(tx, "approved_by", None)),
        "created_at": tx.created_at,
        "updated_at": tx.updated_at,
        "approved_at": tx.approved_at,
        "payment_gateway_reference": tx.payment_gateway_reference,
        "utr_number": tx.utr_number,
        "reference_id": tx.reference_id,
        "legacy_wallet_transaction_id": tx.legacy_wallet_transaction_id,
        "remarks": tx.remarks,
        "metadata": meta,
        "linked_transaction_flow": _money_flow_steps(tx, entries),
        "ledger_entries": [_entry_payload(e) for e in entries] if include_entries else [],
    }


def _money_flow_steps(tx, entries=None):
    meta = tx.metadata or {}
    category = str(tx.category or "")
    steps = []
    if category == "PACKAGE_PURCHASE":
        steps = ["Package Purchase", "Sponsor Commission", "Level Bonus", "Wallet Credit", "Withdrawal", "Settlement"]
    elif category in {"MLM_INCOME", "SPONSOR_INCOME", "MATRIX_INCOME", "REWARD_DISTRIBUTION", "FRANCHISE_REWARD"}:
        steps = ["Source Event", "MLM Rule", "Payout Calculation", "Wallet Credit", "Withdrawal Eligibility"]
    elif category == "WITHDRAWAL":
        steps = ["Withdrawal Request", "Approval", "Wallet Debit", "Payout Reference", "Settlement"]
    elif category in {"VOUCHER_CREATE", "VOUCHER_REDEEM"}:
        steps = ["Voucher Created", "Wallet Hold/Debit", "Redemption", "Wallet Credit/Refund", "Expiry Review"]
    elif category == "ADD_MONEY":
        steps = ["Payment Proof", "Admin Approval", "Wallet Credit", "Ledger Reconcile"]
    else:
        steps = ["Transaction Created", "Ledger Posting", "Balance Updated", "Reconciliation"]
    return [
        {
            "label": label,
            "status": "completed" if tx.status in {"COMPLETED", "REVERSED"} else ("failed" if tx.status == "FAILED" else "pending"),
            "reference": tx.transaction_ref if index == 0 else (meta.get("flow_id") or tx.flow_id or tx.reference_id or ""),
        }
        for index, label in enumerate(steps)
    ]


def _finance_filtered_queryset(request, user_id=None):
    qs = (
        FinancialTransaction.objects
        .select_related("user", "created_by", "approved_by")
        .prefetch_related(
            Prefetch(
                "ledger_entries",
                queryset=LedgerEntry.objects.select_related("wallet_account", "user").order_by("id"),
                to_attr="prefetched_ledger_entries",
            )
        )
        .order_by("-created_at", "-id")
    )
    if user_id is not None:
        qs = qs.filter(Q(user_id=user_id) | Q(ledger_entries__user_id=user_id)).distinct()
    q = str(request.query_params.get("q") or "").strip()
    wallet_type = str(request.query_params.get("wallet_type") or request.query_params.get("source_type") or "").strip()
    category = str(request.query_params.get("category") or request.query_params.get("type") or "").strip()
    source_module = str(request.query_params.get("source_module") or "").strip()
    mlm_income_type = str(request.query_params.get("mlm_income_type") or "").strip()
    status_filter = str(request.query_params.get("status") or "").strip()
    package = str(request.query_params.get("package") or "").strip()
    voucher = str(request.query_params.get("voucher") or "").strip()
    withdrawal = str(request.query_params.get("withdrawal") or "").strip()
    if q:
        qs = qs.filter(
            Q(transaction_ref__icontains=q) |
            Q(flow_id__icontains=q) |
            Q(source_id__icontains=q) |
            Q(reference_id__icontains=q) |
            Q(utr_number__icontains=q) |
            Q(user__username__icontains=q) |
            Q(user__prefixed_id__icontains=q)
        )
    if wallet_type:
        qs = qs.filter(ledger_entries__wallet_account__wallet_type=wallet_type).distinct()
    if category:
        qs = qs.filter(category=category)
    if source_module:
        qs = qs.filter(source_module__icontains=source_module)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if mlm_income_type:
        qs = qs.filter(category__in=["MLM_INCOME", "SPONSOR_INCOME", "MATRIX_INCOME", "REWARD_DISTRIBUTION"])
    if package:
        qs = qs.filter(Q(source_module__icontains="PACKAGE") | Q(source_id__icontains=package))
    if voucher:
        qs = qs.filter(Q(source_module__icontains="VOUCHER") | Q(reference_id__icontains=voucher))
    if withdrawal:
        qs = qs.filter(Q(source_module__icontains="WITHDRAWAL") | Q(reference_id__icontains=withdrawal) | Q(source_id__icontains=withdrawal))
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    return qs


def _paginate(request, qs, *, default_page_size=50, max_page_size=200):
    try:
        page = max(1, int(request.query_params.get("page") or 1))
        page_size = min(max_page_size, max(1, int(request.query_params.get("page_size") or default_page_size)))
    except Exception:
        page, page_size = 1, default_page_size
    total = qs.count()
    start = (page - 1) * page_size
    return page, page_size, total, qs[start:start + page_size]


def _csv_response(filename, headers, rows):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def _voucher_payload(v):
    return {
        "id": v.id,
        "code": v.code,
        "voucher_type": v.voucher_type,
        "amount": str(v.amount),
        "status": v.status,
        "creator_id": v.creator_id,
        "creator_username": getattr(v.creator, "username", "") if getattr(v, "creator", None) else "",
        "assigned_to_id": v.assigned_to_id,
        "assigned_to_username": getattr(v.assigned_to, "username", "") if getattr(v, "assigned_to", None) else "",
        "redeemed_by_id": v.redeemed_by_id,
        "redeemed_by_username": getattr(v.redeemed_by, "username", "") if getattr(v, "redeemed_by", None) else "",
        "created_at": v.created_at,
        "expires_at": v.expires_at,
        "redeemed_at": v.redeemed_at,
        "expired_at": v.expired_at,
        "note": v.note,
    }


class AdminWalletListView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        q = str(request.query_params.get("q") or "").strip()
        category = str(request.query_params.get("category") or "").strip()
        try:
            page = max(1, int(request.query_params.get("page") or 1))
            page_size = min(100, max(1, int(request.query_params.get("page_size") or 25)))
        except Exception:
            page, page_size = 1, 25

        users = CustomUser.objects.select_related("wallet").order_by("-id")
        if q:
            users = users.filter(
                Q(username__icontains=q) |
                Q(prefixed_id__icontains=q) |
                Q(full_name__icontains=q) |
                Q(phone__icontains=q) |
                Q(email__icontains=q)
            )
        if category:
            users = users.filter(category=category)
        total = users.count()
        start = (page - 1) * page_size
        page_users = list(users[start:start + page_size])
        bucket_map = _wallet_bucket_sums_for_users([u.id for u in page_users])
        rows = [
            _wallet_summary_payload(
                u,
                buckets=bucket_map.get(u.id, _empty_wallet_buckets()),
                create_missing_wallet=False,
            )
            for u in page_users
        ]
        return Response({"count": total, "page": page, "page_size": page_size, "results": rows})


class AdminWalletDetailView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request, user_id: int):
        user = CustomUser.objects.filter(pk=user_id).first()
        if not user:
            return Response({"detail": "User not found."}, status=404)
        payload = _wallet_summary_payload(user)
        accounts = list(WalletAccount.objects.filter(user=user).order_by("wallet_type"))
        account_ids = [a.id for a in accounts]
        entry_totals = {
            row["wallet_account_id"]: row
            for row in (
                LedgerEntry.objects
                .filter(wallet_account_id__in=account_ids)
                .values("wallet_account_id")
                .annotate(
                    total_credits=Sum("amount", filter=Q(direction="CREDIT")),
                    total_debits=Sum("amount", filter=Q(direction="DEBIT")),
                )
            )
        }
        payload["finance_accounts"] = [
            {
                "id": account.id,
                "wallet_type": account.wallet_type,
                "current_balance": str(account.current_balance),
                "available_balance": str(account.available_balance),
                "locked_balance": str(account.locked_balance),
                "pending_balance": str(account.pending_balance),
                "total_credits": str(Decimal(str(entry_totals.get(account.id, {}).get("total_credits") or "0.00")).quantize(Decimal("0.01"))),
                "total_debits": str(Decimal(str(entry_totals.get(account.id, {}).get("total_debits") or "0.00")).quantize(Decimal("0.01"))),
                "status": account.status,
                "updated_at": account.updated_at,
            }
            for account in accounts
        ]
        category_rows = (
            FinancialTransaction.objects
            .filter(Q(user=user) | Q(ledger_entries__user=user))
            .values("category")
            .annotate(total=Sum("net_amount"), count=Count("id", distinct=True))
            .order_by("category")
        )
        payload["income_breakdown"] = [
            {"category": row["category"], "total": str(row["total"] or "0.00"), "count": row["count"]}
            for row in category_rows
        ]
        payload["transaction_timeline"] = [
            _finance_tx_payload(tx, include_entries=True)
            for tx in _finance_filtered_queryset(request, user_id=user_id)[:25]
        ]
        payload["audit_timeline"] = [
            {
                "id": a.id,
                "actor": _admin_name(getattr(a, "actor_user", None)),
                "action": a.action,
                "module": a.resource_type,
                "reference_id": a.resource_id,
                "ip_address": a.ip_address,
                "device": a.user_agent,
                "created_at": a.created_at,
            }
            for a in AuditLog.objects.select_related("actor_user").filter(
                Q(actor_user=user) | Q(resource_id=str(user.id))
            ).order_by("-created_at", "-id")[:25]
        ]
        return Response(payload)


class AdminWalletLedgerView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request, user_id: int | None = None):
        if str(request.query_params.get("legacy") or "").lower() in {"1", "true", "yes"}:
            qs = WalletTransaction.objects.select_related("user").order_by("-created_at", "-id")
            if user_id is not None:
                qs = qs.filter(user_id=user_id)
            q = str(request.query_params.get("q") or "").strip()
            tx_type = str(request.query_params.get("type") or "").strip()
            source_type = str(request.query_params.get("source_type") or "").strip()
            if q:
                qs = qs.filter(
                    Q(user__username__icontains=q) |
                    Q(user__prefixed_id__icontains=q) |
                    Q(source_id__icontains=q) |
                    Q(source_type__icontains=q) |
                    Q(type__icontains=q)
                )
            if tx_type:
                qs = qs.filter(type=tx_type)
            if source_type:
                qs = qs.filter(source_type=source_type)
            date_from = request.query_params.get("date_from")
            date_to = request.query_params.get("date_to")
            if date_from:
                qs = qs.filter(created_at__date__gte=date_from)
            if date_to:
                qs = qs.filter(created_at__date__lte=date_to)
            if str(request.query_params.get("export") or "").lower() == "csv":
                rows = [_tx_payload(x) for x in qs[:10000]]
                return _csv_response(
                    "legacy-wallet-ledger.csv",
                    ["id", "user", "type", "source_type", "source_id", "amount", "balance_after", "created_at"],
                    [
                        [
                            row.get("id"),
                            row.get("username"),
                            row.get("type"),
                            row.get("source_type"),
                            row.get("source_id"),
                            row.get("amount"),
                            row.get("balance_after"),
                            row.get("created_at"),
                        ]
                        for row in rows
                    ],
                )
            page, page_size, total, rows = _paginate(request, qs)
            return Response({"count": total, "page": page, "page_size": page_size, "results": [_tx_payload(x) for x in rows]})

        qs = _finance_filtered_queryset(request, user_id=user_id)
        if str(request.query_params.get("export") or "").lower() == "csv":
            rows = [_finance_tx_payload(x, include_entries=True) for x in qs[:10000]]
            return _csv_response(
                "finance-ledger.csv",
                ["transaction_id", "user", "wallet_type", "category", "source_module", "gross_amount", "service_charge", "gst", "tds", "net_amount", "status", "created_by", "approved_by", "created_at", "reference_id"],
                [
                    [
                        row.get("transaction_ref"),
                        row.get("username"),
                        row.get("wallet_type"),
                        row.get("category"),
                        row.get("source_module"),
                        row.get("gross_amount"),
                        row.get("service_charge"),
                        row.get("gst_amount"),
                        row.get("tds_amount"),
                        row.get("net_amount"),
                        row.get("status"),
                        row.get("created_by"),
                        row.get("approved_by"),
                        row.get("created_at"),
                        row.get("reference_id"),
                    ]
                    for row in rows
                ],
            )
        page, page_size, total, rows = _paginate(request, qs)
        return Response({"count": total, "page": page, "page_size": page_size, "results": [_finance_tx_payload(x, include_entries=True) for x in rows]})


class AdminWalletAdjustPocketView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def post(self, request, user_id: int):
        user = CustomUser.objects.filter(pk=user_id).first()
        if not user:
            return Response({"detail": "User not found."}, status=404)
        action = str(request.data.get("action") or "").strip().lower()
        pocket = str(request.data.get("pocket") or "main").strip().lower()
        note = str(request.data.get("note") or "").strip()
        try:
            amount = Decimal(str(request.data.get("amount") or "0")).quantize(Decimal("0.01"))
        except Exception:
            return Response({"detail": "Invalid amount."}, status=400)
        if action not in {"credit", "debit"} or amount <= 0:
            return Response({"detail": "Use action credit/debit and amount > 0."}, status=400)

        credit_types = {
            "main": "ADJUSTMENT_CREDIT",
            "coupon": "COUPON_WALLET_CREDIT",
            "shopping": "SHOPPING_WALLET_CREDIT",
            "self_package": "INTERNAL_WALLET_CREDIT",
            "withdrawal": "WITHDRAWAL_WALLET_CREDIT",
            "package_purchase_coupon": "PACKAGE_COUPON_WALLET_CREDIT",
        }
        debit_types = {
            "main": "ADJUSTMENT_DEBIT",
            "coupon": "COUPON_WALLET_DEBIT",
            "shopping": "SHOPPING_WALLET_DEBIT",
            "self_package": "INTERNAL_WALLET_DEBIT",
            "withdrawal": "WITHDRAWAL_DEBIT",
            "package_purchase_coupon": "PACKAGE_COUPON_WALLET_DEBIT",
        }
        if pocket not in credit_types:
            return Response({"detail": "Invalid pocket."}, status=400)

        signed = amount if action == "credit" else amount * Decimal("-1")
        tx_type = credit_types[pocket] if action == "credit" else debit_types[pocket]
        with transaction.atomic():
            w = Wallet.get_or_create_for_user(user)
            w = Wallet.objects.select_for_update().get(pk=w.pk)
            if pocket == "main":
                if action == "debit" and (w.main_balance or Decimal("0")) < amount:
                    return Response({"detail": "Insufficient main wallet balance."}, status=400)
                w.main_balance = (w.main_balance or Decimal("0")) + signed
            elif pocket == "withdrawal":
                if action == "debit" and (w.withdrawable_balance or Decimal("0")) < amount:
                    return Response({"detail": "Insufficient withdrawal pocket balance."}, status=400)
                w.withdrawable_balance = (w.withdrawable_balance or Decimal("0")) + signed
            elif pocket == "self_package":
                if action == "debit" and (w.self_account_balance or Decimal("0")) < amount:
                    return Response({"detail": "Insufficient self package pocket balance."}, status=400)
                w.self_account_balance = (w.self_account_balance or Decimal("0")) + signed
            else:
                if action == "debit" and _wallet_bucket_sums(user).get(pocket, Decimal("0")) < amount:
                    return Response({"detail": f"Insufficient {pocket} balance."}, status=400)

            w.save(update_fields=["main_balance", "withdrawable_balance", "self_account_balance", "updated_at"])
            tx = WalletTransaction.objects.create(
                user=user,
                amount=signed,
                balance_after=w.balance,
                type=tx_type,
                source_type="ADMIN_MANUAL",
                source_id=str(getattr(request.user, "id", "")),
                meta={"pocket": pocket, "note": note, "by_admin": getattr(request.user, "username", "")},
            )

            # Sync to WalletAccount via WalletEngine
            if pocket in {"coupon", "shopping", "self_package", "package_purchase_coupon"}:
                try:
                    from accounts.finance_constants import WalletTypes, FinanceCategories, LedgerDirections
                    from accounts.wallet_engine import WalletEngine, LedgerPosting

                    pocket_wallet_map = {
                        "coupon": WalletTypes.COUPON_POCKET,
                        "shopping": WalletTypes.SHOPPING_REBIRTH,
                        "self_package": WalletTypes.SELF_PACKAGE_POCKET,
                        "package_purchase_coupon": WalletTypes.PACKAGE_PURCHASE_COUPON,
                    }
                    wtype = pocket_wallet_map[pocket]
                    system_user = WalletEngine.get_system_user()
                    
                    if action == "credit":
                        postings = [
                            LedgerPosting(system_user, WalletTypes.SYSTEM, LedgerDirections.DEBIT, amount, metadata={"counterparty_user_id": user.id}),
                            LedgerPosting(user, wtype, LedgerDirections.CREDIT, amount, metadata={"pocket": pocket, "note": note, "by_admin": getattr(request.user, "username", "")}),
                        ]
                    else:
                        postings = [
                            LedgerPosting(user, wtype, LedgerDirections.DEBIT, amount, metadata={"pocket": pocket, "note": note, "by_admin": getattr(request.user, "username", "")}),
                            LedgerPosting(system_user, WalletTypes.SYSTEM, LedgerDirections.CREDIT, amount, metadata={"counterparty_user_id": user.id}),
                        ]
                    
                    WalletEngine.post_transaction(
                        category=FinanceCategories.ADMIN_ADJUSTMENT,
                        user=user,
                        source_module="ADMIN_MANUAL",
                        source_id=str(getattr(request.user, "id", "")),
                        destination_module=wtype,
                        gross_amount=amount,
                        net_amount=amount,
                        idempotency_key=f"admin_adjust:{tx.id}",
                        legacy_wallet_transaction=tx,
                        created_by=request.user,
                        approved_by=request.user,
                        remarks=note or f"Admin manual adjustment for {pocket}",
                        metadata={"pocket": pocket, "action": action, "by_admin": getattr(request.user, "username", "")},
                        postings=postings,
                    )
                except Exception:
                    pass

        return Response(_wallet_summary_payload(user))


def _admin_reward_consumer_payload(user):
    return {
        "id": user.id,
        "username": user.username,
        "prefixed_id": getattr(user, "prefixed_id", "") or "",
        "full_name": getattr(user, "full_name", "") or "",
        "phone": getattr(user, "phone", "") or "",
        "email": getattr(user, "email", "") or "",
        "category": getattr(user, "category", "") or "",
    }


def _resolve_admin_reward_consumer(identifier):
    value = str(identifier or "").strip()
    if not value:
        return None
    query = Q(username__iexact=value) | Q(prefixed_id__iexact=value) | Q(unique_id__iexact=value)
    digits = "".join(ch for ch in value if ch.isdigit())
    if digits:
        query = query | Q(phone__iexact=digits)
    return CustomUser.objects.filter(query, category="consumer").first()


class AdminTeamRewardValidateConsumerView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        username = request.query_params.get("username") or request.query_params.get("q") or ""
        user = _resolve_admin_reward_consumer(username)
        if not user:
            return Response({"detail": "Consumer username not found."}, status=404)
        return Response({"consumer": _admin_reward_consumer_payload(user)}, status=200)


class AdminMatrixReanchorUserView(APIView):
    permission_classes = [IsAdminOrStaff]

    def post(self, request):
        data = request.data or {}
        username = data.get("username") or data.get("user_id") or data.get("q")
        target_sponsor_input = data.get("target_sponsor_id") or data.get("sponsor_id") or data.get("target_sponsor")
        pool_type = str(data.get("pool_type") or "FIVE_150").upper()

        if not username:
            return Response({"detail": "username or user_id is required."}, status=400)

        user = CustomUser.objects.filter(Q(username=username) | Q(id=str(username) if str(username).isdigit() else 0) | Q(phone=username)).first()
        if not user:
            return Response({"detail": "User not found."}, status=404)

        if target_sponsor_input:
            target_sponsor = CustomUser.objects.filter(Q(username=target_sponsor_input) | Q(id=str(target_sponsor_input) if str(target_sponsor_input).isdigit() else 0) | Q(phone=target_sponsor_input)).first()
        else:
            target_sponsor = getattr(user, "registered_by", None)

        if not target_sponsor:
            return Response({"detail": "Target sponsor not found."}, status=404)

        from business.models import AutoPoolAccount
        success = AutoPoolAccount.reanchor_user_to_sponsor(user, target_sponsor, pool_type)
        if success:
            return Response({
                "message": f"Successfully re-anchored user {user.username} under sponsor {target_sponsor.username} in {pool_type} matrix.",
                "user": user.username,
                "target_sponsor": target_sponsor.username,
                "pool_type": pool_type
            }, status=200)
        else:
            return Response({"detail": "Failed to re-anchor user. Ensure target sponsor has an active matrix account."}, status=400)


class AdminTeamRewardCreditView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    REWARD_TYPES = {
        "franchise_reference": {
            "source_type": "ADMIN_FRANCHISE_REFERENCE_REWARD",
            "label": "Franchise Reference Reward",
        },
        "zonal": {
            "source_type": "ADMIN_ZONAL_REWARD",
            "label": "Zonal Reward",
        },
    }

    def post(self, request):
        data = request.data or {}
        reward_type = str(data.get("reward_type") or "").strip().lower()
        config = self.REWARD_TYPES.get(reward_type)
        if not config:
            return Response({"detail": "reward_type must be franchise_reference or zonal."}, status=400)

        user = _resolve_admin_reward_consumer(data.get("username") or data.get("consumer") or data.get("q"))
        if not user:
            return Response({"detail": "Consumer username not found."}, status=404)

        try:
            amount = Decimal(str(data.get("amount") or "0")).quantize(Decimal("0.01"))
        except Exception:
            return Response({"detail": "Invalid amount."}, status=400)
        if amount <= 0:
            return Response({"detail": "Amount must be greater than 0."}, status=400)

        note = str(data.get("note") or "").strip()
        with transaction.atomic():
            wallet = Wallet.get_or_create_for_user(user)
            wallet.credit(
                amount,
                tx_type="REWARD_CREDIT",
                source_type=config["source_type"],
                source_id=str(getattr(request.user, "id", "")),
                meta={
                    "reward_type": reward_type,
                    "reward_label": config["label"],
                    "note": note,
                    "by_admin": getattr(request.user, "username", "") or "",
                    "admin_user_id": getattr(request.user, "id", None),
                    "no_withhold": True,
                },
            )

        return Response(
            {
                "detail": f"{config['label']} credited successfully.",
                "consumer": _admin_reward_consumer_payload(user),
                "reward_type": reward_type,
                "amount": str(amount),
                "wallet": _wallet_summary_payload(user),
            },
            status=201,
        )


class AdminWalletVoucherListView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        now = timezone.now()
        due = ConsumerVoucher.objects.filter(status=ConsumerVoucher.STATUS_ACTIVE, expires_at__lte=now)
        for voucher in due[:500]:
            self._expire_and_refund(voucher)

        qs = ConsumerVoucher.objects.select_related("creator", "assigned_to", "redeemed_by").order_by("-created_at", "-id")
        q = str(request.query_params.get("q") or "").strip()
        status_filter = str(request.query_params.get("status") or "").strip().upper()
        voucher_type = str(request.query_params.get("voucher_type") or "").strip().upper()
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(creator__username__icontains=q) | Q(assigned_to__username__icontains=q) | Q(redeemed_by__username__icontains=q))
        if status_filter:
            qs = qs.filter(status=status_filter)
        if voucher_type:
            qs = qs.filter(voucher_type=voucher_type)
        try:
            page = max(1, int(request.query_params.get("page") or 1))
            page_size = min(100, max(1, int(request.query_params.get("page_size") or 50)))
        except Exception:
            page, page_size = 1, 50
        total = qs.count()
        start = (page - 1) * page_size
        return Response({"count": total, "page": page, "page_size": page_size, "results": [_voucher_payload(v) for v in qs[start:start + page_size]]})

    @staticmethod
    def _expire_and_refund(voucher):
        if voucher.status != ConsumerVoucher.STATUS_ACTIVE or voucher.refund_transaction_id:
            return
        with transaction.atomic():
            voucher = ConsumerVoucher.objects.select_for_update().get(pk=voucher.pk)
            if voucher.status != ConsumerVoucher.STATUS_ACTIVE or voucher.refund_transaction_id:
                return
            w = Wallet.get_or_create_for_user(voucher.creator)
            w = Wallet.objects.select_for_update().get(pk=w.pk)
            # Remove direct legacy balance credit to prevent double-crediting to MAIN pocket.
            # Double-entry posting is handled below via post_system_credit to the COUPON_POCKET.
            tx = WalletTransaction.objects.create(
                user=voucher.creator,
                amount=voucher.amount,
                balance_after=w.balance,
                type="COUPON_WALLET_REFUND",
                source_type="VOUCHER_EXPIRED",
                source_id=str(voucher.id),
                meta={"voucher_code": voucher.code, "voucher_type": voucher.voucher_type, "by_admin_job": True},
            )
            try:
                from accounts.finance_constants import WalletTypes, FinanceCategories
                from accounts.wallet_engine import WalletEngine
                WalletEngine.post_system_credit(
                    user=voucher.creator,
                    wallet_type=WalletTypes.COUPON_POCKET,
                    amount=voucher.amount,
                    category=FinanceCategories.REFUND,
                    source_module="CONSUMER_VOUCHER",
                    source_id=str(voucher.id),
                    idempotency_key=f"voucher_expire_refund:{voucher.id}",
                    legacy_wallet_transaction=tx,
                    actor=None,
                    remarks="Voucher expired, refunded to coupon pocket",
                    metadata={"voucher_code": voucher.code, "voucher_type": voucher.voucher_type},
                )
            except Exception:
                pass
            voucher.status = ConsumerVoucher.STATUS_EXPIRED
            voucher.expired_at = timezone.now()
            voucher.refund_transaction = tx
            voucher.save(update_fields=["status", "expired_at", "refund_transaction"])


class AdminWalletVoucherCancelRefundView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def post(self, request, pk: int):
        voucher = ConsumerVoucher.objects.filter(pk=pk).select_related("creator").first()
        if not voucher:
            return Response({"detail": "Voucher not found."}, status=404)
        if voucher.status != ConsumerVoucher.STATUS_ACTIVE:
            return Response({"detail": "Only active vouchers can be cancelled/refunded."}, status=400)
        with transaction.atomic():
            voucher = ConsumerVoucher.objects.select_for_update().get(pk=pk)
            w = Wallet.get_or_create_for_user(voucher.creator)
            w = Wallet.objects.select_for_update().get(pk=w.pk)
            # Remove direct legacy balance credit to prevent double-crediting to MAIN pocket.
            # Double-entry posting is handled below via post_system_credit to the COUPON_POCKET.
            tx = WalletTransaction.objects.create(
                user=voucher.creator,
                amount=voucher.amount,
                balance_after=w.balance,
                type="COUPON_WALLET_REFUND",
                source_type="VOUCHER_CANCEL_REFUND",
                source_id=str(voucher.id),
                meta={"voucher_code": voucher.code, "by_admin": getattr(request.user, "username", ""), "reason": str(request.data.get("reason") or "")},
            )
            try:
                from accounts.finance_constants import WalletTypes, FinanceCategories
                from accounts.wallet_engine import WalletEngine
                WalletEngine.post_system_credit(
                    user=voucher.creator,
                    wallet_type=WalletTypes.COUPON_POCKET,
                    amount=voucher.amount,
                    category=FinanceCategories.REFUND,
                    source_module="CONSUMER_VOUCHER",
                    source_id=str(voucher.id),
                    idempotency_key=f"voucher_cancel_refund:{voucher.id}",
                    legacy_wallet_transaction=tx,
                    actor=request.user,
                    remarks=f"Voucher cancelled/refunded by admin: {getattr(request.user, 'username', '')}",
                    metadata={"voucher_code": voucher.code, "reason": str(request.data.get("reason") or "")},
                )
            except Exception:
                pass
            voucher.status = ConsumerVoucher.STATUS_CANCELLED
            voucher.refund_transaction = tx
            voucher.save(update_fields=["status", "refund_transaction"])
        return Response(_voucher_payload(voucher))


class AdminWalletReconcileView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        q = str(request.query_params.get("q") or "").strip()
        users = CustomUser.objects.select_related("wallet").order_by("-id")
        if q:
            users = users.filter(Q(username__icontains=q) | Q(prefixed_id__icontains=q) | Q(full_name__icontains=q) | Q(phone__icontains=q))
        limit = min(500, max(1, int(request.query_params.get("limit") or 100)))
        page_users = list(users[:limit])
        finance_accounts = list(WalletAccount.objects.select_related("user").order_by("-updated_at", "-id")[:limit])
        account_ids = [a.id for a in finance_accounts]
        user_ids = list(set([u.id for u in page_users] + [fa.user_id for fa in finance_accounts if fa.user_id]))
        bucket_map = _wallet_bucket_sums_for_users(user_ids)
        
        # Pre-fetch transactions for all users in a single query
        from collections import defaultdict
        user_txs = defaultdict(list)
        for tx in WalletTransaction.objects.filter(user_id__in=user_ids).order_by('id'):
            user_txs[tx.user_id].append(tx)

        ledger_map = {}
        legacy_main_map = {}
        legacy_self_map = {}
        for uid in user_ids:
            sim_main = Decimal('0.00')
            sim_self = Decimal('0.00')
            for tx in user_txs[uid]:
                amt = tx.amount
                tx_type = tx.type
                
                # Calculate Main Balance
                if tx_type in [
                    'INCOME_CREDIT_75', 'DIRECT_REF_BONUS', 'WELCOME_BONUS', 'LEVEL_BONUS',
                    'AUTOPOOL_BONUS_FIVE', 'AUTOPOOL_BONUS_THREE', 'GLOBAL_ROYALTY',
                    'LIFETIME_WITHDRAWAL_BONUS', 'REWARD_CREDIT', 'FRANCHISE_INCOME',
                    'COMMISSION_CREDIT', 'ADJUSTMENT_CREDIT'
                ] and amt > 0:
                    sim_main += amt
                elif tx_type in [
                    'WITHDRAWAL_WALLET_TRANSFER_OUT', 'COUPON_WALLET_TRANSFER_OUT',
                    'ADJUSTMENT_DEBIT', 'INTERNAL_WALLET_TRANSFER_OUT'
                ]:
                    sim_main = max(Decimal('0.00'), sim_main - abs(amt))
                elif tx_type == 'INTERNAL_WALLET_DEBIT':
                    wallet_source = (tx.meta or {}).get('wallet_source') or ''
                    if wallet_source == 'internal':
                        sim_main = max(Decimal('0.00'), sim_main - abs(amt))
                        
                # Calculate Self Balance
                if tx_type == 'SELF_ACCOUNT_CREDIT':
                    sim_self += amt
                elif tx_type in ['SELF_ACCOUNT_DEBIT', 'AUTO_PURCHASE_DEBIT']:
                    sim_self = max(Decimal('0.00'), sim_self - abs(amt))
                    
            ledger_map[uid] = (sim_main + sim_self).quantize(Decimal("0.01"))
            legacy_main_map[uid] = sim_main.quantize(Decimal("0.01"))
            legacy_self_map[uid] = sim_self.quantize(Decimal("0.01"))

        rows = []
        for user in page_users:
            try:
                w = user.wallet
            except Exception:
                w = None
            ledger_total = legacy_main_map.get(user.id, Decimal("0.00"))
            wallet_balance = Decimal(str(getattr(w, "main_balance", 0) or 0))
            diff = (wallet_balance - Decimal(str(ledger_total))).quantize(Decimal("0.01"))
            rows.append({
                **_wallet_summary_payload(
                    user,
                    buckets=bucket_map.get(user.id, _empty_wallet_buckets()),
                    create_missing_wallet=False,
                ),
                "ledger_total": str(ledger_total),
                "balance_vs_ledger_diff": str(diff),
                "status": "OK" if abs(diff) <= Decimal("1.00") else "MISMATCH",
            })
        mismatches = sum(1 for r in rows if r["status"] != "OK")

        ledger_totals = {
            row["wallet_account_id"]: (
                Decimal(str(row.get("credit_total") or "0.00")) -
                Decimal(str(row.get("debit_total") or "0.00"))
            ).quantize(Decimal("0.01"))
            for row in (
                LedgerEntry.objects
                .filter(wallet_account_id__in=account_ids)
                .values("wallet_account_id")
                .annotate(
                    credit_total=Sum("amount", filter=Q(direction="CREDIT")),
                    debit_total=Sum("amount", filter=Q(direction="DEBIT")),
                )
            )
        }
        finance_rows = []
        for account in finance_accounts:
            ledger_balance = ledger_totals.get(account.id, Decimal("0.00"))
            
            # Add the legacy starting balance to double-entry ledger balance
            legacy_start = Decimal("0.00")
            from accounts.finance_constants import WalletTypes
            if account.wallet_type == WalletTypes.MAIN:
                legacy_start = legacy_main_map.get(account.user_id, Decimal("0.00"))
            elif account.wallet_type == WalletTypes.SELF_PACKAGE_POCKET:
                legacy_start = legacy_self_map.get(account.user_id, Decimal("0.00"))
            elif account.wallet_type == WalletTypes.WITHDRAWAL_WALLET:
                try:
                    w = account.user.wallet
                    legacy_start = Decimal(str(w.withdrawable_balance or "0.00"))
                except Exception:
                    legacy_start = Decimal("0.00")
            
            total_derived = ledger_balance + legacy_start
            stored_balance = Decimal(str(account.current_balance or "0.00")).quantize(Decimal("0.01"))
            diff = (stored_balance - total_derived).quantize(Decimal("0.01"))
            finance_rows.append({
                "wallet_account_id": account.id,
                "user_id": account.user_id,
                "username": getattr(account.user, "username", "") if account.user_id else "",
                "wallet_type": account.wallet_type,
                "stored_balance": str(stored_balance),
                "ledger_balance": str(total_derived.quantize(Decimal("0.01"))),
                "diff": str(diff),
                "status": "OK" if abs(diff) <= Decimal("25.00") else "MISMATCH",
            })
        finance_mismatches = sum(1 for r in finance_rows if r["status"] != "OK")
        return Response({
            "count": len(rows),
            "mismatches": mismatches,
            "results": rows,
            "finance_ledger": {
                "count": len(finance_rows),
                "mismatches": finance_mismatches,
                "results": finance_rows,
            },
        })


class AdminFinanceOverviewView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        today = timezone.now().date()
        month_start = today.replace(day=1)
        tx = FinancialTransaction.objects.all()
        today_tx = tx.filter(created_at__date=today)
        month_tx = tx.filter(created_at__date__gte=month_start)
        withdrawals = WithdrawalRequest.objects.all()
        vouchers = ConsumerVoucher.objects.all()
        risk = _risk_alerts()
        data = {
            "transaction_volume": str(tx.aggregate(total=Sum("gross_amount")).get("total") or "0.00"),
            "today_volume": str(today_tx.aggregate(total=Sum("gross_amount")).get("total") or "0.00"),
            "month_volume": str(month_tx.aggregate(total=Sum("gross_amount")).get("total") or "0.00"),
            "pending_withdrawals": withdrawals.filter(status="pending").count(),
            "failed_transactions": tx.filter(status="FAILED").count(),
            "mlm_payout_volume": str(tx.filter(category__in=["MLM_INCOME", "SPONSOR_INCOME", "MATRIX_INCOME", "REWARD_DISTRIBUTION"]).aggregate(total=Sum("net_amount")).get("total") or "0.00"),
            "voucher_activity": vouchers.count(),
            "tax_collected": str(tx.aggregate(total=Sum("gst_amount")).get("total") or "0.00"),
            "tds_deducted": str(tx.aggregate(total=Sum("tds_amount")).get("total") or "0.00"),
            "company_revenue": str(tx.aggregate(total=Sum("charges_amount")).get("total") or "0.00"),
            "active_users": CustomUser.objects.filter(account_active=True).count(),
            "suspicious_activity_alerts": len(risk),
            "alerts": risk[:8],
        }
        return Response(data)


class AdminFinanceTaxDashboardView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        period = str(request.query_params.get("period") or "daily").lower()
        if period == "yearly":
            trunc = TruncYear("created_at")
        elif period == "monthly":
            trunc = TruncMonth("created_at")
        else:
            trunc = TruncDate("created_at")
        qs = _finance_filtered_queryset(request).order_by()
        rows = (
            qs.annotate(period_bucket=trunc).values("period_bucket")
            .annotate(
                gross_amount=Sum("gross_amount"),
                service_charge=Sum("charges_amount"),
                gst_amount=Sum("gst_amount"),
                tds_amount=Sum("tds_amount"),
                net_amount=Sum("net_amount"),
                transfer_charges=Sum("charges_amount", filter=Q(category="WALLET_TRANSFER")),
                withdrawal_charges=Sum("charges_amount", filter=Q(category="WITHDRAWAL")),
                voucher_fees=Sum("charges_amount", filter=Q(category__in=["VOUCHER_CREATE", "VOUCHER_REDEEM"])),
                package_charges=Sum("charges_amount", filter=Q(category="PACKAGE_PURCHASE")),
                count=Count("id"),
            )
            .order_by("-period_bucket")[:370]
        )
        summary = qs.aggregate(
            gross_amount=Sum("gross_amount"),
            service_charge=Sum("charges_amount"),
            gst_amount=Sum("gst_amount"),
            tds_amount=Sum("tds_amount"),
            net_amount=Sum("net_amount"),
            platform_revenue=Sum("charges_amount"),
        )
        if str(request.query_params.get("export") or "").lower() == "csv":
            return _csv_response(
                "tax-service-charge-summary.csv",
                ["period", "gross_amount", "service_charge", "gst_amount", "tds_amount", "net_amount", "transfer_charges", "withdrawal_charges", "voucher_fees", "package_charges", "count"],
                [
                    [
                        row.get("period_bucket"),
                        row.get("gross_amount") or "0.00",
                        row.get("service_charge") or "0.00",
                        row.get("gst_amount") or "0.00",
                        row.get("tds_amount") or "0.00",
                        row.get("net_amount") or "0.00",
                        row.get("transfer_charges") or "0.00",
                        row.get("withdrawal_charges") or "0.00",
                        row.get("voucher_fees") or "0.00",
                        row.get("package_charges") or "0.00",
                        row.get("count") or 0,
                    ]
                    for row in rows
                ],
            )
        return Response({
            "period": period,
            "summary": {k: str(v or "0.00") for k, v in summary.items()},
            "results": [
                {("period" if k == "period_bucket" else k): (str(v or "0.00") if k not in {"period_bucket", "count"} else v) for k, v in row.items()}
                for row in rows
            ],
        })


def _risk_alerts():
    alerts = []
    recent = timezone.now() - timedelta(days=7)
    duplicate_utrs = (
        WalletUploadRequest.objects
        .filter(requested_at__gte=recent)
        .exclude(utr="")
        .values("utr")
        .annotate(count=Count("id"))
        .filter(count__gt=1)
        [:10]
    )
    for row in duplicate_utrs:
        alerts.append({"type": "DUPLICATE_UTR", "severity": "high", "message": f"Duplicate UTR uploaded {row.get('utr')}", "count": row["count"]})
    duplicate_withdrawals = (
        WithdrawalRequest.objects
        .filter(requested_at__gte=recent)
        .values("user_id", "amount", "status")
        .annotate(count=Count("id"))
        .filter(count__gt=1)[:10]
    )
    for row in duplicate_withdrawals:
        alerts.append({"type": "DUPLICATE_WITHDRAWAL", "severity": "medium", "message": f"Repeated withdrawal user #{row['user_id']} amount {row['amount']}", "count": row["count"]})
    high_frequency = (
        WalletTransaction.objects
        .filter(created_at__gte=timezone.now() - timedelta(hours=24))
        .values("user_id")
        .annotate(count=Count("id"), volume=Sum("amount"))
        .filter(count__gte=15)[:10]
    )
    for row in high_frequency:
        alerts.append({"type": "HIGH_FREQUENCY_WALLET_MOVEMENT", "severity": "medium", "message": f"High wallet activity for user #{row['user_id']}", "count": row["count"], "volume": str(row.get("volume") or "0.00")})
    duplicate_vouchers = (
        ConsumerVoucher.objects
        .values("code")
        .annotate(count=Count("id"))
        .filter(count__gt=1)[:10]
    )
    for row in duplicate_vouchers:
        alerts.append({"type": "DUPLICATE_VOUCHER_CODE", "severity": "high", "message": f"Duplicate voucher code {row['code']}", "count": row["count"]})
    return alerts


class AdminFinanceRiskView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        return Response({"count": len(_risk_alerts()), "results": _risk_alerts()})


class AdminFinanceFlowView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request, transaction_ref: str):
        tx = (
            FinancialTransaction.objects
            .select_related("user", "created_by", "approved_by")
            .prefetch_related(Prefetch("ledger_entries", queryset=LedgerEntry.objects.select_related("wallet_account", "user").order_by("id"), to_attr="prefetched_ledger_entries"))
            .filter(Q(transaction_ref=transaction_ref) | Q(flow_id=transaction_ref) | Q(reference_id=transaction_ref))
            .first()
        )
        if not tx:
            return Response({"detail": "Transaction flow not found."}, status=404)
        related_query = Q(reference_id=tx.transaction_ref)
        if tx.flow_id:
            related_query |= Q(flow_id=tx.flow_id)
        if tx.source_id:
            related_query |= Q(source_id=tx.source_id)
        related = FinancialTransaction.objects.filter(related_query).exclude(id=tx.id).select_related("user").order_by("created_at", "id")[:50]
        return Response({
            "transaction": _finance_tx_payload(tx, include_entries=True),
            "nodes": _money_flow_steps(tx, getattr(tx, "prefetched_ledger_entries", [])),
            "related_transactions": [_finance_tx_payload(row, include_entries=True) for row in related],
        })


class AdminAuditTimelineView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        qs = AuditLog.objects.select_related("actor_user").order_by("-created_at", "-id")
        q = str(request.query_params.get("q") or "").strip()
        module = str(request.query_params.get("module") or "").strip()
        action = str(request.query_params.get("action") or "").strip()
        if q:
            qs = qs.filter(Q(actor_user__username__icontains=q) | Q(resource_id__icontains=q) | Q(resource_type__icontains=q) | Q(action__icontains=q))
        if module:
            qs = qs.filter(resource_type__icontains=module)
        if action:
            qs = qs.filter(action__icontains=action)
        page, page_size, total, rows = _paginate(request, qs, default_page_size=50, max_page_size=100)
        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "results": [
                {
                    "id": row.id,
                    "actor": _admin_name(getattr(row, "actor_user", None)),
                    "actor_user_id": row.actor_user_id,
                    "action": row.action,
                    "module": row.resource_type,
                    "reference_id": row.resource_id,
                    "before_json": row.before_json,
                    "after_json": row.after_json,
                    "ip_address": row.ip_address,
                    "device": row.user_agent,
                    "created_at": row.created_at,
                }
                for row in rows
            ],
        })


class AdminECouponBulkCreateView(APIView):
    """
    Bulk-generate E-Coupons with prefix (default 'ELC') and sequential serials.
    Creates CouponBatch for traceability and CouponCode rows with issued_channel='e_coupon'.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("ecoupons")]

    def post(self, request):
        from django.db import transaction

        data = request.data or {}
        prefix = (data.get("prefix") or "ELC").strip().upper()
        try:
            quantity = int(data.get("quantity") or 0)
        except Exception:
            quantity = 0
        if quantity <= 0:
            return Response({"detail": "quantity must be a positive integer"}, status=400)

        serial_start_in = data.get("serialStart")
        try:
            serial_start_in = int(serial_start_in) if serial_start_in is not None else None
        except Exception:
            serial_start_in = None

        # Optional unit value per coupon (defaults to 150)
        try:
            unit_value = Decimal(str(data.get("value") or "150"))
        except Exception:
            unit_value = Decimal("150.00")

        # Ensure there is a parent Coupon record for E-Coupons under the given prefix
        coupon, _ = Coupon.objects.get_or_create(
            code=prefix,
            defaults={
                "title": f"{prefix} E-Coupon",
                "description": "Electronic coupon",
                "campaign": prefix,
                "issuer_id": getattr(request.user, "id", None) or 1,
            },
        )

        # Determine starting serial (continue from latest if not provided)
        last = (
            CouponCode.objects.filter(coupon=coupon, issued_channel="e_coupon", code__startswith=prefix)
            .order_by("-serial")
            .first()
        )
        next_serial = (getattr(last, "serial", None) or 0) + 1
        serial_start = serial_start_in or max(1, next_serial)
        serial_end = serial_start + quantity - 1
        serial_width = 7  # ELC + 7 digits -> ELC0000001

        with transaction.atomic():
            batch = CouponBatch.objects.create(
                coupon=coupon,
                prefix=prefix,
                serial_start=serial_start,
                serial_end=serial_end,
                serial_width=serial_width,
                created_by=getattr(request.user, "id", None) and request.user or None,
            )

            objs = []
            for serial in range(serial_start, serial_end + 1):
                code_str = f"{prefix}{str(serial).zfill(serial_width)}"
                objs.append(
                    CouponCode(
                        code=code_str,
                        coupon=coupon,
                        issued_channel="e_coupon",
                        assigned_employee=None,
                        assigned_agency=None,
                        batch=batch,
                        serial=serial,
                        value=unit_value,
                        issued_by=request.user,
                        status="AVAILABLE",
                    )
                )
            # Ignore duplicates silently (idempotence across partial retries)
            CouponCode.objects.bulk_create(objs, ignore_conflicts=True)

        return Response(
            {
                "batchId": batch.id,
                "prefix": prefix,
                "serialStart": serial_start,
                "serialEnd": serial_end,
                "quantity": quantity,
                "codePreviewFirst": f"{prefix}{str(serial_start).zfill(serial_width)}",
                "codePreviewLast": f"{prefix}{str(serial_end).zfill(serial_width)}",
            },
            status=201,
        )


class AdminECouponAssignView(APIView):
    """
    Assign E-Coupons either by a serial range or by quantity (earliest available) to
    an Agency/Subfranchise/Employee.

    Body:
    {
      "method": "range" | "quantity",
      "prefix": "ELC",
      // when method=range
      "startSerial": 1,
      "endSerial": 100,
      // when method=quantity
      "quantity": 50,
      "entityType": "agency" | "subfranchise" | "employee",
      "entityId": 123
    }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("ecoupons")]

    def post(self, request):
        from django.db import transaction

        data = request.data or {}
        method = (data.get("method") or "").strip().lower()
        prefix = (data.get("prefix") or "ELC").strip().upper()
        entity_type = (data.get("entityType") or "").strip().lower()
        try:
            entity_id = int(data.get("entityId") or 0)
        except Exception:
            entity_id = 0

        if method not in ("range", "quantity"):
            return Response({"detail": "method must be 'range' or 'quantity'"}, status=400)
        if entity_type not in ("agency", "subfranchise", "employee"):
            return Response({"detail": "entityType must be one of: agency, subfranchise, employee"}, status=400)
        if entity_id <= 0:
            return Response({"detail": "entityId is required"}, status=400)

        target = CustomUser.objects.filter(id=entity_id).first()
        if not target:
            return Response({"detail": "Target entity not found"}, status=404)

        coupon = Coupon.objects.filter(code=prefix).first()
        if not coupon:
            return Response({"detail": f"No coupon definition found for prefix {prefix}. Create a batch first."}, status=400)

        base_qs = CouponCode.objects.filter(
            coupon=coupon, issued_channel="e_coupon", status="AVAILABLE", code__startswith=prefix
        ).order_by("serial")

        assigned = 0
        assigned_range = None

        with transaction.atomic():
            if method == "range":
                try:
                    start_serial = int(data.get("startSerial"))
                    end_serial = int(data.get("endSerial"))
                except Exception:
                    return Response({"detail": "startSerial and endSerial must be integers"}, status=400)
                if start_serial > end_serial:
                    return Response({"detail": "startSerial must be <= endSerial"}, status=400)

                qs = base_qs.filter(serial__gte=start_serial, serial__lte=end_serial)
                codes = list(qs)
                if not codes:
                    return Response({"detail": "No AVAILABLE codes in the requested range"}, status=400)

                if entity_type == "employee":
                    for c in codes:
                        c.assigned_employee = target
                        c.status = "ASSIGNED_EMPLOYEE"
                        c.assigned_agency = None
                else:
                    # agency or subfranchise -> use assigned_agency slot (model supports agency/employee)
                    for c in codes:
                        c.assigned_agency = target
                        c.status = "ASSIGNED_AGENCY"
                        c.assigned_employee = None
                CouponCode.objects.bulk_update(codes, ["assigned_employee", "assigned_agency", "status"])
                assigned = len(codes)
                assigned_range = {"startSerial": start_serial, "endSerial": end_serial}

            else:  # quantity
                try:
                    qty = int(data.get("quantity") or 0)
                except Exception:
                    qty = 0
                if qty <= 0:
                    return Response({"detail": "quantity must be a positive integer"}, status=400)

                codes = list(base_qs[:qty])
                if not codes:
                    return Response({"detail": "No AVAILABLE codes to assign"}, status=400)

                if entity_type == "employee":
                    for c in codes:
                        c.assigned_employee = target
                        c.status = "ASSIGNED_EMPLOYEE"
                        c.assigned_agency = None
                else:
                    for c in codes:
                        c.assigned_agency = target
                        c.status = "ASSIGNED_AGENCY"
                        c.assigned_employee = None
                CouponCode.objects.bulk_update(codes, ["assigned_employee", "assigned_agency", "status"])
                assigned = len(codes)
                if assigned:
                    assigned_range = {"startSerial": codes[0].serial, "endSerial": codes[-1].serial}

        return Response(
            {
                "assigned": assigned,
                "entityId": entity_id,
                "entityType": entity_type,
                "prefix": prefix,
                "range": assigned_range,
            },
            status=200,
        )


class AdminKYCList(ListAPIView):
    """
    List KYC records with filters and search.
    Filters:
      - status=pending|verified
      - user (id or username contains)
      - state (id)
      - pincode (contains)
      - date_from, date_to on updated_at
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("kyc")]
    serializer_class = AdminKYCSerializer

    def get_queryset(self):
        qs = UserKYC.objects.select_related("user").all()

        status_in = (self.request.query_params.get("status") or "").strip().lower()
        user_q = (self.request.query_params.get("user") or "").strip()
        state_id = (self.request.query_params.get("state") or "").strip()
        pincode = (self.request.query_params.get("pincode") or "").strip()
        category = (self.request.query_params.get("category") or "").strip()
        audience = (self.request.query_params.get("audience") or "").strip().lower()
        date_from = (self.request.query_params.get("date_from") or "").strip()
        date_to = (self.request.query_params.get("date_to") or "").strip()
        ordering = (self.request.query_params.get("ordering") or "-updated_at").strip()

        if audience == "franchise":
            qs = qs.filter(user__category__in=[
                "agency_state_coordinator",
                "agency_state",
                "agency_district_coordinator",
                "agency_district",
                "agency_pincode_coordinator",
                "agency_pincode",
            ])
        elif audience == "consumer":
            qs = qs.filter(user__category="consumer")

        if category:
            c_key = str(category).strip().lower().replace(" ", "_").replace("-", "_")
            if "cordinator" in c_key:
                c_key = c_key.replace("cordinator", "coordinator")
            qs = qs.filter(user__category__iexact=c_key)

        if status_in == "pending":
            qs = qs.filter(verified=False)
        elif status_in == "verified":
            qs = qs.filter(verified=True)
        elif status_in == "submitted":
            qs = qs.filter(
                (Q(bank_name__isnull=False) & ~Q(bank_name=""))
                | (Q(bank_account_number__isnull=False) & ~Q(bank_account_number=""))
                | (Q(ifsc_code__isnull=False) & ~Q(ifsc_code=""))
            )

        if user_q:
            if user_q.isdigit():
                qs = qs.filter(user_id=int(user_q))
            else:
                qs = qs.filter(
                    Q(user__username__icontains=user_q)
                    | Q(user__full_name__icontains=user_q)
                    | Q(user__phone__icontains=user_q)
                )

        if state_id and state_id.isdigit():
            qs = qs.filter(user__state_id=int(state_id))

        if pincode:
            qs = qs.filter(user__pincode__icontains=pincode)

        # date range on updated_at
        if date_from:
            qs = qs.filter(updated_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(updated_at__date__lte=date_to)

        if ordering:
            qs = qs.order_by(ordering)

        return qs


class AdminKYCVerifyView(APIView):
    """
    Verify a user's KYC (sets verified=True, stamps verified_by/verified_at).
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("kyc")]

    def patch(self, request, user_id: int):
        kyc = UserKYC.objects.select_related("user").filter(user_id=user_id).first()
        if not kyc:
            return Response({"detail": "KYC not found for user"}, status=404)
        if not kyc.verified:
            kyc.verified = True
            kyc.verified_by = request.user
            kyc.verified_at = timezone.now()
            kyc.save(update_fields=["verified", "verified_by", "verified_at", "updated_at"])
        data = AdminKYCSerializer(kyc).data
        return Response(data, status=200)


class AdminKYCRejectView(APIView):
    """
    Reject KYC (currently sets verified=False). For explicit rejection tracking,
    extend the model with a 'status' field in a future migration.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("kyc")]

    def patch(self, request, user_id: int):
        kyc = UserKYC.objects.select_related("user").filter(user_id=user_id).first()
        if not kyc:
            return Response({"detail": "KYC not found for user"}, status=404)
        # Set to not verified; keep verified_at as-is to denote no verification timestamp
        if kyc.verified:
            kyc.verified = False
            kyc.verified_by = request.user  # audit who decided
            kyc.save(update_fields=["verified", "verified_by", "updated_at"])
        data = AdminKYCSerializer(kyc).data
        return Response(data, status=200)


class AdminWithdrawalList(ListAPIView):
    """
    List Withdrawal Requests with filters.
    Filters:
      - status=pending|approved|rejected
      - user (id or username contains)
      - date_from, date_to (requested_at)
      - min_amount, max_amount
      - method=upi|bank
      - ordering (default -requested_at)
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("withdrawals")]
    serializer_class = AdminWithdrawalSerializer

    def get_queryset(self):
        qs = WithdrawalRequest.objects.select_related("user").all()

        status_in = (self.request.query_params.get("status") or "").strip().lower()
        user_q = (self.request.query_params.get("user") or "").strip()
        date_from = (self.request.query_params.get("date_from") or "").strip()
        date_to = (self.request.query_params.get("date_to") or "").strip()
        min_amount = (self.request.query_params.get("min_amount") or "").strip()
        max_amount = (self.request.query_params.get("max_amount") or "").strip()
        method = (self.request.query_params.get("method") or "").strip().lower()
        ordering = (self.request.query_params.get("ordering") or "-requested_at").strip()

        if status_in in ("pending", "approved", "rejected"):
            qs = qs.filter(status=status_in)

        if user_q:
            if user_q.isdigit():
                qs = qs.filter(Q(user_id=int(user_q)) | Q(user__username__icontains=user_q))
            else:
                qs = qs.filter(
                    Q(user__username__icontains=user_q)
                    | Q(user__full_name__icontains=user_q)
                    | Q(user__phone__icontains=user_q)
                )

        if date_from:
            qs = qs.filter(requested_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(requested_at__date__lte=date_to)

        try:
            if min_amount:
                from decimal import Decimal as D
                qs = qs.filter(amount__gte=D(min_amount))
            if max_amount:
                from decimal import Decimal as D
                qs = qs.filter(amount__lte=D(max_amount))
        except Exception:
            pass

        if method in ("upi", "bank"):
            qs = qs.filter(method=method)

        if ordering:
            qs = qs.order_by(ordering)

        return qs


class AdminWithdrawalApproveView(APIView):
    """
    Approve a pending withdrawal and debit user's wallet atomically.
    Body: { "payout_ref": "optional reference", "note": "optional" }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("withdrawals")]

    def patch(self, request, pk: int):
        obj = WithdrawalRequest.objects.select_related("user").filter(pk=pk).first()
        if not obj:
            return Response({"detail": "Withdrawal not found"}, status=404)
        payout_ref = (request.data or {}).get("payout_ref") or ""
        note = (request.data or {}).get("note") or ""
        try:
            # Attach extra note if provided
            if note:
                obj.note = (obj.note or "") + f"\nApproved Note: {note}"
                obj.save(update_fields=["note"])
            obj.approve(actor=request.user, payout_ref=payout_ref)
        except Exception as e:
            return Response({"detail": str(e)}, status=400)
        return Response(AdminWithdrawalSerializer(obj).data, status=200)

# Admin health ping for auth/namespace diagnostics
class AdminPingView(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request):
        user = request.user
        # Compute module access map (default-deny)
        modules = {}
        try:
            if getattr(user, "is_superuser", False):
                modules = {k: True for k in MODULE_KEYS}
            else:
                modules = {k: has_admin_module_access(user, k) for k in MODULE_KEYS}
        except Exception:
            modules = {k: False for k in MODULE_KEYS}
        return Response(
            {
                "ok": True,
                "user": getattr(user, "username", None),
                "id": getattr(user, "id", None),
                "is_staff": bool(getattr(user, "is_staff", False)),
                "is_superuser": bool(getattr(user, "is_superuser", False)),
                "modules": modules,
            },
            status=200,
        )


# ====================
# Admin Support Portal
# ====================

class AdminSupportTicketList(ListAPIView):
    """
    List support tickets with powerful filters:
      - status: open|in_progress|resolved|rejected|closed
      - type: KYC_REVERIFY|GENERAL
      - user: id or username/full_name/phone contains
      - search: subject/message contains
      - ordering: default -updated_at
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("support")]
    serializer_class = AdminSupportTicketSerializer

    def get_queryset(self):
        qs = SupportTicket.objects.select_related("user", "admin_assignee").all()
        status_in = (self.request.query_params.get("status") or "").strip().lower()
        type_in = (self.request.query_params.get("type") or "").strip().upper()
        user_q = (self.request.query_params.get("user") or "").strip()
        search = (self.request.query_params.get("search") or "").strip()
        ordering = (self.request.query_params.get("ordering") or "-updated_at").strip()

        if status_in in ("open", "in_progress", "resolved", "rejected", "closed"):
            qs = qs.filter(status=status_in)
        if type_in in ("KYC_REVERIFY", "GENERAL"):
            qs = qs.filter(type=type_in)
        if user_q:
            if user_q.isdigit():
                qs = qs.filter(Q(user_id=int(user_q)) | Q(user__username__icontains=user_q))
            else:
                qs = qs.filter(
                    Q(user__username__icontains=user_q)
                    | Q(user__full_name__icontains=user_q)
                    | Q(user__phone__icontains=user_q)
                )
        if search:
            qs = qs.filter(Q(subject__icontains=search) | Q(message__icontains=search))
        if ordering:
            qs = qs.order_by(ordering)
        return qs


class AdminSupportTicketUpdate(APIView):
    """
    Update ticket fields: status, admin_assignee, resolution_note.
    Body:
      {
        "status": "open|in_progress|resolved|rejected|closed",
        "admin_assignee": 123,   // user id of staff/admin
        "resolution_note": "text to append/replace"
      }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("support")]

    def patch(self, request, pk: int):
        ticket = SupportTicket.objects.select_related("user").filter(pk=pk).first()
        if not ticket:
            return Response({"detail": "Ticket not found"}, status=404)

        data = request.data or {}
        status_in = (data.get("status") or "").strip().lower()
        admin_assignee_id = data.get("admin_assignee")
        resolution_note = (data.get("resolution_note") or "").strip()

        if admin_assignee_id is not None:
            try:
                aid = int(admin_assignee_id)
                assignee = CustomUser.objects.filter(id=aid, is_staff=True).first() or CustomUser.objects.filter(id=aid, is_superuser=True).first()
                if not assignee:
                    return Response({"detail": "admin_assignee must be an admin/staff user id"}, status=400)
                ticket.admin_assignee = assignee
            except Exception:
                return Response({"detail": "admin_assignee must be integer id"}, status=400)

        if status_in in ("open", "in_progress", "resolved", "rejected", "closed"):
            ticket.status = status_in

        if resolution_note:
            # append note
            ticket.resolution_note = (ticket.resolution_note or "") + (("\n" if ticket.resolution_note else "") + resolution_note)

        ticket.save(update_fields=["admin_assignee", "status", "resolution_note", "updated_at"])
        return Response(AdminSupportTicketSerializer(ticket).data, status=200)


class AdminSupportTicketMessageCreate(APIView):
    """
    Post an admin message to a ticket thread. Also moves status to in_progress when currently open.
    Body: { "message": "..." }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("support")]

    def post(self, request, pk: int):
        ticket = SupportTicket.objects.select_related("user").filter(pk=pk).first()
        if not ticket:
            return Response({"detail": "Ticket not found"}, status=404)
        msg = (request.data or {}).get("message")
        if not msg or not str(msg).strip():
            return Response({"detail": "message is required"}, status=400)
        m = SupportTicketMessage.objects.create(ticket=ticket, author=request.user, message=str(msg).strip())
        if ticket.status == "open":
            ticket.status = "in_progress"
            ticket.save(update_fields=["status", "updated_at"])
        return Response(AdminSupportTicketMessageSerializer(m).data, status=201)


class AdminSupportTicketApproveKYC(APIView):
    """
    Approve KYC re-verification request:
      - Sets user's kyc.kyc_reopen_allowed = True
      - Moves ticket to resolved (unless explicitly set otherwise)
      - Appends resolution note if provided
    Body: { "note": "optional" }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("support")]

    def post(self, request, pk: int):
        ticket = SupportTicket.objects.select_related("user").filter(pk=pk).first()
        if not ticket:
            return Response({"detail": "Ticket not found"}, status=404)
        if ticket.type != "KYC_REVERIFY":
            return Response({"detail": "Not a KYC re-verification ticket"}, status=400)

        # Ensure KYC exists and enable reopen flag
        kyc, _ = UserKYC.objects.get_or_create(user=ticket.user)
        if not kyc.kyc_reopen_allowed:
            kyc.kyc_reopen_allowed = True
            kyc.save(update_fields=["kyc_reopen_allowed", "updated_at"])

        note = (request.data or {}).get("note") or ""
        if note:
            ticket.resolution_note = (ticket.resolution_note or "") + (("\n" if ticket.resolution_note else "") + str(note))

        # Auto-assign current admin if not set
        if not ticket.admin_assignee_id:
            ticket.admin_assignee = request.user

        ticket.status = "resolved"
        ticket.save(update_fields=["status", "resolution_note", "admin_assignee", "updated_at"])
        return Response(AdminSupportTicketSerializer(ticket).data, status=200)


class AdminMatrixProgressList(ListAPIView):
    """
    List matrix progress per user and pool.
    Filters:
      - pool=FIVE_150|THREE_150|THREE_50
      - user (id or username/full_name/phone contains)
      - state (id), pincode (contains)
      - ordering (default -updated_at)
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("autopool")]
    serializer_class = AdminMatrixProgressSerializer

    def get_queryset(self):
        qs = UserMatrixProgress.objects.select_related("user").all()

        pool = (self.request.query_params.get("pool") or "").strip().upper()
        user_q = (self.request.query_params.get("user") or "").strip()
        state_id = (self.request.query_params.get("state") or "").strip()
        pincode = (self.request.query_params.get("pincode") or "").strip()
        ordering = (self.request.query_params.get("ordering") or "-updated_at").strip()

        if pool in ("FIVE_150", "THREE_150", "THREE_50"):
            qs = qs.filter(pool_type=pool)

        if user_q:
            if user_q.isdigit():
                qs = qs.filter(Q(user_id=int(user_q)) | Q(user__username__icontains=user_q))
            else:
                qs = qs.filter(
                    Q(user__username__icontains=user_q)
                    | Q(user__full_name__icontains=user_q)
                    | Q(user__phone__icontains=user_q)
                )

        if state_id and state_id.isdigit():
            qs = qs.filter(user__state_id=int(state_id))
        if pincode:
            qs = qs.filter(user__pincode__icontains=pincode)

        if ordering:
            qs = qs.order_by(ordering)
        return qs


class AdminMatrixTree(APIView):
    """
    Returns sponsor-based downline tree for a root user, limited by max_depth.
    Query:
      - pool=FIVE_150|THREE_150|THREE_50 (affects default max_depth only)
      - root_user_id (int, required)
      - max_depth (optional override; default 6 for FIVE, 15 for THREE)
    Response:
      { id, username, full_name, level, children:[...] }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("autopool")]

    def get(self, request):
        try:
            root_id = int(request.query_params.get("root_user_id") or "0")
        except Exception:
            return Response({"detail": "root_user_id must be integer"}, status=400)
        if root_id <= 0:
            return Response({"detail": "root_user_id is required"}, status=400)

        pool = (request.query_params.get("pool") or "").strip().upper()
        default_depth = 6 if pool == "FIVE_150" else 15
        try:
            max_depth = int(request.query_params.get("max_depth") or default_depth)
        except Exception:
            max_depth = default_depth
        max_depth = max(1, min(max_depth, 20))  # hard safety cap

        # Prefetch children by registered_by in batches to reduce queries
        def build_node(user, level, visited=None):
            if visited is None:
                visited = set()
            if getattr(user, "id", None) in visited:
                return None
            visited.add(user.id)
            node = {
                "id": user.id,
                "username": user.username,
                "full_name": user.full_name,
                "level": level,
                "children": [],
            }
            if level >= max_depth:
                return node
            # Build robust sponsor tokens including dashless variant of prefixed_id
            tokens = set()
            try:
                pid = (getattr(user, "prefixed_id", "") or "").strip()
                if pid:
                    tokens.add(pid)
                    tokens.add(pid.replace("-", ""))
                uname = (getattr(user, "username", "") or "").strip()
                if uname:
                    tokens.add(uname)
                uid = (getattr(user, "unique_id", "") or "").strip()
                if uid:
                    tokens.add(uid)
                phone_digits = "".join(ch for ch in str(getattr(user, "phone", "") or "") if ch.isdigit())
                if phone_digits:
                    tokens.add(phone_digits)
            except Exception:
                pass

            # Initial candidate query by registered_by or sponsor token match
            children_q = Q(registered_by_id=user.id)
            for t in tokens:
                children_q = children_q | Q(sponsor_id__iexact=t)

            candidates = list(
                CustomUser.objects.filter(children_q, category="consumer")
                .exclude(id=user.id)
                .only("id", "username", "full_name", "registered_by_id", "sponsor_id")
                .order_by("id")
                .distinct()
            )

            # Verify each sponsor_id token resolves back to the current user before accepting
            def _owner_id_by_token(token: str):
                if not token:
                    return None
                q = Q(prefixed_id__iexact=token) | Q(username__iexact=token) | Q(unique_id__iexact=token)
                t_no_dash = "".join(ch for ch in str(token) if ch.isalnum())
                if t_no_dash and t_no_dash != token:
                    q = q | Q(prefixed_id__iexact=t_no_dash)
                digits = "".join(ch for ch in str(token) if ch.isdigit())
                if digits:
                    q = q | Q(phone__iexact=digits) | Q(username__iexact=digits)
                u2 = CustomUser.objects.filter(q).only("id").first()
                return getattr(u2, "id", None)

            children = []
            for c in candidates:
                if getattr(c, "registered_by_id", None) == user.id:
                    children.append(c)
                else:
                    sid = (getattr(c, "sponsor_id", "") or "").strip()
                    if _owner_id_by_token(sid) == user.id:
                        children.append(c)

            for c in children:
                cn = build_node(c, level + 1, visited)
                if cn:
                    node["children"].append(cn)
            return node

        root = CustomUser.objects.filter(id=root_id).first()
        if not root:
            return Response({"detail": "root user not found"}, status=404)
        tree = build_node(root, 1)
        return Response(tree, status=200)


class AdminMatrix5Tree(APIView):
    """
    Entry-based matrix tree for AutoPoolAccount.

    Query params:
      - pool: FIVE_150 | THREE_150 | THREE_50 (default FIVE_150)
      - start_entry_id: int (optional) -> start BFS from this AutoPoolAccount.id
      - display_user_id: int (optional) -> if provided, start at this user's earliest entry in the pool.
          Special rule: if display_user_id == 32 and the user has no entry, fall back to sentinel root.
      - root_user_id: int (optional) -> alias for display_user_id (backward compatibility with older clients)
      - identifier: str (optional) -> resolve to a user (id/username/email/unique_id/phone/sponsor_id) and start at their earliest entry
      - max_depth: int (optional) -> defaults to configured levels for the pool; capped to configured levels and 20.

    Response (entry-based):
      {
        "account_id": <int>,           # AutoPoolAccount.id
        "owner_id": <int>,
        "username": <str>,
        "level": <int>,                # entry.level
        "position": <int|null>,        # sibling position under parent
        "status": "ACTIVE",
        "children": [ ...same shape... ]
      }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("autopool")]

    def get(self, request):
        # Params
        pool = (request.query_params.get("pool") or "FIVE_150").strip().upper()
        if pool not in ("FIVE_150", "THREE_150", "THREE_50"):
            pool = "FIVE_150"

        try:
            start_entry_id = int(request.query_params.get("start_entry_id") or "0")
        except Exception:
            start_entry_id = 0
        try:
            display_user_id = int(request.query_params.get("display_user_id") or "0")
        except Exception:
            display_user_id = 0
        # Backward-compat: accept root_user_id as an alias for display_user_id (older clients)
        if display_user_id <= 0:
            try:
                display_user_id = int(request.query_params.get("root_user_id") or "0")
            except Exception:
                display_user_id = display_user_id

        # Optional identifier support (admin search bar): resolve to user then use earliest entry
        ident = (request.query_params.get("identifier") or "").strip()
        if display_user_id <= 0 and ident:
            user = None
            # Prefer exact prefixed_id (sponsor code)
            try:
                user = CustomUser.objects.filter(prefixed_id__iexact=ident).first() or None
            except Exception:
                user = None
            # Numeric id path
            if not user:
                digits = "".join(ch for ch in ident if ch.isdigit())
                try:
                    if digits and digits == ident and digits.isdigit():
                        user = CustomUser.objects.filter(id=int(digits)).first()
                except Exception:
                    user = user
            # Username/email/unique_id and phone digits fallback
            if not user:
                try:
                    q = (Q(username__iexact=ident) | Q(email__iexact=ident) | Q(unique_id__iexact=ident))
                    digits = "".join(ch for ch in ident if ch.isdigit())
                    if digits:
                        q = q | Q(phone__iexact=digits) | Q(username__iexact=digits)
                    user = CustomUser.objects.filter(q).first()
                except Exception:
                    user = user
            if user:
                try:
                    display_user_id = int(getattr(user, "id", 0) or 0)
                except Exception:
                    display_user_id = 0

        # Configured depth
        cfg = CommissionConfig.get_solo()
        default_levels = cfg.get_matrix_five_levels() if pool == "FIVE_150" else cfg.get_matrix_three_levels()
        try:
            max_depth = int(request.query_params.get("max_depth") or default_levels)
        except Exception:
            max_depth = default_levels
        max_depth = max(1, min(int(max_depth), int(default_levels), 20))

        fanout = 5 if pool == "FIVE_150" else 3

        # Resolve sentinel to guarantee a structural root exists
        try:
            from business.services.placement import _ensure_sentinel_root
            sentinel = _ensure_sentinel_root(pool)
        except Exception:
            sentinel = AutoPoolAccount.objects.filter(pool_type=pool, parent_account__isnull=True).order_by("id").first()

        # Resolve special head/root consumer id (defaults to 32 when unavailable)
        try:
            from business.models import RootConsumerConfig
            _rc = RootConsumerConfig.get_solo()
            _ru = _rc.get_root_user()
            head_user_id = int(getattr(_ru, "id", 0) or 0) or 32
        except Exception:
            head_user_id = 32

        # Determine start account id:
        root_acc = None
        if start_entry_id > 0:
            root_acc = (
                AutoPoolAccount.objects.select_related("owner", "parent_account")
                .filter(id=start_entry_id, pool_type=pool, status="ACTIVE")
                .first()
            )
        if not root_acc and display_user_id > 0:
            # Prefer earliest ACTIVE entry for the requested user; fallback to sentinel only if none exists
            root_acc = (
                AutoPoolAccount.objects.select_related("owner")
                .filter(owner_id=display_user_id, pool_type=pool, status="ACTIVE")
                .order_by("id")
                .first()
            )
            if not root_acc and display_user_id == head_user_id and sentinel and getattr(sentinel, "pool_type", None) == pool:
                root_acc = AutoPoolAccount.objects.select_related("owner").filter(id=sentinel.id).first()
        if not root_acc:
            # Fallback to sentinel for this pool
            if sentinel and getattr(sentinel, "pool_type", None) == pool:
                root_acc = AutoPoolAccount.objects.select_related("owner").filter(id=sentinel.id).first()

        if not root_acc:
            return Response({"detail": "No matrix root available for the requested pool."}, status=404)

        # Disable special compression; always show actual L1 children including head-owned
        head_hide_self = False

        # BFS over AutoPoolAccount graph, width-before-depth, per-parent fanout cap
        def serialize_node(acc, rel_level):
            return {
                "account_id": acc.id,
                "owner_id": getattr(acc.owner, "id", None),
                "username": getattr(acc.owner, "username", None),
                "username_key": getattr(acc, "username_key", None),
                "level": int(rel_level),  # relative to requested root (root=1)
                "abs_level": int(getattr(acc, "level", 0) or 0),  # absolute persisted level
                "position": getattr(acc, "position", None),
                "status": getattr(acc, "status", "ACTIVE"),
                "team_count": 0,  # to be annotated after BFS as number of descendant entries
                "children": [],
            }

        root = serialize_node(root_acc, 1)
        # Expose pool fanout for client consumers (fixed width: 5 for FIVE_150, 3 for THREE_x)
        try:
            root["fanout"] = int(fanout)
        except Exception:
            pass

        nodes_by_account = {int(root_acc.id): root}
        current_parent_ids = [int(root_acc.id)]
        rel_levels = {int(root_acc.id): 1}
        levels_used = 1  # count root as level 1 for response budget
        # Track head-owned L1 account ids (used to compress display under sentinel)
        head_l1_ids: set[int] = set()

        while current_parent_ids and levels_used < max_depth:
            # Fetch children for all current parents in a single query
            qs_rows = (
                AutoPoolAccount.objects.select_related("owner")
                .filter(
                    pool_type=pool,
                    status="ACTIVE",
                    parent_account_id__in=current_parent_ids,
                )
            )
            rows = None
            # Special: when starting at sentinel for head user, compress out head-owned immediate children
            if head_hide_self and levels_used == 1 and len(current_parent_ids) == 1 and int(current_parent_ids[0]) == int(getattr(root_acc, "id", 0) or 0):
                # Build ordered L1 under sentinel without filtering head, then expand head-owned nodes into their children
                l1 = list(qs_rows.order_by("parent_account_id", "position", "id"))
                # Cache the head-owned L1 ids so we can attach their children directly under root for display
                head_l1_ids = {int(getattr(r, "id", 0) or 0) for r in l1 if getattr(r, "owner_id", None) == head_user_id}
                expanded = []
                for r in l1:
                    if getattr(r, "owner_id", None) == head_user_id:
                        # Pull this node's ACTIVE children (left-to-right)
                        gc = list(
                            AutoPoolAccount.objects.select_related("owner")
                            .filter(pool_type=pool, status="ACTIVE", parent_account_id=int(getattr(r, "id", 0) or 0))
                            .order_by("position", "id")
                        )
                        expanded.extend(gc)
                    else:
                        expanded.append(r)
                rows = expanded
            else:
                if head_hide_self:
                    qs_rows = qs_rows.exclude(owner_id=head_user_id)
                rows = list(qs_rows.order_by("parent_account_id", "position", "id"))
            if not rows:
                break

            # Per-parent child cap (fanout)
            counts = {}
            next_parent_ids = []
            # Determine if we're in the first BFS level under the requested root (sentinel)
            first_level_special = bool(head_hide_self and levels_used == 1 and len(current_parent_ids) == 1 and int(current_parent_ids[0]) == int(getattr(root_acc, "id", 0) or 0))
            for acc in rows:
                pid_raw = getattr(acc, "parent_account_id", None)
                # When compressing head-owned L1, attach their children directly under the root for display
                if first_level_special and pid_raw is not None and int(pid_raw) in head_l1_ids:
                    pid_eff = int(getattr(root_acc, "id", 0) or 0)
                else:
                    pid_eff = int(pid_raw) if pid_raw is not None else None
                if pid_eff is None or pid_eff not in nodes_by_account:
                    # parent may have been pruned; skip
                    continue
                used = counts.get(pid_eff, 0)
                if used >= fanout:
                    continue
                parent_node = nodes_by_account[pid_eff]
                parent_rel = int(rel_levels.get(pid_eff, levels_used))
                child_rel = parent_rel + 1
                child_node = serialize_node(acc, child_rel)
                parent_node["children"].append(child_node)
                nodes_by_account[int(acc.id)] = child_node
                rel_levels[int(acc.id)] = child_rel
                counts[pid_eff] = used + 1
                next_parent_ids.append(int(acc.id))

            if not next_parent_ids:
                break
            current_parent_ids = next_parent_ids
            levels_used += 1

        # Post-process: annotate team_count per entry as total number of descendant entries
        def _annotate_team_count(node: dict) -> int:
            try:
                kids = node.get("children") or []
            except Exception:
                kids = []
            total = 0
            for ch in kids:
                total += 1 + _annotate_team_count(ch)
            try:
                node["team_count"] = int(total)
            except Exception:
                node["team_count"] = 0
            return total

        try:
            _annotate_team_count(root)
        except Exception:
            pass

        # Override root team_count with actual DB total (BFS is depth-limited so
        # the recursive count only reflects visible nodes, not the real total).
        try:
            real_total = AutoPoolAccount.objects.filter(
                pool_type=pool, status="ACTIVE", parent_account__isnull=False,
            ).count()
            root["team_count"] = int(real_total)
        except Exception:
            pass

        return Response(root, status=200)


class AdminAutopoolSummary(APIView):
    """
    Summary for Auto Commission Pool and Matrix progress.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("autopool")]

    def get(self, request):
        # Progress aggregates
        agg = (
            UserMatrixProgress.objects.values("pool_type")
            .annotate(
                users=Count("id"),
                total_earned=Sum("total_earned"),
            )
            .order_by()
        )
        progress = {row["pool_type"]: {
            "users": row["users"],
            "total_earned": float(row["total_earned"] or 0),
        } for row in agg}

        # Active accounts by pool (best-effort)
        accounts = (
            AutoPoolAccount.objects.values("pool_type", "status")
            .annotate(c=Count("id"))
            .order_by()
        )
        acc_map = {}
        for row in accounts:
            pool = row["pool_type"]
            acc_map.setdefault(pool, {"ACTIVE": 0, "PENDING": 0, "CLOSED": 0})
            acc_map[pool][row["status"]] = row["c"]

        # Compose
        data = {
            "progress": progress,
            "accounts": acc_map,
        }
        return Response(data, status=200)


class AdminAutopoolTransactionList(ListAPIView):
    """
    List recent Auto Pool and related commission transactions in a table-friendly format.
    Columns: TR (prefixed_id), Username, Sponsor ID, Amount (gross), Type, Main Wallet, Withdrawable (net), Date.
    Filters:
      - types: comma-separated WalletTransaction.type values (optional, defaults to commission/autopool related)
      - user: user id or username/full_name/phone contains
      - date_from, date_to: created_at (date)
      - ordering: default -created_at
      - page, page_size: pagination
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("autopool")]
    serializer_class = AdminAutopoolTxnSerializer

    def get_queryset(self):
        qs = WalletTransaction.objects.select_related("user").all()

        # Default transaction types focused on commissions/autopool flows
        default_types = (
            "COMMISSION_CREDIT",
            "DIRECT_REF_BONUS",
            "LEVEL_BONUS",
            "AUTOPOOL_BONUS_FIVE",
            "AUTOPOOL_BONUS_THREE",
            "FRANCHISE_INCOME",
            "GLOBAL_ROYALTY",
            # Include coupon redeem credit since it may be a source for pool entries
            "REDEEM_ECOUPON_CREDIT",
        )
        types_param = (self.request.query_params.get("types") or "").strip()
        if types_param:
            types_list = [t.strip() for t in types_param.split(",") if t.strip()]
            if types_list:
                qs = qs.filter(type__in=types_list)
        else:
            qs = qs.filter(type__in=default_types)

        user_q = (self.request.query_params.get("user") or "").strip()
        if user_q:
            if user_q.isdigit():
                qs = qs.filter(Q(user_id=int(user_q)) | Q(user__username__icontains=user_q))
            else:
                qs = qs.filter(
                    Q(user__username__icontains=user_q)
                    | Q(user__full_name__icontains=user_q)
                    | Q(user__phone__icontains=user_q)
                )

        date_from = (self.request.query_params.get("date_from") or "").strip()
        date_to = (self.request.query_params.get("date_to") or "").strip()
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        ordering = (self.request.query_params.get("ordering") or "-created_at").strip()
        if ordering:
            qs = qs.order_by(ordering)
        return qs

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        try:
            page = int(request.query_params.get("page") or 1)
        except Exception:
            page = 1
        try:
            page_size = int(request.query_params.get("page_size") or 50)
        except Exception:
            page_size = 50
        page = max(1, page)
        page_size = max(1, min(page_size, 200))

        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = self.get_serializer(qs[start:end], many=True, context=self.get_serializer_context())
        return Response({"count": total, "results": serializer.data}, status=200)


class AdminMatrixAccountsList(APIView):
    """
    Admin: List AutoPoolAccount entries (matrix accounts) with filters.
    Query params:
      - pool: FIVE_150 | THREE_150 | THREE_50
      - user: id or username/full_name/phone contains
      - source_type: string (e.g., ECOUPON)
      - source_id: string (e.g., coupon id)
      - date_from, date_to: created_at (date)
      - ordering: -created_at (default) | created_at | level | -level | position | -position
      - page (default 1), page_size (default 25, max 200)
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("autopool")]

    def get(self, request):
        try:
            page = int(request.query_params.get("page") or 1)
        except Exception:
            page = 1
        try:
            page_size = int(request.query_params.get("page_size") or 25)
        except Exception:
            page_size = 25
        page = max(1, page)
        page_size = max(1, min(page_size, 200))

        pool = (request.query_params.get("pool") or "").strip().upper()
        user_q = (request.query_params.get("user") or "").strip()
        source_type = (request.query_params.get("source_type") or "").strip()
        source_id = (request.query_params.get("source_id") or "").strip()
        date_from = (request.query_params.get("date_from") or "").strip()
        date_to = (request.query_params.get("date_to") or "").strip()
        ordering = (request.query_params.get("ordering") or "-created_at").strip()

        qs = (AutoPoolAccount.objects
              .select_related("owner", "parent_account", "parent_account__owner")
              .all())

        if pool in ("FIVE_150", "THREE_150", "THREE_50"):
            qs = qs.filter(pool_type=pool)

        if user_q:
            if user_q.isdigit():
                qs = qs.filter(Q(owner_id=int(user_q)) | Q(owner__username__icontains=user_q))
            else:
                qs = qs.filter(
                    Q(owner__username__icontains=user_q)
                    | Q(owner__full_name__icontains=user_q)
                    | Q(owner__phone__icontains=user_q)
                )

        if source_type:
            qs = qs.filter(source_type=source_type)
        if source_id:
            qs = qs.filter(source_id=source_id)

        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        allowed_order = {"created_at", "-created_at", "level", "-level", "position", "-position", "id", "-id"}
        if ordering not in allowed_order:
            ordering = "-created_at"
        qs = qs.order_by(ordering)

        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        items = qs[start:end]

        results = []
        for a in items:
            parent_owner = None
            try:
                po = getattr(a.parent_account, "owner", None)
                parent_owner = {
                    "id": getattr(po, "id", None),
                    "username": getattr(po, "username", None),
                } if po else None
            except Exception:
                parent_owner = None

            results.append({
                "id": a.id,
                "pool_type": a.pool_type,
                "owner_id": a.owner_id,
                "username": getattr(a.owner, "username", None),
                "parent_account_id": a.parent_account_id,
                "parent_owner": parent_owner,
                "level": a.level,
                "position": a.position,
                "source_type": a.source_type or "",
                "source_id": a.source_id or "",
                "created_at": a.created_at,
            })

        return Response({"count": total, "page": page, "page_size": page_size, "results": results}, status=200)


class AdminMatrixAccountStats(APIView):
    """
    Admin: Level-wise and total commission stats for a matrix account (by account_id)
           or by (pool + source_type + source_id) tuple.
    Query params:
      - account_id: int
        OR
      - pool: FIVE_150 | THREE_150 | THREE_50
      - source_type: string
      - source_id: string
    Response:
      {
        "pool": "...",
        "source_type": "...",
        "source_id": "...",
        "tx_count": 10,
        "total_amount": "123.45",
        "per_level": {
          "1": { "count": 5, "amount": "50.00" },
          ...
        },
        "account": { ... }  // when account_id provided
      }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("autopool")]

    def get(self, request):
        account_id = (request.query_params.get("account_id") or "").strip()
        pool = (request.query_params.get("pool") or "").strip().upper()
        src_type = (request.query_params.get("source_type") or "").strip()
        src_id = (request.query_params.get("source_id") or "").strip()

        account_info = None
        if account_id:
            try:
                acc = AutoPoolAccount.objects.select_related("owner", "parent_account", "parent_account__owner").get(pk=int(account_id))
            except Exception:
                return Response({"detail": "account not found"}, status=404)
            pool = acc.pool_type
            src_type = acc.source_type or ""
            src_id = acc.source_id or ""
            account_info = {
                "id": acc.id,
                "pool_type": acc.pool_type,
                "owner_id": acc.owner_id,
                "username": getattr(acc.owner, "username", None),
                "parent_account_id": acc.parent_account_id,
                "level": acc.level,
                "position": acc.position,
                "source_type": src_type,
                "source_id": src_id,
                "created_at": acc.created_at,
            }

        if not (pool and src_type and src_id):
            return Response({"detail": "Provide account_id or (pool + source_type + source_id)."}, status=400)

        if pool == "FIVE_150":
            tx_types = ("AUTOPOOL_BONUS_FIVE",)
        elif pool in ("THREE_150", "THREE_50"):
            tx_types = ("AUTOPOOL_BONUS_THREE",)
        else:
            return Response({"detail": "Invalid pool"}, status=400)

        qs = (WalletTransaction.objects
              .select_related("user")
              .filter(type__in=tx_types, source_type=src_type, source_id=str(src_id))
              .order_by("id"))

        from decimal import Decimal as D
        total_amount = D("0.00")
        per_level = {}
        tx_count = 0

        for t in qs:
            tx_count += 1
            amt = D(str(getattr(t, "amount", "0") or "0"))
            total_amount += amt
            meta = getattr(t, "meta", {}) or {}
            try:
                lvl = int(meta.get("level_index") or 0)
            except Exception:
                lvl = 0
            key = str(lvl)
            row = per_level.get(key) or {"count": 0, "amount": D("0.00")}
            row["count"] += 1
            row["amount"] = row["amount"] + amt
            per_level[key] = row

        # Format amounts to strings
        per_level_out = {}
        for k, v in per_level.items():
            per_level_out[k] = {"count": v["count"], "amount": f'{D(v["amount"]):.2f}'}

        payload = {
            "pool": pool,
            "source_type": src_type,
            "source_id": str(src_id),
            "tx_count": tx_count,
            "total_amount": f'{D(total_amount):.2f}',
            "per_level": dict(sorted(per_level_out.items(), key=lambda x: int(x[0]) if str(x[0]).isdigit() else 9999)),
        }
        if account_info:
            payload["account"] = account_info
        return Response(payload, status=200)


class AdminWithdrawalRejectView(APIView):
    """
    Reject a pending withdrawal without wallet mutation.
    Body: { "reason": "required/optional" }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("withdrawals")]

    def patch(self, request, pk: int):
        obj = WithdrawalRequest.objects.select_related("user").filter(pk=pk).first()
        if not obj:
            return Response({"detail": "Withdrawal not found"}, status=404)
        reason = (request.data or {}).get("reason") or ""
        try:
            obj.reject(actor=request.user, reason=reason)
        except Exception as e:
            return Response({"detail": str(e)}, status=400)
        return Response(AdminWithdrawalSerializer(obj).data, status=200)


class AdminWithdrawalDistributionPreviewView(APIView):
    """
    Preview Direct Refer Withdraw Commission distribution for a given user and amount.
    Query params:
      - user_id: int (preferred)
      - user or username: str (optional alternative)
      - amount: number (required)
    Response: same schema as business.services.withdrawals.compute_withdraw_distribution(...)
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("withdrawals")]

    def get(self, request):
        user_id = (request.query_params.get("user_id") or "").strip()
        user_q = (request.query_params.get("user") or request.query_params.get("username") or "").strip()
        amount = request.query_params.get("amount")
        if amount is None or str(amount).strip() == "":
            return Response({"detail": "amount is required"}, status=400)

        user = None
        if user_id and str(user_id).isdigit():
            user = CustomUser.objects.filter(id=int(user_id)).first()

        if not user and user_q:
            # Try by username, then prefixed sponsor id, then phone digits
            u = CustomUser.objects.filter(username__iexact=user_q).first()
            if not u:
                u = CustomUser.objects.filter(prefixed_id__iexact=user_q).first()
            if not u:
                digits = "".join(ch for ch in str(user_q) if ch.isdigit())
                if digits:
                    u = CustomUser.objects.filter(phone__iexact=digits).first()
            user = u

        if not user:
            return Response({"detail": "user not found"}, status=404)

        try:
            from business.services.withdrawals import compute_withdraw_distribution
            data = compute_withdraw_distribution(user, amount)
            return Response(data, status=200)
        except Exception as e:
            return Response({"detail": str(e)}, status=400)

# ================================
# Master Level Commission (L0..L5)
# ================================
class AdminLevelCommissionView(APIView):
    """
    GET: Return current Direct and L1..L5 fixed level commissions (rupees) from CommissionConfig.referral_join_fixed_json
         { direct, l1, l2, l3, l4, l5, updated_at }
    PATCH: Update any subset of the above keys with numeric values
         Body e.g. { "direct": 15, "l1": 2, "l2": 1, "l3": 1, "l4": 0.5, "l5": 0.5 }
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("commissions")]

    def _serialize(self, cfg):
        fixed = dict(getattr(cfg, "referral_join_fixed_json", {}) or {})
        def _f(k, d):
            from decimal import Decimal as D
            try:
                return float(D(str(fixed.get(k, d))))
            except Exception:
                return float(d)
        payload = {
            "direct": _f("direct", 15),
            "l1": _f("l1", 2),
            "l2": _f("l2", 1),
            "l3": _f("l3", 1),
            "l4": _f("l4", 0.5),
            "l5": _f("l5", 0.5),
            "updated_at": getattr(cfg, "updated_at", None),
        }
        return payload

    def get(self, request):
        cfg = CommissionConfig.get_solo()
        return Response(self._serialize(cfg), status=200)

    def patch(self, request):
        if not getattr(request.user, "is_superuser", False):
            return Response({"detail": "Only superuser can modify commission config."}, status=403)
        from decimal import Decimal as D
        cfg = CommissionConfig.get_solo()
        fixed = dict(getattr(cfg, "referral_join_fixed_json", {}) or {})
        data = request.data or {}
        # Accept only known keys
        for k in ("direct", "l1", "l2", "l3", "l4", "l5"):
            if k in data:
                try:
                    v = D(str(data.get(k)))
                    if v < 0:
                        return Response({"detail": f"{k} must be >= 0"}, status=400)
                    # Quantize to 2 decimals string to avoid float drift in JSON
                    fixed[k] = float(v.quantize(D("0.01")))
                except Exception:
                    return Response({"detail": f"{k} must be a number"}, status=400)
        cfg.referral_join_fixed_json = fixed
        try:
            cfg.save(update_fields=["referral_join_fixed_json", "updated_at"])
        except Exception:
            cfg.save()
        return Response(self._serialize(cfg), status=200)


class AdminLevelCommissionSeedView(APIView):
    """
    POST: Reset Direct and L1..L5 to defaults {15, 2, 1, 1, 0.5, 0.5}
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("commissions")]

    def post(self, request):
        if not getattr(request.user, "is_superuser", False):
            return Response({"detail": "Only superuser can modify commission config."}, status=403)
        cfg = CommissionConfig.get_solo()
        cfg.referral_join_fixed_json = {
            "direct": 15,
            "l1": 2,
            "l2": 1,
            "l3": 1,
            "l4": 0.5,
            "l5": 0.5,
        }
        try:
            cfg.save(update_fields=["referral_join_fixed_json", "updated_at"])
        except Exception:
            cfg.save()
        return Response({"ok": True, "defaults": cfg.referral_join_fixed_json}, status=200)


class AdminMatrixCommissionConfig(APIView):
    """
    GET:
      - When no ?product= is provided: returns global typed fields on CommissionConfig
        {
          five_matrix_levels, five_matrix_amounts_json, five_matrix_percents_json,
          three_matrix_levels, three_matrix_amounts_json, three_matrix_percents_json,
          updated_at
        }
      - When ?product=coupon150|rs759|750 is provided: returns per‑product overrides under
        master_commission_json.consumer_matrix_5/3[productKey] with fallback to globals.
        productKey: "150" for coupon150, "759" for rs759, "750" for 750.

    PATCH:
      - No product: same as before, updates typed fields via AdminAutopoolConfigSerializer.
      - With product: updates master_commission_json.consumer_matrix_5/3[productKey].
        Accepts any subset: five_matrix_levels, five_matrix_amounts_json, five_matrix_percents_json,
                            three_matrix_levels, three_matrix_amounts_json, three_matrix_percents_json.
        Arrays are coerced to 2 decimals; when levels provided, arrays are padded/truncated to match.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("commissions")]

    def _product_key(self, request):
        p = (request.query_params.get("product") or "").strip().lower()
        if p in ("coupon150", "coupon_150", "150", "prime150", "prime_150"):
            return "150"
        if p in ("rs759", "759"):
            return "759"
        if p in ("750", "prime750", "prime_750", "rs750"):
            return "750"
        return None

    def _to_num_list(self, arr):
        out = []
        try:
            for x in (arr or []):
                from decimal import Decimal as D
                out.append(float(D(str(x)).quantize(D("0.01"))))
        except Exception:
            pass
        return out

    def get(self, request):
        cfg = CommissionConfig.get_solo()
        pk = self._product_key(request)
        if not pk:
            ser = AdminAutopoolConfigSerializer(cfg)
            return Response(ser.data, status=200)

        master = dict(getattr(cfg, "master_commission_json", {}) or {})
        cm5_all = dict(master.get("consumer_matrix_5", {}) or {})
        cm3_all = dict(master.get("consumer_matrix_3", {}) or {})
        cm5 = dict(cm5_all.get(pk, {}) or {})
        cm3 = dict(cm3_all.get(pk, {}) or {})

        five_levels = int(cm5.get("levels") or getattr(cfg, "five_matrix_levels", 6) or 6)
        three_levels = int(cm3.get("levels") or getattr(cfg, "three_matrix_levels", 15) or 15)

        resp = {
            "five_matrix_levels": five_levels,
            "five_matrix_amounts_json": self._to_num_list(cm5.get("fixed_amounts") or getattr(cfg, "five_matrix_amounts_json", []) or []),
            "five_matrix_percents_json": self._to_num_list(cm5.get("percents") or getattr(cfg, "five_matrix_percents_json", []) or []),
            "three_matrix_levels": three_levels,
            "three_matrix_amounts_json": self._to_num_list(cm3.get("fixed_amounts") or getattr(cfg, "three_matrix_amounts_json", []) or []),
            "three_matrix_percents_json": self._to_num_list(cm3.get("percents") or getattr(cfg, "three_matrix_percents_json", []) or []),
            "updated_at": getattr(cfg, "updated_at", None),
        }
        return Response(resp, status=200)

    def patch(self, request):
        if not getattr(request.user, "is_superuser", False):
            return Response({"detail": "Only superuser can modify commission config."}, status=403)
        cfg = CommissionConfig.get_solo()
        pk = self._product_key(request)
        if not pk:
            ser = AdminAutopoolConfigSerializer(cfg, data=request.data, partial=True)
            if ser.is_valid():
                obj = ser.save()
                return Response(AdminAutopoolConfigSerializer(obj).data, status=200)
            return Response(ser.errors, status=400)

        data = request.data or {}
        master = dict(getattr(cfg, "master_commission_json", {}) or {})
        cm5_all = dict(master.get("consumer_matrix_5", {}) or {})
        cm3_all = dict(master.get("consumer_matrix_3", {}) or {})
        cm5 = dict(cm5_all.get(pk, {}) or {})
        cm3 = dict(cm3_all.get(pk, {}) or {})

        from decimal import Decimal as D

        def coerce_int(v, default_val):
            try:
                iv = int(v)
                return iv if iv > 0 else int(default_val)
            except Exception:
                return int(default_val)

        # five levels
        if "five_matrix_levels" in data:
            cm5["levels"] = coerce_int(data.get("five_matrix_levels"), cm5.get("levels") or getattr(cfg, "five_matrix_levels", 6) or 6)
        # amounts/percents (coerce to 2 decimals)
        if "five_matrix_amounts_json" in data:
            cm5["fixed_amounts"] = self._to_num_list(data.get("five_matrix_amounts_json"))
        if "five_matrix_percents_json" in data:
            cm5["percents"] = self._to_num_list(data.get("five_matrix_percents_json"))

        # three levels
        if "three_matrix_levels" in data:
            cm3["levels"] = coerce_int(data.get("three_matrix_levels"), cm3.get("levels") or getattr(cfg, "three_matrix_levels", 15) or 15)
        if "three_matrix_amounts_json" in data:
            cm3["fixed_amounts"] = self._to_num_list(data.get("three_matrix_amounts_json"))
        if "three_matrix_percents_json" in data:
            cm3["percents"] = self._to_num_list(data.get("three_matrix_percents_json"))

        # Normalize lengths to levels if both present
        def normalize(block):
            levels = int(block.get("levels") or 0)
            if levels > 0:
                for key in ("fixed_amounts", "percents"):
                    arr = list(block.get(key) or [])
                    if arr:
                        if len(arr) > levels:
                            block[key] = arr[:levels]
                        elif len(arr) < levels:
                            block[key] = arr + [0.0] * (levels - len(arr))

        normalize(cm5)
        normalize(cm3)

        cm5_all[pk] = cm5
        cm3_all[pk] = cm3
        master["consumer_matrix_5"] = cm5_all
        master["consumer_matrix_3"] = cm3_all
        # Keep commissions block in sync with master keys so policy.enable_3/enable_5 reflect latest changes
        try:
            from business.services.commission_policy import CommissionPolicy
            master["commissions"] = CommissionPolicy._synth_from_master(master)
        except Exception:
            pass
        cfg.master_commission_json = master
        try:
            cfg.save(update_fields=["master_commission_json", "updated_at"])
        except Exception:
            cfg.save()
        # Return fresh product-specific payload
        request._request.GET._mutable = True if hasattr(request._request.GET, "_mutable") else False  # noop safety
        return self.get(request)


class AdminMasterCommissionConfig(APIView):
    """
    GET:
      - No product: return global master commission view (tax, withdrawal %, company user, upline %, global geo %)
      - With ?product=coupon150|rs759|750: additionally include per-product keys under master_commission_json:
        {
          direct_bonus: { sponsor, self },
          geo_mode: "fixed" | "percent" | "",
          geo_fixed: { role -> ₹ },
          geo: percent overrides for this product (fallback to global if missing)
        }
    PATCH:
      - No product: existing behavior (tax, withdrawal.sponsor_percent, upline %, global geo %)
      - With product: accept any subset of:
          { direct_bonus: { sponsor, self },
            geo_mode: "fixed" | "percent",
            geo_fixed: { sub_franchise, pincode, pincode_coord, district, district_coord, state, state_coord, employee, royalty },
            geo: { same keys as above }  # product-specific percent overrides
          }
        Global keys (tax, withdrawal, upline) are also allowed and update as usual.
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("commissions")]

    def _float(self, v):
        from decimal import Decimal as D
        try:
            return float(D(str(v)))
        except Exception:
            try:
                return float(v)
            except Exception:
                return 0.0

    def _product_key(self, request):
        p = (request.query_params.get("product") or "").strip().lower()
        if p in ("coupon150", "coupon_150", "150", "prime150", "prime_150"):
            return "150"
        if p in ("rs759", "759"):
            return "759"
        if p in ("750", "prime750", "prime_750", "rs750"):
            return "750"
        return None

    def get(self, request):
        cfg = CommissionConfig.get_solo()
        cu = cfg.get_company_user()
        upline_list = cfg.get_level_percents()
        try:
            upline = {
                "l1": self._float(upline_list[0]) if len(upline_list) > 0 else 0.0,
                "l2": self._float(upline_list[1]) if len(upline_list) > 1 else 0.0,
                "l3": self._float(upline_list[2]) if len(upline_list) > 2 else 0.0,
                "l4": self._float(upline_list[3]) if len(upline_list) > 3 else 0.0,
                "l5": self._float(upline_list[4]) if len(upline_list) > 4 else 0.0,
            }
        except Exception:
            upline = {"l1": 0, "l2": 0, "l3": 0, "l4": 0, "l5": 0}
        geo_raw = cfg.get_geo_percents()
        geo_global = {k: self._float(v) for k, v in (geo_raw.items() if isinstance(geo_raw, dict) else [])}

        payload = {
            "tax": {"percent": self._float(cfg.get_tax_percent())},
            "withdrawal": {"sponsor_percent": self._float(cfg.get_withdrawal_sponsor_percent())},
            "withdrawals_window": {
                "enabled": bool(getattr(cfg, "withdrawals_enabled", True)),
                "weekday": int(getattr(cfg, "withdrawals_weekday", 2) or 0),
                "start_time": (getattr(cfg, "withdrawals_start_time", None).strftime("%H:%M") if getattr(cfg, "withdrawals_start_time", None) else "00:00"),
                "end_time": (getattr(cfg, "withdrawals_end_time", None).strftime("%H:%M") if getattr(cfg, "withdrawals_end_time", None) else "23:59"),
            },
            "company_user": ({"id": getattr(cu, "id", None), "username": getattr(cu, "username", None)} if cu else None),
            "upline": upline,
            "geo": geo_global,
            "updated_at": getattr(cfg, "updated_at", None),
        }

        pk = self._product_key(request)
        if pk:
            master = dict(getattr(cfg, "master_commission_json", {}) or {})
            direct_all = dict(master.get("direct_bonus", {}) or {})
            geo_mode_all = dict(master.get("geo_mode", {}) or {})
            geo_fixed_all = dict(master.get("geo_fixed", {}) or {})
            geo_pct_all = dict(master.get("geo_percent", {}) or {})

            direct = dict(direct_all.get(pk, {}) or {})
            geo_mode = str(geo_mode_all.get(pk, "") or "")
            geo_fixed = dict(geo_fixed_all.get(pk, {}) or {})
            geo_pct = dict(geo_pct_all.get(pk, {}) or {})

            # Compose per-product view (geo percent per-product falls back to global)
            payload["direct_bonus"] = {
                "sponsor": self._float(direct.get("sponsor", 0)),
                "self": self._float(direct.get("self", 0)),
            }
            payload["geo_mode"] = geo_mode
            payload["geo_fixed"] = {k: self._float(v) for k, v in geo_fixed.items()} if geo_fixed else {}
            payload["geo"] = ({k: self._float(v) for k, v in geo_pct.items()} if geo_pct else payload["geo"])

            # Expose per-product base amount and (for 150) coupon activation count for admin UI
            try:
                products_all = dict(master.get("products", {}) or {})
                prod_row = dict(products_all.get(pk, {}) or {})
                if "base_amount" in prod_row:
                    payload["product_base_amount"] = self._float(prod_row.get("base_amount"))
            except Exception:
                pass
            if pk == "150":
                try:
                    comm = dict(master.get("commissions", {}) or {})
                    p150c = dict(comm.get("prime_150", {}) or {})
                    cnode = dict(p150c.get("coupons", {}) or {})
                    if "activation_count" in cnode:
                        try:
                            payload["coupon_activation_count"] = int(cnode.get("activation_count") or 0)
                        except Exception:
                            payload["coupon_activation_count"] = 0
                except Exception:
                    pass
                # Expose matrix open config for 150
                try:
                    products_all = dict(master.get("products", {}) or {})
                    row150 = dict(products_all.get("150", {}) or {})
                    mode150 = str(row150.get("matrix_open_mode", "FIRST_TIME_ONLY") or "FIRST_TIME_ONLY").strip().upper()
                    try:
                        count150 = int(row150.get("matrix_open_count", 1))
                    except Exception:
                        count150 = 1
                    payload["product_matrix_open_mode"] = mode150
                    payload["matrix_open_mode"] = mode150
                    payload["product_matrix_open_count"] = max(0, int(count150))
                    payload["matrix_open_count"] = max(0, int(count150))
                except Exception:
                    pass

            if pk == "750":
                try:
                    products_all = dict(master.get("products", {}) or {})
                    row750 = dict(products_all.get("750", {}) or {})
                    aoc = row750.get("activation_open_count", None)
                    if aoc is None:
                        # Backward compatible top-level map support
                        aoc_map = dict(master.get("activation_open_count", {}) or {})
                        aoc = aoc_map.get("750", None)
                    if aoc is not None:
                        try:
                            payload["activation_open_count"] = int(aoc)
                        except Exception:
                            payload["activation_open_count"] = 0
                    # Also expose matrix open config for 750
                    mode750 = str(row750.get("matrix_open_mode", "FIRST_TIME_ONLY") or "FIRST_TIME_ONLY").strip().upper()
                    try:
                        count750 = int(row750.get("matrix_open_count", row750.get("activation_open_count", 1)))
                    except Exception:
                        try:
                            count750 = int(row750.get("activation_open_count", 1))
                        except Exception:
                            count750 = 1
                    payload["product_matrix_open_mode"] = mode750
                    payload["matrix_open_mode"] = mode750
                    payload["product_matrix_open_count"] = max(0, int(count750))
                    payload["matrix_open_count"] = max(0, int(count750))
                except Exception:
                    pass

            # Monthly 759 product-specific config (first-month vs subsequent months, levels, agency toggle, base amount)
            if pk == "759":
                m759 = dict(master.get("monthly_759", {}) or {})
                def _f2(val, d=0):
                    try:
                        from decimal import Decimal as D
                        return float(D(str(val)))
                    except Exception:
                        try:
                            return float(val)
                        except Exception:
                            return float(d)
                levels = m759.get("levels_fixed") or []
                try:
                    lvls = [_f2(x, 0) for x in levels]
                except Exception:
                    lvls = []
                payload["monthly_759"] = {
                    "direct_first_month": _f2(m759.get("direct_first_month", 250)),
                    "direct_monthly": _f2(m759.get("direct_monthly", 50)),
                    "levels_fixed": lvls,
                    "agency_enabled": bool(m759.get("agency_enabled", True)),
                    "base_amount": _f2(m759.get("base_amount", 759)),
                    "matrix_open_mode": str(m759.get("matrix_open_mode", "FIRST_MONTH_ONLY") or "FIRST_MONTH_ONLY").strip().upper(),
                }

        # Include minimal commissions.prime_150.rewards.points_amount and commissions.prime_750 for UI consumers
        try:
            master2 = dict(getattr(cfg, "master_commission_json", {}) or {})
            comm_all2 = dict(master2.get("commissions", {}) or {})
            # prime_150 rewards.points_amount
            p150_2 = dict(comm_all2.get("prime_150", {}) or {})
            rewards_2 = dict(p150_2.get("rewards", {}) or {})
            # prime_750 (base_package + multiplier)
            p750_2 = dict(comm_all2.get("prime_750", {}) or {})
            base_pkg = str(p750_2.get("base_package", "prime_150") or "prime_150").strip().lower()
            try:
                mul_in = int(p750_2.get("multiplier", 1))
                if mul_in <= 0:
                    mul_in = 1
            except Exception:
                mul_in = 1
            payload.setdefault("commissions", {})
            payload["commissions"]["prime_150"] = {"rewards": {"points_amount": self._float(rewards_2.get("points_amount", 0))}}
            payload["commissions"]["prime_750"] = {"base_package": "prime_150", "multiplier": int(mul_in) if base_pkg == "prime_150" else 1}
        except Exception:
            pass
        return Response(payload, status=200)

    def patch(self, request):
        if not getattr(request.user, "is_superuser", False):
            return Response({"detail": "Only superuser can modify commission config."}, status=403)
        from decimal import Decimal as D
        data = request.data or {}
        cfg = CommissionConfig.get_solo()
        master = dict(getattr(cfg, "master_commission_json", {}) or {})
        pk = self._product_key(request)

        # Allow updating commissions.prime_150.rewards.points_amount directly from payload
        cm = data.get("commissions")
        if isinstance(cm, dict):
            # commissions.prime_150.rewards.points_amount
            try:
                p150_in = dict(cm.get("prime_150", {}) or {})
                r_in = p150_in.get("rewards")
                if isinstance(r_in, dict) and "points_amount" in r_in:
                    v = D(str(r_in.get("points_amount")))
                    if v < 0:
                        return Response({"detail": "commissions.prime_150.rewards.points_amount must be >= 0"}, status=400)
                    # Merge into master["commissions"] so CommissionPolicy._synth_from_master preserves it
                    comm = dict(master.get("commissions", {}) or {})
                    p150_cur = dict(comm.get("prime_150", {}) or {})
                    r_cur = dict(p150_cur.get("rewards", {}) or {})
                    r_cur["points_amount"] = float(v)
                    p150_cur["rewards"] = r_cur
                    # Ensure keys exist to avoid later synthesis dropping structure
                    if "direct" not in p150_cur:
                        p150_cur["direct"] = p150_cur.get("direct", {"sponsor": 0, "self": 0})
                    if "coupons" not in p150_cur:
                        p150_cur["coupons"] = p150_cur.get("coupons", {})
                    comm["prime_150"] = p150_cur
                    master["commissions"] = comm
            except Exception:
                return Response({"detail": "commissions.prime_150.rewards.points_amount must be a number"}, status=400)

            # commissions.prime_750 (base_package, multiplier)
            p750_in = cm.get("prime_750")
            if isinstance(p750_in, dict):
                # base_package must be 'prime_150' (only allowed)
                bp = str(p750_in.get("base_package", "prime_150") or "prime_150").strip().lower()
                if bp not in ("prime_150", ""):
                    return Response({"detail": "commissions.prime_750.base_package must be 'prime_150'"}, status=400)

                mv = None
                if "multiplier" in p750_in:
                    try:
                        mv = int(p750_in.get("multiplier"))
                        if mv <= 0:
                            return Response({"detail": "commissions.prime_750.multiplier must be a positive integer"}, status=400)
                    except Exception:
                        return Response({"detail": "commissions.prime_750.multiplier must be an integer"}, status=400)

                comm2 = dict(master.get("commissions", {}) or {})
                p750_cur = dict(comm2.get("prime_750", {}) or {})
                # enforce allowed base_package
                p750_cur["base_package"] = "prime_150"
                if mv is not None:
                    p750_cur["multiplier"] = int(mv)
                elif "multiplier" not in p750_cur:
                    # ensure there is a multiplier; default to 1 if absent
                    p750_cur["multiplier"] = int(p750_cur.get("multiplier", 1) or 1)
                comm2["prime_750"] = p750_cur
                master["commissions"] = comm2

        # Global updates (always permitted)
        tax = data.get("tax")
        if isinstance(tax, dict) and "percent" in tax:
            try:
                p = D(str(tax.get("percent")))
                if p < 0:
                    return Response({"detail": "tax.percent must be >= 0"}, status=400)
                t = dict(master.get("tax") or {})
                t["percent"] = float(p)
                master["tax"] = t
            except Exception:
                return Response({"detail": "tax.percent must be a number"}, status=400)

        if "tax_company_user_id" in data:
            try:
                tid = int(data.get("tax_company_user_id") or 0)
            except Exception:
                return Response({"detail": "tax_company_user_id must be integer id"}, status=400)
            if tid > 0:
                from accounts.models import CustomUser
                user = CustomUser.objects.filter(id=tid).first()
                if not user:
                    return Response({"detail": "tax_company_user_id not found"}, status=400)
                cfg.tax_company_user = user
            else:
                cfg.tax_company_user = None

        wd = data.get("withdrawal")
        if isinstance(wd, dict) and "sponsor_percent" in wd:
            try:
                sp = D(str(wd.get("sponsor_percent")))
                if sp < 0:
                    return Response({"detail": "withdrawal.sponsor_percent must be >= 0"}, status=400)
                w = dict(master.get("withdrawal") or {})
                w["sponsor_percent"] = float(sp)
                master["withdrawal"] = w
            except Exception:
                return Response({"detail": "withdrawal.sponsor_percent must be a number"}, status=400)

        # Withdrawals window (global)
        wwin = data.get("withdrawals_window")
        if isinstance(wwin, dict):
            # enabled
            if "enabled" in wwin:
                try:
                    cfg.withdrawals_enabled = bool(wwin.get("enabled"))
                except Exception:
                    cfg.withdrawals_enabled = True
            # weekday (0=Mon..6=Sun)
            if "weekday" in wwin:
                try:
                    wday = int(wwin.get("weekday"))
                except Exception:
                    return Response({"detail": "withdrawals_window.weekday must be an integer 0..6"}, status=400)
                if wday < 0 or wday > 6:
                    return Response({"detail": "withdrawals_window.weekday must be between 0 and 6"}, status=400)
                cfg.withdrawals_weekday = wday
            # start_time / end_time in HH:MM (24h)
            from datetime import time as _time
            def _parse_hhmm(val, field_name):
                if val is None:
                    return None
                s = str(val).strip()
                try:
                    parts = s.split(":")
                    if len(parts) != 2:
                        raise ValueError("bad")
                    hh = int(parts[0])
                    mm = int(parts[1])
                    if hh < 0 or hh > 23 or mm < 0 or mm > 59:
                        raise ValueError("range")
                    return _time(hh, mm)
                except Exception:
                    raise ValueError(f"{field_name} must be in HH:MM 24-hour format")

            if "start_time" in wwin:
                try:
                    cfg.withdrawals_start_time = _parse_hhmm(wwin.get("start_time"), "withdrawals_window.start_time")
                except ValueError as ve:
                    return Response({"detail": str(ve)}, status=400)
            if "end_time" in wwin:
                try:
                    cfg.withdrawals_end_time = _parse_hhmm(wwin.get("end_time"), "withdrawals_window.end_time")
                except ValueError as ve:
                    return Response({"detail": str(ve)}, status=400)

            # Persist withdrawals window fields immediately.
            # Note: These fields live on CommissionConfig model (not in master_commission_json).
            try:
                cfg.save(update_fields=[
                    "withdrawals_enabled",
                    "withdrawals_weekday",
                    "withdrawals_start_time",
                    "withdrawals_end_time",
                    "updated_at",
                ])
            except Exception:
                try:
                    cfg.save()
                except Exception:
                    pass

        up = data.get("upline")
        if isinstance(up, dict):
            u = dict(master.get("upline") or {})
            for k in ("l1", "l2", "l3", "l4", "l5"):
                if k in up:
                    try:
                        val = D(str(up.get(k)))
                        if val < 0:
                            return Response({"detail": f"upline.{k} must be >= 0"}, status=400)
                        u[k] = float(val)
                    except Exception:
                        return Response({"detail": f"upline.{k} must be a number"}, status=400)
            master["upline"] = u

        # Global geo percents
        geo = data.get("geo")
        if isinstance(geo, dict) and not pk:
            g_cur = dict(master.get("geo") or {})
            allowed = {"sub_franchise", "pincode", "pincode_coord", "district", "district_coord", "state", "state_coord", "employee", "royalty"}
            for k, v in geo.items():
                if k not in allowed:
                    continue
                try:
                    vv = D(str(v))
                    if vv < 0:
                        return Response({"detail": f"geo.{k} must be >= 0"}, status=400)
                    g_cur[k] = float(vv)
                except Exception:
                    return Response({"detail": f"geo.{k} must be a number"}, status=400)
            master["geo"] = g_cur

        # Product-specific updates
        if pk:
            direct_all = dict(master.get("direct_bonus", {}) or {})
            geo_mode_all = dict(master.get("geo_mode", {}) or {})
            geo_fixed_all = dict(master.get("geo_fixed", {}) or {})
            geo_pct_all = dict(master.get("geo_percent", {}) or {})

            # direct_bonus
            db = data.get("direct_bonus")
            if isinstance(db, dict):
                row = dict(direct_all.get(pk, {}) or {})
                for k in ("sponsor", "self"):
                    if k in db:
                        try:
                            v = D(str(db.get(k)))
                            if v < 0:
                                return Response({"detail": f"direct_bonus.{k} must be >= 0"}, status=400)
                            row[k] = float(v.quantize(D("0.01")))
                        except Exception:
                            return Response({"detail": f"direct_bonus.{k} must be a number"}, status=400)
                direct_all[pk] = row
                master["direct_bonus"] = direct_all

            # geo_mode
            if "geo_mode" in data:
                gm = str(data.get("geo_mode") or "").strip().lower()
                if gm not in ("fixed", "percent", ""):
                    return Response({"detail": "geo_mode must be 'fixed' or 'percent'"}, status=400)
                geo_mode_all[pk] = gm
                master["geo_mode"] = geo_mode_all

            # geo_fixed (rupees)
            gf = data.get("geo_fixed")
            if isinstance(gf, dict):
                allowed = {"sub_franchise", "pincode", "pincode_coord", "district", "district_coord", "state", "state_coord", "employee", "royalty"}
                row = dict(geo_fixed_all.get(pk, {}) or {})
                for k, v in gf.items():
                    if k not in allowed:
                        continue
                    try:
                        vv = D(str(v))
                        if vv < 0:
                            return Response({"detail": f"geo_fixed.{k} must be >= 0"}, status=400)
                        row[k] = float(vv.quantize(D("0.01")))
                    except Exception:
                        return Response({"detail": f"geo_fixed.{k} must be a number"}, status=400)
                geo_fixed_all[pk] = row
                master["geo_fixed"] = geo_fixed_all

            # product-specific geo percents (optional)
            geo_p = data.get("geo")
            if isinstance(geo_p, dict):
                allowed = {"sub_franchise", "pincode", "pincode_coord", "district", "district_coord", "state", "state_coord", "employee", "royalty"}
                row = dict(geo_pct_all.get(pk, {}) or {})
                for k, v in geo_p.items():
                    if k not in allowed:
                        continue
                    try:
                        vv = D(str(v))
                        if vv < 0:
                            return Response({"detail": f"geo.{k} must be >= 0"}, status=400)
                        row[k] = float(vv.quantize(D("0.01")))
                    except Exception:
                        return Response({"detail": f"geo.{k} must be a number"}, status=400)
                geo_pct_all[pk] = row
                master["geo_percent"] = geo_pct_all

            # Monthly 759 product-specific config (accept nested monthly_759 payload)
            if pk == "759":
                m759 = dict(master.get("monthly_759", {}) or {})
                mm = data.get("monthly_759")
                if isinstance(mm, dict):
                    from decimal import Decimal as D
                    # direct_first_month
                    if "direct_first_month" in mm:
                        try:
                            v = D(str(mm.get("direct_first_month")))
                            if v < 0:
                                return Response({"detail": "monthly_759.direct_first_month must be >= 0"}, status=400)
                            m759["direct_first_month"] = float(v)
                        except Exception:
                            return Response({"detail": "monthly_759.direct_first_month must be a number"}, status=400)
                    # direct_monthly
                    if "direct_monthly" in mm:
                        try:
                            v = D(str(mm.get("direct_monthly")))
                            if v < 0:
                                return Response({"detail": "monthly_759.direct_monthly must be >= 0"}, status=400)
                            m759["direct_monthly"] = float(v)
                        except Exception:
                            return Response({"detail": "monthly_759.direct_monthly must be a number"}, status=400)
                    # base_amount
                    if "base_amount" in mm:
                        try:
                            v = D(str(mm.get("base_amount")))
                            if v < 0:
                                return Response({"detail": "monthly_759.base_amount must be >= 0"}, status=400)
                            m759["base_amount"] = float(v)
                        except Exception:
                            return Response({"detail": "monthly_759.base_amount must be a number"}, status=400)
                    # levels_fixed (limit to 5 entries)
                    if "levels_fixed" in mm:
                        lst = mm.get("levels_fixed")
                        if lst is None:
                            m759["levels_fixed"] = []
                        elif isinstance(lst, (list, tuple)):
                            new = []
                            for x in lst[:5]:
                                try:
                                    v = D(str(x))
                                    if v < 0:
                                        v = D("0")
                                    new.append(float(v))
                                except Exception:
                                    new.append(0.0)
                            m759["levels_fixed"] = new
                        else:
                            return Response({"detail": "monthly_759.levels_fixed must be an array"}, status=400)
                    # agency_enabled (boolean)
                    if "agency_enabled" in mm:
                        m759["agency_enabled"] = bool(mm.get("agency_enabled"))
                    # matrix_open_mode (enum)
                    if "matrix_open_mode" in mm:
                        mv = str(mm.get("matrix_open_mode") or "").strip().upper()
                        if mv not in ("FIRST_MONTH_ONLY", "EVERY_PURCHASE", "NEVER"):
                            return Response({"detail": "monthly_759.matrix_open_mode must be FIRST_MONTH_ONLY, EVERY_PURCHASE, or NEVER"}, status=400)
                        m759["matrix_open_mode"] = mv
                # Persist back only when product=759 to avoid UnboundLocalError for other products
                master["monthly_759"] = m759

            # product base amount (per-product)
            if "product_base_amount" in data:
                try:
                    v = D(str(data.get("product_base_amount")))
                    if v < 0:
                        return Response({"detail": "product_base_amount must be >= 0"}, status=400)
                    pro_all = dict(master.get("products") or {})
                    row = dict(pro_all.get(pk, {}) or {})
                    row["base_amount"] = float(v)
                    pro_all[pk] = row
                    master["products"] = pro_all
                except Exception:
                    return Response({"detail": "product_base_amount must be a number"}, status=400)

            # coupons activation count (only applicable for 150 product)
            if "coupon_activation_count" in data and pk == "150":
                try:
                    iv = int(data.get("coupon_activation_count"))
                    if iv < 0:
                        return Response({"detail": "coupon_activation_count must be >= 0"}, status=400)
                    comm = dict(master.get("commissions", {}) or {})
                    p150 = dict(comm.get("prime_150", {}) or {})
                    coupons = dict(p150.get("coupons", {}) or {})
                    coupons["activation_count"] = iv
                    p150["coupons"] = coupons
                    if "direct" not in p150:
                        p150["direct"] = {"sponsor": 0, "self": 0}
                    if "rewards" not in p150:
                        p150["rewards"] = {"points_amount": 0}
                    comm["prime_150"] = p150
                    master["commissions"] = comm
                except Exception:
                    return Response({"detail": "coupon_activation_count must be an integer"}, status=400)

            # activation open count (only applicable for 750 product)
            if "activation_open_count" in data and pk == "750":
                try:
                    iv = int(data.get("activation_open_count"))
                    if iv < 0:
                        return Response({"detail": "activation_open_count must be >= 0"}, status=400)
                    pro_all = dict(master.get("products") or {})
                    row = dict(pro_all.get(pk, {}) or {})
                    row["activation_open_count"] = int(iv)
                    pro_all[pk] = row
                    master["products"] = pro_all
                    # Optional: mirror under top-level map for backward compatibility
                    aoc_map = dict(master.get("activation_open_count", {}) or {})
                    aoc_map["750"] = int(iv)
                    master["activation_open_count"] = aoc_map
                except Exception:
                    return Response({"detail": "activation_open_count must be an integer"}, status=400)

            # matrix open controls (applicable for 150 and 750 products)
            # Accept both product_matrix_open_* (preferred) and matrix_open_* (legacy UI) keys.
            # Also normalize common alias/typo values to keep UI lenient (e.g., PURCHASE_EVERTIME -> EVERY_PURCHASE).
            if pk in ("150", "750"):
                mode_key = None
                if "product_matrix_open_mode" in data:
                    mode_key = "product_matrix_open_mode"
                elif "matrix_open_mode" in data:
                    mode_key = "matrix_open_mode"
                if mode_key:
                    raw = str(data.get(mode_key) or "")
                    gm_in = raw.strip().upper().replace("-", "_").replace(" ", "_")
                    aliases = {
                        "PURCHASE_EVERYTIME": "EVERY_PURCHASE",
                        "PURCHASE_EVERTIME": "EVERY_PURCHASE",
                        "EVERYTIME_PURCHASE": "EVERY_PURCHASE",
                        "EVERYTIME": "EVERY_PURCHASE",
                        "EACH_PURCHASE": "EVERY_PURCHASE",
                        "FIRST_TIME": "FIRST_TIME_ONLY",
                        "FIRSTTIME_ONLY": "FIRST_TIME_ONLY",
                        "FIRST": "FIRST_TIME_ONLY",
                    }
                    gm = aliases.get(gm_in, gm_in)
                    allowed = ("FIRST_TIME_ONLY", "EVERY_PURCHASE", "NEVER")
                    if gm not in allowed:
                        return Response({"detail": f"{mode_key} must be FIRST_TIME_ONLY, EVERY_PURCHASE, or NEVER"}, status=400)
                    pro_all = dict(master.get("products") or {})
                    row = dict(pro_all.get(pk, {}) or {})
                    row["matrix_open_mode"] = gm
                    pro_all[pk] = row
                    master["products"] = pro_all

                count_key = None
                if "product_matrix_open_count" in data:
                    count_key = "product_matrix_open_count"
                elif "matrix_open_count" in data:
                    count_key = "matrix_open_count"
                if count_key:
                    try:
                        iv = int(data.get(count_key))
                    except Exception:
                        return Response({"detail": f"{count_key} must be an integer"}, status=400)
                    if iv < 0:
                        return Response({"detail": f"{count_key} must be >= 0"}, status=400)
                    pro_all = dict(master.get("products") or {})
                    row = dict(pro_all.get(pk, {}) or {})
                    row["matrix_open_count"] = int(iv)
                    pro_all[pk] = row
                    master["products"] = pro_all

        # keep commissions in sync with master keys
        try:
            from business.services.commission_policy import CommissionPolicy
            master["commissions"] = CommissionPolicy._synth_from_master(master)
        except Exception:
            pass

        cfg.master_commission_json = master
        try:
            if "tax_company_user_id" in data:
                cfg.save(update_fields=["master_commission_json", "tax_company_user", "updated_at"])
            else:
                cfg.save(update_fields=["master_commission_json", "updated_at"])
        except Exception:
            cfg.save()

        return self.get(request)

class AdminRewardPointsConfig(APIView):
    """
    GET: Return admin-configurable Reward Points schedule used by /business/rewards/points/
         {
           "tiers": [ { "count": 1, "points": 1000 }, ..., { "count": 5, "points": 110000 } ],
           "after": { "base_count": 5, "per_coupon": 20000 },
           "updated_at": "..."
         }
    PATCH: Update any subset of the above keys with validation.
           - tiers: array of {count:int>=1, points:int>=0}, sorted and unique by count
           - after.base_count: int >= max(tiers.count)
           - after.per_coupon: int >= 0
    """
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("commissions")]

    def _default_config(self):
        return {
            "tiers": [
                {"count": 1, "points": 1000},
                {"count": 2, "points": 10000},
                {"count": 3, "points": 30000},
                {"count": 4, "points": 60000},
                {"count": 5, "points": 110000},
            ],
            "after": {"base_count": 5, "per_coupon": 20000},
        }

    def _serialize(self, cfg):
        conf = dict(getattr(cfg, "reward_points_config_json", {}) or {})
        # Fill defaults if missing/invalid
        try:
            tiers = conf.get("tiers") or []
            after = conf.get("after") or {}
            if not isinstance(tiers, list) or not tiers:
                raise ValueError("tiers missing")
            norm_tiers = []
            seen = set()
            for t in tiers:
                c = int(t.get("count"))
                p = int(t.get("points"))
                if c < 1 or p < 0:
                    raise ValueError("invalid tier")
                if c in seen:
                    continue
                seen.add(c)
                norm_tiers.append({"count": c, "points": p})
            norm_tiers.sort(key=lambda x: x["count"])
            base_count = int(after.get("base_count", norm_tiers[-1]["count"]))
            per_coupon = int(after.get("per_coupon", 0))
            if base_count < norm_tiers[-1]["count"] or per_coupon < 0:
                raise ValueError("invalid after")
            conf = {"tiers": norm_tiers, "after": {"base_count": base_count, "per_coupon": per_coupon}}
        except Exception:
            conf = self._default_config()
        conf["updated_at"] = getattr(cfg, "updated_at", None)
        return conf

    def get(self, request):
        cfg = CommissionConfig.get_solo()
        return Response(self._serialize(cfg), status=200)

    def patch(self, request):
        if not getattr(request.user, "is_superuser", False):
            return Response({"detail": "Only superuser can modify commission config."}, status=403)
        cfg = CommissionConfig.get_solo()
        existing = self._serialize(cfg)
        data = request.data or {}

        # Start from existing; override selectively
        tiers_in = data.get("tiers", None)
        after_in = data.get("after", None)

        new_conf = {
            "tiers": existing.get("tiers") or self._default_config()["tiers"],
            "after": existing.get("after") or self._default_config()["after"],
        }

        # Validate tiers if provided
        if tiers_in is not None:
            if not isinstance(tiers_in, list) or not tiers_in:
                return Response({"detail": "tiers must be a non-empty array"}, status=400)
            try:
                norm = []
                seen = set()
                for t in tiers_in:
                    c = int(t.get("count"))
                    p = int(t.get("points"))
                    if c < 1 or p < 0:
                        return Response({"detail": "each tier must have count>=1 and points>=0"}, status=400)
                    if c in seen:
                        return Response({"detail": f"duplicate count in tiers: {c}"}, status=400)
                    seen.add(c)
                    norm.append({"count": c, "points": p})
                norm.sort(key=lambda x: x["count"])
                new_conf["tiers"] = norm
            except Exception:
                return Response({"detail": "invalid tiers payload"}, status=400)

        # Validate after if provided
        if after_in is not None:
            try:
                base_count = int(after_in.get("base_count", new_conf["after"]["base_count"]))
                per_coupon = int(after_in.get("per_coupon", new_conf["after"]["per_coupon"]))
                max_tier = max(t["count"] for t in (new_conf["tiers"] or [{"count": 1}]))
                if base_count < max_tier:
                    return Response({"detail": f"after.base_count must be >= max tier count ({max_tier})"}, status=400)
                if per_coupon < 0:
                    return Response({"detail": "after.per_coupon must be >= 0"}, status=400)
                new_conf["after"] = {"base_count": base_count, "per_coupon": per_coupon}
            except Exception:
                return Response({"detail": "invalid after payload"}, status=400)

        cfg.reward_points_config_json = new_conf
        try:
            cfg.save(update_fields=["reward_points_config_json", "updated_at"])
        except Exception:
            cfg.save()
        return Response(self._serialize(cfg), status=200)


class AdminWalletDebugView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        from decimal import Decimal
        from django.db.models import Q
        from accounts.models import Wallet, WalletTransaction, CustomUser

        username = (request.query_params.get("username") or "").strip()
        if not username:
            return Response({"detail": "Username is required"}, status=400)
            
        user = CustomUser.objects.filter(Q(username=username) | Q(prefixed_id=username) | Q(phone=username)).first()
        if not user:
            return Response({"detail": "User not found"}, status=400)
            
        try:
            w = user.wallet
        except Exception:
            w = Wallet.objects.create(user=user)
            
        txs = WalletTransaction.objects.filter(user=user).order_by('created_at', 'id')
        
        calc_main = Decimal("0.00")
        calc_self = Decimal("0.00")
        
        upload_sources = ["WALLET_UPLOAD", "UPLOAD_TO_WALLET", "PACKAGE_UPLOAD", "PACKAGE_BUY_UPLOAD"]
        withdrawal_types = ["WITHDRAWABLE_CREDIT", "WITHDRAWAL_DEBIT", "WITHDRAWAL_WALLET_TRANSFER_OUT"]
        shopping_types = ["SHOPPING_WALLET_CREDIT", "SHOPPING_WALLET_DEBIT", "SHOPPING_WALLET_TRANSFER_OUT"]
        coupon_redeem_types = [
            "AUTO_ECOUPON_ISSUED", "ECOUPON_WALLET_DEBIT", "COUPON_PURCHASE_CREDIT",
            "REDEEM_ECOUPON_CREDIT", "PRODUCT_WALLET_CREDIT", "PRODUCT_PURCHASE_DEBIT"
        ]
        
        for tx in txs:
            amt = tx.amount
            t = tx.type
            m = tx.meta or {}
            if t == 'SELF_ACCOUNT_CREDIT':
                calc_self += amt
            elif t == 'SELF_ACCOUNT_DEBIT':
                calc_self -= Decimal("250.00")
                if calc_self < Decimal("0.00"):
                    calc_self = Decimal("0.00")
            elif t == 'AUTO_PURCHASE_DEBIT' and tx.source_type == 'SELF_250_PACK':
                pass
            elif t in ['INTERNAL_WALLET_CREDIT', 'INTERNAL_WALLET_DEBIT']:
                is_add_money = (
                    tx.source_type in upload_sources
                    or m.get("wallet") == "ADD_MONEY"
                    or m.get("destination_wallet") == "ADD_MONEY_POCKET"
                    or m.get("legacy_wallet_type") == "ADD_MONEY_POCKET"
                    or m.get("wallet_source") == "package_upload"
                )
                if not is_add_money:
                    calc_main += amt
            elif t in withdrawal_types or t in shopping_types or t in coupon_redeem_types:
                pass
            else:
                calc_main += amt

        if user.id == 32 or user.username == "9999999999":
            calc_self = Decimal("0.00")

        if calc_main < Decimal("0.00"):
            calc_main = Decimal("0.00")
            
        cached_main = Decimal(str(w.main_balance or '0.00')).quantize(Decimal('0.01'))
        cached_self = Decimal(str(w.self_account_balance or '0.00')).quantize(Decimal('0.01'))
        cached_total = Decimal(str(w.balance or '0.00')).quantize(Decimal('0.01'))
        
        diff = cached_total - (calc_main + calc_self)
        
        recent_txs = txs.order_by('-created_at', '-id')[:50]
        tx_list = []
        for tx in recent_txs:
            tx_list.append({
                "id": tx.id,
                "amount": str(tx.amount),
                "balance_after": str(tx.balance_after),
                "type": tx.type,
                "created_at": tx.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "meta": tx.meta
            })
            
        return Response({
            "user": {
                "id": user.id,
                "full_name": user.full_name,
                "username": user.username,
                "phone": user.phone
            },
            "cached": {
                "main": str(cached_main),
                "self": str(cached_self),
                "total": str(cached_total)
            },
            "calculated": {
                "main": str(calc_main),
                "self": str(calc_self),
                "total": str(calc_main + calc_self)
            },
            "status": "OK" if diff == Decimal('0.00') else "MISMATCH",
            "diff": str(diff),
            "transactions": tx_list
        })

    def post(self, request):
        from decimal import Decimal
        from django.db import transaction as db_transaction
        from django.db.models import Q
        from accounts.models import WalletTransaction, CustomUser

        action = request.data.get("action")
        username = (request.data.get("username") or "").strip()
        if not username:
            return Response({"detail": "Username is required"}, status=400)
            
        user = CustomUser.objects.filter(Q(username=username) | Q(prefixed_id=username) | Q(phone=username)).first()
        if not user:
            return Response({"detail": "User not found"}, status=400)
            
        w = user.wallet
        
        if action == "sync":
            with db_transaction.atomic():
                txs = WalletTransaction.objects.filter(user=user).order_by('created_at', 'id')
                calc_main = Decimal("0.00")
                calc_self = Decimal("0.00")
                
                upload_sources = ["WALLET_UPLOAD", "UPLOAD_TO_WALLET", "PACKAGE_UPLOAD", "PACKAGE_BUY_UPLOAD"]
                withdrawal_types = ["WITHDRAWABLE_CREDIT", "WITHDRAWAL_DEBIT", "WITHDRAWAL_WALLET_TRANSFER_OUT"]
                shopping_types = ["SHOPPING_WALLET_CREDIT", "SHOPPING_WALLET_DEBIT", "SHOPPING_WALLET_TRANSFER_OUT"]
                coupon_redeem_types = [
                    "AUTO_ECOUPON_ISSUED", "ECOUPON_WALLET_DEBIT", "COUPON_PURCHASE_CREDIT",
                    "REDEEM_ECOUPON_CREDIT", "PRODUCT_WALLET_CREDIT", "PRODUCT_PURCHASE_DEBIT"
                ]
                
                for tx in txs:
                    amt = tx.amount
                    t = tx.type
                    m = tx.meta or {}
                    if t == 'SELF_ACCOUNT_CREDIT':
                        calc_self += amt
                    elif t == 'SELF_ACCOUNT_DEBIT':
                        calc_self -= Decimal("250.00")
                        if calc_self < Decimal("0.00"):
                            calc_self = Decimal("0.00")
                    elif t == 'AUTO_PURCHASE_DEBIT' and tx.source_type == 'SELF_250_PACK':
                        pass
                    elif t in ['INTERNAL_WALLET_CREDIT', 'INTERNAL_WALLET_DEBIT']:
                        is_add_money = (
                            tx.source_type in upload_sources
                            or m.get("wallet") == "ADD_MONEY"
                            or m.get("destination_wallet") == "ADD_MONEY_POCKET"
                            or m.get("legacy_wallet_type") == "ADD_MONEY_POCKET"
                            or m.get("wallet_source") == "package_upload"
                        )
                        if not is_add_money:
                            calc_main += amt
                    elif t in withdrawal_types or t in shopping_types or t in coupon_redeem_types:
                        pass
                    else:
                        calc_main += amt

                if user.id == 32 or user.username == "9999999999":
                    calc_self = Decimal("0.00")

                if calc_main < Decimal("0.00"):
                    calc_main = Decimal("0.00")
                    
                w.main_balance = calc_main
                w.self_account_balance = calc_self
                w.save(update_fields=["main_balance", "self_account_balance", "updated_at"])
                
            return Response({"detail": "Wallet balance successfully synced with ledger."})
            
        elif action == "adjust":
            amount = request.data.get("amount")
            pocket = request.data.get("pocket")
            adj_type = request.data.get("type")
            reason = (request.data.get("reason") or "").strip()
            
            if not amount:
                return Response({"detail": "Amount is required"}, status=400)
            if pocket not in ("main", "self"):
                return Response({"detail": "Pocket must be main or self"}, status=400)
            if not adj_type:
                return Response({"detail": "Adjustment type is required"}, status=400)
            if not reason:
                return Response({"detail": "Reason/comment is required"}, status=400)
                
            try:
                amt_dec = Decimal(str(amount))
            except Exception:
                return Response({"detail": "Invalid amount format"}, status=400)
                
            with db_transaction.atomic():
                if pocket == "main":
                    w.main_balance = (w.main_balance or Decimal("0")) + amt_dec
                else:
                    w.self_account_balance = (w.self_account_balance or Decimal("0")) + amt_dec
                    
                w.save(update_fields=["main_balance", "self_account_balance", "updated_at"])
                
                meta = {
                    "reason": reason,
                    "ledger": "MAIN" if pocket == "main" else "SELF_ACCOUNT",
                    "admin_adjusted": True
                }
                WalletTransaction.objects.create(
                    user=user,
                    amount=amt_dec,
                    balance_after=w.balance,
                    type=adj_type,
                    meta=meta
                )
                
            return Response({"detail": "Manual adjustment applied successfully."})
            
        else:
            return Response({"detail": "Invalid action"}, status=400)


class AdminDailySalesReportView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        from decimal import Decimal
        from django.db.models import Sum, Count
        from django.utils import timezone
        from mlm_ranks.models import RankUpgrade
        from business.models import SubscriptionActivation

        from_date_str = request.query_params.get("from")
        to_date_str = request.query_params.get("to")
        
        if not from_date_str or not to_date_str:
            return Response({"detail": "Both from and to dates are required"}, status=400)
            
        try:
            from_date = timezone.datetime.strptime(from_date_str, "%Y-%m-%d").date()
            to_date = timezone.datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Dates must be in YYYY-MM-DD format"}, status=400)
            
        upgrades = RankUpgrade.objects.filter(
            payment_status='SUCCESS',
            upgraded_at__date__range=(from_date, to_date)
        ).values('upgraded_at__date').annotate(
            count=Count('id'),
            total=Sum('upgrade_amount')
        )
        
        packages = SubscriptionActivation.objects.filter(
            created_at__date__range=(from_date, to_date)
        ).values('created_at__date').annotate(
            count=Count('id'),
            total=Sum('amount')
        )
        
        daily_sales = {}
        curr = from_date
        while curr <= to_date:
            date_key = curr.strftime("%Y-%m-%d")
            daily_sales[date_key] = {
                "date": date_key,
                "packages_count": 0,
                "packages_amount": "0.00",
                "upgrades_count": 0,
                "upgrades_amount": "0.00",
                "total_amount": "0.00"
            }
            curr += timezone.timedelta(days=1)
            
        for up in upgrades:
            dt = up['upgraded_at__date']
            if dt:
                date_key = dt.strftime("%Y-%m-%d")
                if date_key in daily_sales:
                    daily_sales[date_key]["upgrades_count"] = up['count'] or 0
                    daily_sales[date_key]["upgrades_amount"] = str(up['total'] or Decimal("0.00"))
                    
        for pkg in packages:
            dt = pkg['created_at__date']
            if dt:
                date_key = dt.strftime("%Y-%m-%d")
                if date_key in daily_sales:
                    daily_sales[date_key]["packages_count"] = pkg['count'] or 0
                    daily_sales[date_key]["packages_amount"] = str(pkg['total'] or Decimal("0.00"))
                    
        grand_total = Decimal("0.00")
        total_pkg_count = 0
        total_pkg_amt = Decimal("0.00")
        total_upg_count = 0
        total_upg_amt = Decimal("0.00")
        
        for date_key, data in daily_sales.items():
            pkg_amt = Decimal(data["packages_amount"])
            upg_amt = Decimal(data["upgrades_amount"])
            tot = pkg_amt + upg_amt
            data["total_amount"] = f"{tot:.2f}"
            grand_total += tot
            total_pkg_count += data["packages_count"]
            total_pkg_amt += pkg_amt
            total_upg_count += data["upgrades_count"]
            total_upg_amt += upg_amt
            
        results = sorted(daily_sales.values(), key=lambda x: x["date"], reverse=True)
        
        return Response({
            "results": results,
            "summary": {
                "total_revenue": f"{grand_total:.2f}",
                "packages_count": total_pkg_count,
                "packages_amount": f"{total_pkg_amt:.2f}",
                "upgrades_count": total_upg_count,
                "upgrades_amount": f"{total_upg_amt:.2f}"
            }
        })


class AdminUserLedgerStatementView(APIView):
    permission_classes = [IsAdminOrStaff, HasAdminModuleAccess("reports_finance")]

    def get(self, request):
        from decimal import Decimal
        from django.db.models import Q
        from accounts.models import Wallet, WalletTransaction, LedgerEntry, CustomUser, WalletAccount

        username = (request.query_params.get("username") or "").strip()
        if not username:
            return Response({"detail": "Username is required"}, status=400)
            
        user = CustomUser.objects.filter(Q(username=username) | Q(prefixed_id=username) | Q(phone=username)).first()
        if not user:
            return Response({"detail": "User not found"}, status=400)
            
        # 1. Fetch current pocket balances
        accounts_qs = WalletAccount.objects.filter(user=user)
        accounts_data = {}
        for acc in accounts_qs:
            accounts_data[acc.wallet_type] = {
                "id": acc.id,
                "current_balance": str(acc.current_balance),
                "available_balance": str(acc.available_balance)
            }
            
        # Ensure all types exist in accounts_data to prevent frontend errors
        from accounts.finance_constants import WalletTypes
        all_types = [WalletTypes.MAIN, WalletTypes.SELF_PACKAGE_POCKET, WalletTypes.COUPON_POCKET, WalletTypes.ADD_MONEY_POCKET, WalletTypes.WITHDRAWAL_WALLET]
        for t in all_types:
            if t not in accounts_data:
                accounts_data[t] = {
                    "id": None,
                    "current_balance": "0.00",
                    "available_balance": "0.00"
                }

        # 2. Fetch all WalletTransactions in chronological order
        txs = WalletTransaction.objects.filter(user=user).order_by('id')
        
        # 3. Trace running balances
        statement_txs = []
        running_main = Decimal("40.00") # Start with registration welcome bonus
        running_self = Decimal("0.00")
        
        # Virtual Welcome Bonus entry
        statement_txs.append({
            "id": "V-START",
            "created_at": user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "",
            "type": "WELCOME_BONUS",
            "amount": "40.00",
            "main_balance": "40.00",
            "self_balance": "0.00",
            "from_user": "-",
            "level": "-",
            "source": "WELCOME",
            "remarks": "Initial welcome/registration bonus"
        })
        
        for tx in txs:
            amt = tx.amount
            tt = tx.type
            meta = tx.meta or {}
            
            from_user = meta.get("from_user") or meta.get("from_user_id") or "-"
            level = meta.get("level_index")
            level_str = f"L{level}" if level is not None else "-"
            
            source = meta.get("source") or meta.get("trigger") or meta.get("reason") or "-"
            
            is_self = False
            if tt == 'SELF_ACCOUNT_CREDIT':
                is_self = True
            elif tt == 'SELF_ACCOUNT_DEBIT':
                is_self = True
            elif tt == 'INTERNAL_WALLET_CREDIT' and meta.get('destination_wallet') == 'SELF_PACKAGE_POCKET':
                is_self = True
            elif tt == 'INTERNAL_WALLET_DEBIT' and meta.get('wallet_source') == 'internal':
                is_self = True
                
            main_impact = "-"
            self_impact = "-"
            remarks = ""
            
            if is_self:
                if amt > 0:
                    running_self += amt
                else:
                    running_self = max(Decimal("0.00"), running_self - abs(amt))
                self_impact = f"{running_self:.2f}"
                remarks = f"Self Account update from {source}"
            else:
                # Ignore Coupon/Withdrawal specific pockets post-transition
                if tt in ['COUPON_WALLET_CREDIT', 'VOUCHER_CREATE_DEBIT', 'VOUCHER_REDEEM_CREDIT', 'PACKAGE_COUPON_WALLET_DEBIT']:
                    remarks = f"Coupon pocket specific update"
                elif tt in ['WITHDRAWAL_WALLET_CREDIT', 'WITHDRAWAL_DEBIT'] and meta.get('target_wallet') == 'withdrawal':
                    remarks = f"Withdrawal pocket specific update"
                elif tt == 'INTERNAL_WALLET_CREDIT' and meta.get('destination_wallet') == 'ADD_MONEY_POCKET':
                    remarks = f"Add money pocket specific update"
                elif tt == 'INTERNAL_WALLET_DEBIT' and meta.get('wallet_source') == 'package_upload':
                    remarks = f"Add money pocket specific update"
                else:
                    if amt > 0:
                        running_main += amt
                    else:
                        if tt == 'ADJUSTMENT_DEBIT' and meta.get('target_wallet') != 'internal':
                            pass
                        else:
                            running_main = max(Decimal("0.00"), running_main - abs(amt))
                    main_impact = f"{running_main:.2f}"
                    remarks = f"Main Wallet update from {source}"
                    
            statement_txs.append({
                "id": tx.id,
                "created_at": tx.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "type": tt,
                "amount": str(amt),
                "main_balance": main_impact,
                "self_balance": self_impact,
                "from_user": str(from_user),
                "level": level_str,
                "source": str(source),
                "remarks": remarks
            })

        # Calculate adjustments needed to match active database balances
        db_main = Decimal(str(accounts_data.get(WalletTypes.MAIN, {}).get("current_balance", "0.00")))
        db_self = Decimal(str(accounts_data.get(WalletTypes.SELF_PACKAGE_POCKET, {}).get("current_balance", "0.00")))
        
        diff_main = db_main - running_main
        diff_self = db_self - running_self
        
        if diff_main != Decimal("0.00") or diff_self != Decimal("0.00"):
            running_main += diff_main
            running_self += diff_self
            
            statement_txs.append({
                "id": "V-ADJ",
                "created_at": "2026-07-31 23:59:59",  # Transition alignment date
                "type": "MIGRATION_POCKET_SPLIT",
                "amount": f"Main: {diff_main:+.2f}, Self: {diff_self:+.2f}",
                "main_balance": f"{running_main:.2f}",
                "self_balance": f"{running_self:.2f}",
                "from_user": "-",
                "level": "-",
                "source": "MIGRATION",
                "remarks": "Split correction/starting balance migration alignment"
            })

        # 4. Fetch Ledger Entries
        ledgers = LedgerEntry.objects.filter(wallet_account__user=user).order_by('id')
        statement_ledgers = []
        for le in ledgers:
            statement_ledgers.append({
                "id": le.id,
                "created_at": le.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "pocket": le.wallet_account.wallet_type,
                "direction": le.direction,
                "amount": str(le.amount),
                "balance_before": str(le.balance_before),
                "balance_after": str(le.balance_after),
                "entry_ref": le.entry_ref,
                "remarks": le.remarks or ""
            })
            
        return Response({
            "user": {
                "id": user.id,
                "full_name": user.full_name,
                "username": user.username,
                "phone": user.phone,
                "is_active": user.is_active,
                "date_joined": user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else ""
            },
            "pockets": accounts_data,
            "transactions": statement_txs,
            "ledgers": statement_ledgers
        })

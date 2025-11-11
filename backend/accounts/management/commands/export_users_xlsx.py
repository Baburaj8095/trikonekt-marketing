import os
from pathlib import Path
from typing import Dict, Tuple

from django.core.management.base import BaseCommand
from django.db.models import Q

from accounts.models import CustomUser, AgencyRegionAssignment

# Excel writer
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover
    openpyxl = None


HEADERS = ["username", "password", "sponsor_id", "pincode", "district", "state"]


SHEETS: Dict[str, Q] = {
    # Names kept short (Excel sheet title limit is 31 chars)
    "Company_TRCM": Q(category="company_manager"),
    "State_Coords": Q(category="agency_state_coordinator"),
    "States": Q(category="agency_state"),
    "District_Coords": Q(category="agency_district_coordinator"),
    "Districts": Q(category="agency_district"),
    "Pincode_Coords": Q(category="agency_pincode_coordinator"),
    "Pincodes": Q(category="agency_pincode"),
    "Sub_Franchises": Q(category="agency_sub_franchise"),
    "Employees": Q(category="employee"),
    "Consumers": Q(category="consumer"),
}


class Command(BaseCommand):
    help = (
        "Export users grouped by category into an Excel workbook (one sheet per group).\n"
        "Sheets: Company_TRCM, State_Coords, States, District_Coords, Districts, Pincode_Coords, "
        "Pincodes, Sub_Franchises, Employees, Consumers.\n"
        "Columns per sheet: username, password (hashed), sponsor_id, pincode, district, state."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default="fixtures/export/users_by_category.xlsx",
            help="Output .xlsx path (relative to backend/ or absolute). Default: fixtures/export/users_by_category.xlsx",
        )
        parser.add_argument(
            "--include-inactive",
            action="store_true",
            help="Include inactive users (default: only active users).",
        )
        parser.add_argument(
            "--skip-empty",
            action="store_true",
            help="Do not include sheets that would be empty.",
        )

    def _resolve_output_path(self, out_path: str) -> Path:
        # Resolve relative to backend/ (manage.py directory)
        base_dir = Path(__file__).resolve().parents[3]  # .../backend/
        out_file = Path(out_path)
        if not out_file.is_absolute():
            out_file = base_dir / out_file
        out_file.parent.mkdir(parents=True, exist_ok=True)
        return out_file

    def _auto_size_columns(self, ws):
        # Set column widths based on max content length
        for col_idx, header in enumerate(HEADERS, start=1):
            max_len = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            # Add padding
            width = min(max_len + 2, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _derive_district_for_user(self, user: CustomUser) -> str:
        # Try to obtain a district text via region assignments if present
        dist = (
            AgencyRegionAssignment.objects.filter(user=user, level="district")
            .order_by("id")
            .values_list("district", flat=True)
            .first()
        )
        return (dist or "").strip()

    def handle(self, *args, **options):
        if openpyxl is None:
            raise RuntimeError(
                "openpyxl is not installed. Install dependencies (e.g., pip install -r requirements.txt) before running."
            )

        out_file = self._resolve_output_path(options["path"])
        include_inactive = bool(options.get("include_inactive"))
        skip_empty = bool(options.get("skip_empty"))

        # Base queryset
        base_qs = CustomUser.objects.select_related("state")
        if not include_inactive:
            base_qs = base_qs.filter(is_active=True)

        wb = openpyxl.Workbook()
        # Remove default sheet to fully control sheet order
        default_title = wb.active.title
        wb.remove(wb[default_title])

        total_written: Dict[str, int] = {}

        for sheet_name, q in SHEETS.items():
            qs = base_qs.filter(q).order_by("id").only(
                "username", "password", "sponsor_id", "pincode", "state", "is_active", "id"
            )

            if skip_empty and not qs.exists():
                continue

            ws = wb.create_sheet(title=sheet_name)
            # Header row
            ws.append(HEADERS)

            count = 0
            # Iterate in memory-efficient way
            for u in qs.iterator():
                state_name = getattr(u.state, "name", "") or ""
                district_name = self._derive_district_for_user(u)  # may be blank

                row = [
                    u.username or "",
                    u.password or "",  # Django stores hashed password; original is not retrievable
                    u.sponsor_id or "",
                    u.pincode or "",
                    district_name,
                    state_name,
                ]
                ws.append(row)
                count += 1

            # Auto-size columns for readability
            self._auto_size_columns(ws)
            total_written[sheet_name] = count

        # If workbook has no sheets (e.g., skip_empty + no data), ensure at least one empty sheet
        if not wb.sheetnames:
            ws = wb.create_sheet(title="Empty")
            ws.append(HEADERS)
            self._auto_size_columns(ws)

        wb.save(out_file)

        # Summary
        parts = [f"{name}={cnt}" for name, cnt in total_written.items()]
        summary = ", ".join(parts) if parts else "no data (created Empty sheet)"
        self.stdout.write(self.style.SUCCESS(f"Exported Excel to {out_file} ({summary})"))
